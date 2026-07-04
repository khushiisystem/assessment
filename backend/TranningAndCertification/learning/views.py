import csv
import logging
from operator import sub
from rest_framework.filters import SearchFilter

from django.conf import settings
from django.core.mail import send_mail
from django.db import transaction
from django.db.models import Count, Prefetch, OuterRef, Subquery, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import generics, permissions, serializers, status, viewsets
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import ValidationError
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from rest_framework.views import APIView
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from rest_framework.decorators import action
from .serializers import CandidateTechnologyProgressSerializer
import base64
from django.core.files.base import ContentFile
from openpyxl import load_workbook


from core.models import User
from core.views import AdminPermission, CandidatePermission

from .models import Assignment, Completion, Question, Technology, UserTechnologyProgress
from .serializers import (
    TechnologySerializer,
    # ModuleSerializer,
    QuestionSerializer,
    AssignmentSerializer,
    CompletionSerializer,
    UserTechnologyProgressSerializer,
    QuestionImportSerializer,
)

logger = logging.getLogger(__name__)


def _is_empty_value(value):
    return value is None or (isinstance(value, str) and not value.strip())


def _load_rows_from_upload(uploaded_file):
    name = (uploaded_file.name or "").lower()
    if name.endswith(".csv"):
        uploaded_file.seek(0)
        decoded = uploaded_file.read().decode("utf-8-sig")
        return [dict(row) for row in csv.DictReader(decoded.splitlines())]

    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(col).strip() if col is not None else "" for col in rows[0]]
    records = []
    for row in rows[1:]:
        record = {}
        for idx, header in enumerate(headers):
            if header:
                record[header] = row[idx] if idx < len(row) and row[idx] is not None else ""
        records.append(record)
    return records

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100


class IsAdminOrOrgAdmin(permissions.BasePermission):
    """
    Allow only admins (staff, org_admin) to access the endpoint.
    """
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        return request.user.is_staff or getattr(request.user, "role", None) in ["org_admin", "super_admin", "manager"]


class TechnologyBulkDeleteAPIView(APIView):
    """
    Bulk delete technologies by IDs.
    Only admins can perform this action.
    
    POST /api/technologies/bulk-delete/
    Body: {"technology_ids": ["uuid1", "uuid2", ...]}
    """
    permission_classes = [IsAdminOrOrgAdmin]

    def post(self, request):
        technology_ids = request.data.get('technology_ids', [])

        if not technology_ids or not isinstance(technology_ids, list):
            logger.warning(
                "Invalid technology_ids provided for bulk delete",
                extra={
                    "user_id": request.user.id,
                    "technology_ids": technology_ids,
                }
            )
            return Response(
                {"error": "technology_ids must be a non-empty list."},
                status=status.HTTP_400_BAD_REQUEST
            )

        logger.info(
            f"Bulk delete initiated for technology_ids : {technology_ids}",
            extra={
                "user_id": request.user.id,
                "technology_count": len(technology_ids),
            }
            )
        deleted_ids = []
        errors = []

        try:
            with transaction.atomic():
                for tech_id in technology_ids:
                    try:
                        tech = Technology.objects.get(id=tech_id)
                        tech_name = tech.name
                        tech.delete()

                        deleted_ids.append({
                        "id": str(tech_id),
                        "name": tech_name
                        })
                        logger.info(
                            f"Technology deleted successfully : {str(tech_id)}",
                            extra={
                                "user_id": request.user.id,
                                "technology_id": str(tech_id),
                                "technology_name": tech_name,
                            }
                        )

                    except Technology.DoesNotExist:
                        logger.warning(
                            "Technology not found during bulk delete",
                            extra={
                                "user_id": request.user.id,
                                "technology_id": str(tech_id),
                            }
                        )
                        errors.append({"id": str(tech_id), "error": "Technology not found."})

                    except Exception as e:
                        logger.exception(
                            f"Unexpected error deleting technology : {e}",
                            extra={
                                "user_id": request.user.id,
                                "technology_id": str(tech_id),
                            }
                        )
                        errors.append({"id": str(tech_id), "error": str(e)})

        except Exception as e:
            logger.exception(
                f"Transaction failed during bulk delete due to : {e}",
                extra={"user_id": request.user.id}
            )
            return Response(
                {"status": "error", "message": "Transaction failed. No technologies were deleted."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        if deleted_ids:
            logger.info(
                "Bulk delete completed",
                extra={
                    "user_id": request.user.id,
                    "deleted_count": len(deleted_ids),
                    "error_count": len(errors),
                }
            )
            return Response({
                "status": "success",
                "message": f"Successfully deleted {len(deleted_ids)} technology(ies).",
                "deleted_count": len(deleted_ids),
                "deleted": deleted_ids,
                "errors": errors,
            }, status=status.HTTP_200_OK)

        logger.error(
            "Bulk delete failed — no technologies were deleted",
            extra={
                "user_id": request.user.id,
                "errors": errors,
            }
        )
        return Response(
                {"status": "error", "message": "No technologies were deleted.", "errors": errors},
                status=status.HTTP_400_BAD_REQUEST
            )


class IsAdminOrReadOnly(permissions.BasePermission):
    """
    Allow only authenticated users to read data,
    and only admins (staff, org_admin) to modify data.
    """
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        if request.method in permissions.SAFE_METHODS:
            return True
        return request.user.is_staff or getattr(request.user, "role", None) in ["org_admin", "super_admin", "manager"]


class TechnologyViewSet(viewsets.ModelViewSet):
    queryset = Technology.objects.all()
    serializer_class = TechnologySerializer
    permission_classes = [IsAdminOrReadOnly]
    pagination_class = StandardResultsSetPagination
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    lookup_field = 'id'
    filterset_fields = []
    filter_backends = [SearchFilter]
    search_fields = ['name', 'category']
    
    def get_queryset(self):
        user = self.request.user
        # Prefetch related data to avoid N+1 queries
        progress_prefetch = Prefetch(
            'usertechnologyprogress_set',
            queryset=UserTechnologyProgress.objects.select_related('user')
        )
        
        # "Enrolled" count = DISTINCT candidates assigned to the technology, not
        # raw Assignment rows. A single candidate has one Assignment row per
        # assigned question, so counting rows hugely inflates the number.
        # Org admins see only their organization's enrolments; super
        # admins/staff see all.
        is_admin_all = user.is_staff or user.is_superuser or user.role == 'super_admin'
        org_id = getattr(user, 'organization_id', None)
        if not is_admin_all and org_id:
            enrolled_annotation = Count(
                'assignment__user',
                distinct=True,
                filter=Q(assignment__user__organization_id=org_id),
            )
        else:
            enrolled_annotation = Count('assignment__user', distinct=True)

        qs = Technology.objects.prefetch_related(
            'questions',
            'assignment_set',
            progress_prefetch
        ).annotate(enrolled_users_count=enrolled_annotation).order_by("-created_at")

        if user.is_staff or user.is_superuser or user.role == 'super_admin':
            logger.info(
                "Admin/super_admin fetching all technologies",
                extra={"user_id": user.id}
            )
            return qs
        
        # --- Handle Individual Subscribers ---
        if getattr(user, 'is_individual', False):
            sub = getattr(user, 'subscription', None)
            if sub and sub.is_valid() and sub.plan.plan_type != 'free':
                # Paid individual: full access to all technologies
                return qs
            # Free / no subscription: show ALL technologies (frontend will lock beyond free limit)
            return qs

        # --- Handle Organization Users ---
        # org_admin AND manager are org staff — both see their org's courses
        # (managers fell through to the candidate branch and saw nothing).
        if user.role in ('org_admin', 'manager'):
            return qs
        #    return qs.filter(organization=user.organization)
  
        logger.info(
            "Candidate fetching assigned technologies for user_id : {user_id}",
            extra={"user_id": user.id}
        )
        return qs.filter(
            id__in=UserTechnologyProgress.objects.all_for_super_admin().filter(user=user).values_list('technology_id', flat=True)
        ).distinct()
    
    def perform_create(self, serializer):
        """Scope new courses to the org for org_admin and manager creators."""
        if self.request.user.role in ('org_admin', 'manager'):
            serializer.save(organization=self.request.user.organization)
        else:
            serializer.save()
    
class TechnologyCandidatesAPIView(APIView):
    permission_classes = [IsAdminOrOrgAdmin]

    def get(self, request, technology_id):
        technology = get_object_or_404(Technology, id=technology_id)

        logger.info(
            "Fetching candidates for technology",
            extra={"user_id": request.user.id, "technology_id": str(technology_id)}
        )

        user = request.user
        if user.role in ('org_admin', 'manager'):
            progress_qs = (
                UserTechnologyProgress.objects.all_for_super_admin()
                .filter(technology=technology, user__organization=user.organization)
                .select_related("user")
            )
            assignments = Assignment.objects.all_for_super_admin().filter(
                technology=technology,
                user__organization=user.organization
            )
        else:
            progress_qs = (
                UserTechnologyProgress.objects.all_for_super_admin()
                .filter(technology=technology)
                .select_related("user")
            )
            assignments = Assignment.objects.all_for_super_admin().filter(technology=technology)

        assignment_map = {
            a.user_id: {
                "assignment_id": a.id,
                "assigned_at": a.assigned_at,
                "due_at": a.due_at
            }
            for a in assignments
        }

        serializer = CandidateTechnologyProgressSerializer(progress_qs, many=True)
        data = serializer.data

        #  Inject assignment data
        for item, obj in zip(data, progress_qs):
            assignment = assignment_map.get(obj.user_id)

            if assignment:
                item["assignment_id"] = assignment["assignment_id"]
                item["assigned_at"] = assignment["assigned_at"]
                item["due_at"] = assignment["due_at"]
            else:
                item["assignment_id"] = None
                item["assigned_at"] = None
                item["due_at"] = None

        logger.info(
            "Candidates fetched successfully",
            extra={
                "user_id": request.user.id,
                "technology_id": str(technology_id),
                "candidate_count": len(data),
            }
        )
        return Response(data)
    
class AllCandidatesActivityView(APIView):
    permission_classes = [IsAdminOrOrgAdmin]

    def get(self, request):
        # Annotate progress with last completion time so we can sort by activity.
        logger.info(
            "Fetching all candidates activity",
            extra={"user_id": request.user.id}
        )

        last_completion_subquery = Subquery(
            Completion.objects
            .filter(
                user=OuterRef('user'),
                question__technology=OuterRef('technology')
            )
            .order_by('-completed_at')
            .values('completed_at')[:1]
        )

        progress_qs = (
            UserTechnologyProgress.objects
            .select_related("user", "technology", "user__organization")
            .annotate(last_active_at=last_completion_subquery)
            .order_by("-last_active_at")
        )

        # Scope by organization: super admins see all orgs (each labelled), org
        # admins see only their organization's activity.
        from organization.context import current_organization_id, current_user_is_super_admin
        is_super = current_user_is_super_admin.get()
        org_id = current_organization_id.get()
        if not is_super:
            progress_qs = progress_qs.filter(user__organization_id=org_id) if org_id else progress_qs.none()

        # Pagination
        paginator = StandardResultsSetPagination()
        paginated_qs = paginator.paginate_queryset(progress_qs, request)

        # Fetch assignments in ONE query for these users/technologies
        user_tech_pairs = [(p.user_id, p.technology_id) for p in paginated_qs]
        
        assignments = Assignment.objects.filter(
            user_id__in=[pair[0] for pair in user_tech_pairs],
            technology_id__in=[pair[1] for pair in user_tech_pairs],
            question__isnull=True  # technology-level assignments only
        ).values("user_id", "technology_id", "assigned_at", "due_at")

        # Build assignment lookup map
        assignment_map = {
            (a["user_id"], a["technology_id"]): a
            for a in assignments
        }

        results = []
        for p in paginated_qs:
            assignment = assignment_map.get((p.user_id, p.technology_id), {})
            results.append({
                "userId": str(p.user.id),
                "name": f"{p.user.first_name} {p.user.last_name}".strip(),
                "email": p.user.email,
                "courseId": str(p.technology.id),
                "courseName": p.technology.name,
                "category": p.technology.category,
                "organization": p.user.organization.name if p.user.organization_id else None,
                "progress": p.progress or 0,
                "completed": p.completed or 0,
                "total": p.total or 0,
                "assigned_at": assignment.get("assigned_at"),
                "due_at": assignment.get("due_at"),
                "last_active_at": p.last_active_at if hasattr(p, 'last_active_at') else None,
            })

        logger.info(
            "All candidates activity fetched",
            extra={"user_id": request.user.id, "result_count": len(results)}
        )
        return paginator.get_paginated_response(results)

class QuestionViewSet(viewsets.ModelViewSet):
    """
    Handles all Question operations directly under Technology:
    - GET /api/technologies/<technology_id>/questions/?difficulty=
    - POST /api/technologies/<technology_id>/questions/
    - PATCH /api/technologies/<technology_id>/questions/<id>/
    - DELETE /api/technologies/<technology_id>/questions/<id>/
    """
    serializer_class = QuestionSerializer
    permission_classes = [IsAdminOrReadOnly]
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        user = self.request.user
        technology_id = self.kwargs.get('technology_id')

        ADMIN_ROLES = {'super_admin', 'org_admin', 'manager'}

        if user.is_staff or user.role in ADMIN_ROLES:
            queryset = Question.objects.all_for_super_admin()
        else:
            queryset = Question.objects.all_for_super_admin()

            # Individual candidates
            if getattr(user, 'is_individual', False):
                sub = getattr(user, 'subscription', None)
                if sub and sub.is_valid() and sub.plan.plan_type != 'free':
                    # Paid: full access
                    pass
                else:
                    # Free: only assigned technologies
                    assigned_tech_ids = Assignment.objects.filter(
                        user=user, 
                        question__isnull=True
                    ).values_list('technology_id', flat=True)
                    queryset = queryset.filter(technology_id__in=assigned_tech_ids)
            else:
                # Org candidate: only assigned technologies
                assigned_tech_ids = UserTechnologyProgress.objects.all_for_super_admin().filter(
                    user=user
                ).values_list('technology_id', flat=True)
                queryset = queryset.filter(technology_id__in=assigned_tech_ids)

        if technology_id:
            queryset = queryset.filter(technology_id=technology_id)

        difficulty = self.request.query_params.get('difficulty')
        if difficulty:
            queryset = queryset.filter(difficulty__iexact=difficulty)

        return queryset.order_by('id')


    def perform_create(self, serializer):
        """Attach question to a specific technology."""
        technology_id = self.kwargs.get('technology_id')
        technology = get_object_or_404(Technology, id=technology_id)
        serializer.save(technology=technology, organization=self.request.user.organization)
        logger.info(
            "Question created",
            extra={
                "user_id": self.request.user.id,
                "technology_id": str(technology_id)} )

    def perform_update(self, serializer):
        """Ensure question remains under same technology during update."""
        technology_id = self.kwargs.get('technology_id')
        technology = get_object_or_404(Technology, id=technology_id)
        serializer.save(technology=technology,organization=self.request.user.organization)
        logger.info(
            "Question updated",
            extra={
                "user_id": self.request.user.id,
                "technology_id": str(technology_id)} )

    def destroy(self, request, *args, **kwargs):
        """Delete question under the correct technology."""
        technology_id = self.kwargs.get('technology_id')
        question = get_object_or_404(Question, id=kwargs['pk'], technology_id=technology_id)
        question.delete()
        logger.info(
            "Question deleted",
            extra={
                "user_id": request.user.id,
                "technology_id": str(technology_id),
                "question_id": str(kwargs['pk'])} )
        return Response(status=status.HTTP_204_NO_CONTENT)

    def get_serializer_context(self):
        """Pass technology instance to serializer for duplicate validation."""
        context = super().get_serializer_context()
        technology_id = self.kwargs.get('technology_id')

        if technology_id:
            technology = get_object_or_404(Technology, id=technology_id)
            context['technology'] = technology

        return context


from rest_framework.permissions import IsAdminUser

class AssignmentListCreateView(generics.ListCreateAPIView):
    serializer_class = AssignmentSerializer
    pagination_class = StandardResultsSetPagination
    # permission_classes = [IsAdminUser]

    def get_queryset(self):
        queryset = Assignment.objects.all()
        user_id = self.request.query_params.get('userId')
        technology_id = self.request.query_params.get('technologyId')
        # module_id = self.request.query_params.get('moduleId')

        if user_id:
            queryset = queryset.filter(user__id=user_id)
        if technology_id:
            queryset = queryset.filter(technology__id=technology_id)
        # if module_id:
        #     queryset = queryset.filter(module__id=module_id)

        # Apply organization-based filtering for org-scoped staff (org_admin,
        # manager). Both see assignments within their own organization.
        user = self.request.user
        if user.role in ('org_admin', 'manager'):
            queryset = queryset.filter(technology__organization=user.organization)
        return queryset.order_by('-assigned_at')

    def perform_create(self, serializer):
        user_id = self.request.data.get('userId')
        technology_id = self.request.data.get('technologyId')
        assigned_by_id = self.request.data.get('assignedBy')

        user = get_object_or_404(User, id=user_id)
        technology = get_object_or_404(Technology, id=technology_id)
        assigned_by = get_object_or_404(User, id=assigned_by_id)

        #  Check duplicate assignment
        if Assignment.objects.filter(user=user, technology=technology, question__isnull=True).exists():
            logger.warning(
                "Duplicate assignment attempt",
                extra={
                    "assigned_by_id": assigned_by_id,
                    "user_id": user_id,
                    "technology_id": technology_id,
                }     )
            raise serializers.ValidationError("This technology is already assigned to the user.")
    
        # Create assignment
        assignment = serializer.save(
            user=user,
            technology=technology,
            assigned_by=assigned_by,
            organization=self.request.user.organization,
            question=None  # technology-level assignment
        )

        #  Create or update progress record
        logger.info(
            "Assignment created",
            extra={
                "assignment_id": assignment.id,
                "user_id": user_id,
                "technology_id": technology_id,
                "assigned_by_id": assigned_by_id,
            }        )
        progress_obj, created = UserTechnologyProgress.objects.get_or_create(
            user=user,
            technology=technology,
            defaults={'progress': 0, 'completed': 0, 'total': technology.questions.count()}
        )

        if not created:
            progress_obj.total = technology.questions.count()
            progress_obj.save()
    
        #  SEND EMAIL AFTER ASSIGNMENT
            logger.info(
                "Progress record updated for existing candidate",
                extra={"user_id": user_id, "technology_id": technology_id}
            )
        else:
            logger.info(
                "New progress record created",
                extra={"user_id": user_id, "technology_id": technology_id} )

        subject = f"New Course Assigned: {technology.name}"
        context = {
             'first_name': user.first_name,
             'technology_name': technology.name,
             'total_questions': technology.questions.count(),
             'assigned_by': "SkilTechy",
            #  'login_url': settings.FRONTEND_URL,
            'login_url': f"{settings.SITE_URL.rstrip('/')}/login?redirect=/candidate/dashboard",
        }

        html_message = render_to_string('emails/technology_assigned.html', context)
        plain_message = strip_tags(html_message)

        try:
            send_mail(
                   subject,
                   plain_message,
                   settings.DEFAULT_FROM_EMAIL,
                   [user.email],
                   fail_silently=False,
                  html_message=html_message,
            )
            logger.info("Assignment email sent", extra={"user_id": user.id, "technology_id": technology.id})
        except Exception:  # noqa: BLE001
           logger.exception(
             "Assignment email failed",
              extra={"user_id": user.id, "technology_id": technology.id},
    )
        return assignment
    

class AssignmentDeleteView(generics.DestroyAPIView):
    serializer_class = AssignmentSerializer
    permission_classes = [IsAdminOrOrgAdmin]

    def get_queryset(self):
        return Assignment.objects.all()

    def destroy(self, request, *args, **kwargs):
        assignment = self.get_object()
        user = assignment.user
        tech = assignment.technology

        # delete the assignment
        logger.info(
            "Deleting assignment",
            extra={
                "user_id": request.user.id,
                "assignment_id": assignment.id,
                "candidate_id": user.id,
                "technology_id": tech.id,
            }  )

        self.perform_destroy(assignment)

        # remove progress so the candidate no longer sees the tech
        # UserTechnologyProgress.objects.filter(user=user, technology=tech).delete()

        logger.info(
            "Assignment and progress deleted successfully",
            extra={
                "candidate_id": user.id,
                "technology_id": tech.id,
            }
        )
        return Response(status=status.HTTP_204_NO_CONTENT)







class CompletionViewSet(viewsets.ModelViewSet):
    """
    Handles marking question completions
    - GET /api/completions/?technologyId=
    - POST /api/completions/
    """
    serializer_class = CompletionSerializer
    pagination_class = StandardResultsSetPagination
    # permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        user = self.request.user
        technology_id = self.request.query_params.get('technologyId')

        queryset = Completion.objects.filter(user=user).select_related('question__technology')

        if technology_id:
            queryset = queryset.filter(question__technology_id=technology_id)
        return queryset.order_by('-completed_at')

    def create(self, request, *args, **kwargs):
        user = request.user
        question_id = request.data.get('questionId')

        # Validate question
        question = get_object_or_404(Question, id=question_id)

        # Prevent duplicates
        completion, created = Completion.objects.get_or_create(
            user=user,
            question=question,
            defaults={"organization": user.organization},
        )

        if not created:
            logger.info(
                "Question already marked as completed",
                extra={"user_id": user.id, "question_id": question_id}
            )
            return Response({'message': 'Already marked as completed'}, status=status.HTTP_200_OK)

        # Update progress automatically
        self.update_progress(user, question.technology)
        logger.info(
            "Question marked as completed",
            extra={
                "user_id": user.id,
                "question_id": question_id,
                "technology_id": question.technology.id})

        serializer = self.get_serializer(completion)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update_progress(self, user, technology):
        """Auto update or create user progress record."""
        total_questions = Question.objects.all_for_super_admin().filter(technology=technology).count()
        completed_questions = Completion.objects.filter(
            user=user, question__technology=technology
        ).count()

        progress_percent = (completed_questions / total_questions * 100) if total_questions > 0 else 0

        progress_obj, _ = UserTechnologyProgress.objects.get_or_create(
            user=user,
            technology=technology,
            defaults={
                'progress': progress_percent,
                'completed': completed_questions,
                'total': total_questions
            }
        )

        progress_obj.completed = completed_questions
        progress_obj.total = total_questions
        progress_obj.progress = progress_percent
        progress_obj.save()

        logger.info(
            "User progress updated",
            extra={
                "user_id": user.id,
                "technology_id": technology.id,
                "progress_percent": progress_percent,
                "completed": completed_questions,
                "total": total_questions,
            } )

class UserTechnologyProgressViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Returns logged-in user's progress per technology
    GET /api/progress/
    """
    serializer_class = UserTechnologyProgressSerializer
    pagination_class = StandardResultsSetPagination
    # permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        logger.info(
            "Fetching technology progress for user_id=%s",
            self.request.user.id,
            extra={"user_id": self.request.user.id}
        )
        return UserTechnologyProgress.objects.all_for_super_admin().filter(user=self.request.user)

class ImportQuestionsAPIView(APIView):
    """
    Import Questions in Bulk using CSV or Excel.
    Columns Required:
        - question
        - answer
        - difficulty (Easy/Medium/Hard)
        - module_level (beginner/basic/intermediate/advanced)
        - reference_link (optional)
        - task_description (optional)
    """
    permission_classes = [IsAdminOrOrgAdmin]   # Only admins allowed
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, technology_id):
        logger.info("Question import started")
        serializer = QuestionImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        file = serializer.validated_data['file']
        tech = get_object_or_404(Technology, id=technology_id)

        # Read file
        try:
            logger.info(
                "Reading import file",
                extra={
                    "user_id": request.user.id,
                    "technology_id": str(technology_id),
                    "file_name": file.name,
                }  )
            rows = _load_rows_from_upload(file)
        except Exception as e:
            logger.exception(
                f"Failed to read import file due to error {e}",
                extra={"user_id": request.user.id, "technology_id": str(technology_id)}
            )
            raise ValidationError({"file": "Failed to process uploaded file."})

        # Normalize header names so Excel/CSV columns like "Question", "question", "Module Level"
        # or "module_level" all work correctly.
        def normalize_header(name):
            return str(name).strip().lower().replace(" ", "_").replace("-", "_")

        rows = [
            {normalize_header(k): v for k, v in row.items()}
            for row in rows
        ]
        original_columns = list(rows[0].keys()) if rows else []

        required_columns = ["question", "answer", "difficulty", "module_level"]
        missing = [col for col in required_columns if col not in original_columns]
        if missing:
            logger.warning(
                "Missing columns in import file",
                extra={
                    "user_id": request.user.id,
                    "technology_id": str(technology_id),
                    "missing_columns": missing,
                    "found_columns": original_columns,
                }  )
            raise ValidationError({
                "file": (
                    f"Missing column(s): {', '.join(missing)}. "
                    f"Found columns: {', '.join(original_columns)}"
                )
            })

        errors = []
        created_count = 0

        allowed_difficulties = {"easy", "medium", "hard"}
        allowed_module_levels = {"beginner", "basic", "intermediate", "advanced"}

        try:
            with transaction.atomic():
                for index, row in enumerate(rows):
                    raw_difficulty = str(row.get('difficulty', '')).strip()
                    raw_module_level = str(row.get('module_level', '')).strip()

                    normalized_difficulty = raw_difficulty.lower()
                    normalized_module_level = raw_module_level.lower().replace(" ", "_").replace("-", "_")

                    if normalized_difficulty not in allowed_difficulties:
                        logger.warning(
                            "Invalid difficulty value in import row",
                            extra={
                                "row_index": index + 1,
                                "value": raw_difficulty,
                                "technology_id": str(technology_id),
                            }
                        )
                        errors.append(
                            f"Row {index+1}: Invalid difficulty '{raw_difficulty}'. "
                            f"Allowed values: Easy, Medium, Hard."
                        )
                        continue

                    if normalized_module_level not in allowed_module_levels:
                        logger.warning(
                            "Invalid module_level value in import row",
                            extra={
                                "row_index": index + 1,
                                "value": raw_module_level,
                                "technology_id": str(technology_id),
                            }
                        )
                        errors.append(
                            f"Row {index+1}: Invalid module_level '{raw_module_level}'. "
                            f"Allowed values: beginner, basic, intermediate, advanced."
                        )
                        continue

                    Question.objects.create(
                        technology=tech,
                        organization=request.user.organization,
                        question=row['question'],
                        answer=row['answer'],
                        difficulty=normalized_difficulty.capitalize(),
                        module_level=normalized_module_level,
                        reference_link=row.get('reference_link'),
                        task_description=row.get('task_description')
                    )
                    created_count += 1
        except Exception:
            logger.exception(
                "Unexpected error during question import transaction",
                extra={"user_id": request.user.id, "technology_id": str(technology_id)}
            )
            raise

        logger.info(
            "Question import completed",
            extra={
                "user_id": request.user.id,
                "technology_id": str(technology_id),
                "created_count": created_count,
                "error_count": len(errors),
            }
        )

        return Response({
            "status": "success",
            "created": created_count,
            "errors": errors
        })


class DownloadQuestionTemplateAPIView(APIView):
    permission_classes = [IsAdminOrOrgAdmin]

    def get(self, request):
        # File Response
        logger.info(
            "Question template CSV download requested",
            extra={"user_id": request.user.id}
        )

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="questions_template.csv"'

        writer = csv.writer(response)
        
        # Header row
        writer.writerow([
            "question",
            "answer",
            "difficulty",
            "module_level",
            "reference_link",
            "task_description"
        ])

        # Sample row
        writer.writerow([
            "What is Python?",
            "Python is a programming language.",
            "Easy",
            "beginner",
            "https://example.com/python",
            "Write a simple Python script."
        ])

        writer.writerow([
            "Explain Django Models.",
            "Django models represent database tables.",
            "Medium",
            "basic",
            "",
            "Create a simple model."
        ])

        return response

class SubmitUserNotesAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, technology_id):
        user = request.user
        notes_url = request.data.get("user_notes")

        if not notes_url:
            logger.warning(
                "user_notes field missing in submit notes request",
                extra={"user_id": user.id, "technology_id": str(technology_id)}
            )
            return Response({"error": "user_notes field is required."}, status=400)

        progress = get_object_or_404(
            UserTechnologyProgress,
            user=user,
            technology_id=technology_id
        )

        # Only allow submit when progress is 100%
        if progress.progress < 100:
            logger.warning(
                "Notes submission blocked — course not complete",
                extra={
                    "user_id": user.id,
                    "technology_id": str(technology_id),
                    "current_progress": progress.progress,
                }
            )
            return Response(
                {"error": "Complete all questions before submitting notes."},
                status=400
            )

        progress.user_notes = notes_url
        progress.save()

        logger.info(
            "User notes submitted successfully",
            extra={"user_id": user.id, "technology_id": str(technology_id)}
        )
        return Response({"message": "Notes submitted successfully!"}, status=200)



class SendReminderEmailAPIView(APIView):
    """
    Send reminder email to a candidate for a specific incomplete learning assignment.
    Only admins can trigger this.
    """
    permission_classes = [AdminPermission]  # Only admins

    def post(self, request, candidate_id, assignment_id):
        # Get candidate
        candidate = get_object_or_404(User, id=candidate_id, role="candidate")

        # Get assignment
        assignment = get_object_or_404(
            Assignment.objects.select_related("technology"),
            id=assignment_id,
            user=candidate
        )

        # Get progress
        progress = UserTechnologyProgress.objects.filter(
            user=candidate,
            technology=assignment.technology
        ).first()

        if not progress:
            logger.warning(
                "No progress record found for assignment reminder",
                extra={
                    "admin_id": request.user.id,
                    "candidate_id": candidate_id,
                    "assignment_id": assignment_id,
                }  )
            return Response(
                {"error": "No progress record found for this assignment."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Determine if reminder is needed
        now = timezone.now()
        needs_reminder = False
        reason = ""

        if progress.progress < 100:
            if assignment.due_at:
                days_until_due = (assignment.due_at - now).days
                if days_until_due <= 7:
                    needs_reminder = True
                    reason = f"Due in {days_until_due} days" if days_until_due > 0 else "Overdue"
            else:
                if progress.progress == 0:
                    needs_reminder = True
                    reason = "Not started"
                elif progress.progress < 100:
                    needs_reminder = True
                    reason = "Incomplete"

        if not needs_reminder:
            logger.info(
                "Reminder not required for this assignment",
                extra={
                    "admin_id": request.user.id,
                    "candidate_id": candidate_id,
                    "assignment_id": assignment_id,
                    "progress": progress.progress,
                }
            )
            return Response(
                {"message": "This assignment does not require a reminder."},
                status=status.HTTP_200_OK
            )

        # Prepare email
        subject = f"Reminder: Complete Your {assignment.technology.name} Assignment"
        message = (
            f"Hello {candidate.first_name},\n\n"
            f"This is a reminder to complete your learning assignment:\n\n"
            f"Assignment ID: {assignment.id}\n"
            f"Technology: {assignment.technology.name}\n"
            f"Progress: {progress.progress:.1f}% ({progress.completed}/{progress.total} questions)\n"
            f"Reason: {reason}\n"
            f"Due Date: {assignment.due_at.strftime('%Y-%m-%d') if assignment.due_at else 'No due date'}\n\n"
            f"Login URL: {settings.SITE_URL}/candidate/my-learning\n\n"
            f"Please log in and complete this assignment.\n\n"
            f"Best Regards,\nYour Training Team"
        )

        # Send email
        try:
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [candidate.email],
                fail_silently=False
            )
            logger.info(
                "Reminder email sent successfully",
                extra={
                    "admin_id": request.user.id,
                    "candidate_id": candidate.id,
                    "assignment_id": assignment.id,
                    "technology_id": assignment.technology.id,
                    "reason": reason,
                }
            )
        except Exception as e:
            logger.exception(
                "Reminder email failed",
                extra={"candidate_id": candidate.id, "assignment_id": assignment.id}
            )
            return Response(
                {"error": f"Failed to send email: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        return Response(
            {
                "message": f"Reminder email sent successfully to {candidate.email}",
                "assignment_details": {
                    "assignment_id": assignment.id,
                    "technology_name": assignment.technology.name,
                    "progress": progress.progress,
                    "reason": reason
                }
            },
            status=status.HTTP_200_OK
        )

class AssignmentUpdateDueDateView(APIView):
    permission_classes = [AdminPermission]

    def patch(self, request, assignment_id):
        assignment = get_object_or_404(Assignment, pk=assignment_id)
        
        due_at_raw = request.data.get("due_at")
        if not due_at_raw:
            logger.warning(
                "due_at missing in update due date request",
                extra={"user_id": request.user.id, "assignment_id": assignment_id}
            )
            return Response({"detail": "due_at is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            from django.utils.dateparse import parse_datetime
            from django.utils import timezone as tz
            due_at = parse_datetime(due_at_raw)
            if due_at is None:
                raise ValueError("Unparseable datetime")
            if tz.is_naive(due_at):
                due_at = tz.make_aware(due_at)
        except (ValueError, TypeError) as exc:
            logger.warning(
                "Invalid due_at format provided",
                extra={
                    "user_id": request.user.id,
                    "assignment_id": assignment_id,
                    "due_at_raw": due_at_raw,
                    "error": str(exc),
                }
            )
            return Response({"detail": f"Invalid due_at format: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

        assignment.due_at = due_at
        assignment.save(update_fields=["due_at"])
        logger.info(
            "Assignment due date updated successfully",
            extra={
                "user_id": request.user.id,
                "assignment_id": assignment_id,
                "due_at": assignment.due_at.isoformat(),
            }   )

        return Response({
            "assignment_id": assignment.id,
            "due_at": assignment.due_at.isoformat(),
            "message": "Due date updated successfully.",
        }, status=status.HTTP_200_OK)
        
class CourseCertificateEmailView(APIView):
    """
    After course completion, send certificate email to candidate with PDF attachment.
    POST /api/course-complete-email/
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        user = request.user
        technology_name = request.data.get("technology_name", "")
        score_display = request.data.get("score_display", "")
        percentage = request.data.get("percentage", 0)
        completion_date = request.data.get("completion_date", "")
        pdf_base64 = request.data.get("pdf_base64", "")

        if not pdf_base64:
            logger.warning(
                "pdf_base64 missing in certificate email request",
                extra={"user_id": user.id, "technology_name": technology_name}
            )
            return Response({"error": "pdf_base64 is required."}, status=400)

        # Base64 decode karke PDF bytes banao
        try:
            # "data:application/pdf;base64,XXXX" format handle karo
            if "," in pdf_base64:
                pdf_base64 = pdf_base64.split(",")[1]
            pdf_bytes = base64.b64decode(pdf_base64)
        except Exception as e:
            logger.exception(
                "Invalid PDF base64 data in certificate email request",
                extra={"user_id": user.id, "technology_name": technology_name}
            )
            return Response({"error": f"Invalid PDF data: {str(e)}"}, status=400)

        # Email template context
        context = {
            "candidate_name": f"{user.first_name} {user.last_name}".strip() or user.email,
            "technology_name": technology_name,
            "score_display": score_display,
            "percentage": percentage,
            "completion_date": completion_date,
        }

        subject = f"🎓 Course Completion Certificate — {technology_name}"
        html_message = render_to_string("emails/course_completion.html", context)
        plain_message = strip_tags(html_message)

        # Email with PDF attachment
        from django.core.mail import EmailMessage

        try:
            email = EmailMessage(
                subject=subject,
                body=plain_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[user.email],
            )
            email.content_subtype = "html"
            email.body = html_message

            # PDF attach karo
            filename = f"Certificate_{technology_name.replace(' ', '_')}.pdf"
            email.attach(filename, pdf_bytes, "application/pdf")

            email.send(fail_silently=False)

            logger.info(
                "Course completion certificate email sent successfully",
                extra={
                    "user_id": user.id,
                    "technology_name": technology_name,
                    "percentage": percentage,
                }
            )
        except Exception as e:
            logger.exception(
                "Course completion certificate email failed",
                extra={"user_id": user.id, "technology_name": technology_name}
            )
            return Response(
                {"error": f"Email send failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        return Response(
            {"message": f"Certificate email sent successfully to {user.email}"},
            status=status.HTTP_200_OK
        )        
class UnlockTechnologyAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    FREE_UNLOCK_LIMIT = 5

    def post(self, request):
        user = request.user
        technology_id = request.data.get('technology_id')

        if not technology_id:
            return Response({"error": "technology_id is required."}, status=400)

        if not getattr(user, 'is_individual', False):
            return Response({"error": "Only individual candidates can unlock technologies."}, status=403)

        # Check paid subscription
        sub = getattr(user, 'subscription', None)
        if sub and sub.is_valid() and sub.plan.plan_type != 'free':
            return Response({"message": "All technologies already unlocked for your plan."}, status=200)

        technology = get_object_or_404(Technology, id=technology_id)

        # Check if already assigned (already unlocked)
        if Assignment.objects.filter(user=user, technology=technology, question__isnull=True).exists():
            return Response({"message": "Technology already unlocked."}, status=200)

        # Check free limit — count self-unlocked technologies
        # Check free limit — count ONLY INCOMPLETE self-unlocked technologies
        from django.db.models import Q

        from django.db.models import OuterRef, Subquery, Q

        # Get IDs of assignments where progress is 100
        completed_assignment_ids = UserTechnologyProgress.objects.filter(
            user=user,
            progress=100,
            technology=OuterRef('technology')
        ).values('technology_id')

        unlocked_count = Assignment.objects.filter(
            user=user, 
            assigned_by=user, 
            question__isnull=True
        ).exclude(
            technology_id__in=completed_assignment_ids
        ).count()

        if unlocked_count >= self.FREE_UNLOCK_LIMIT:
            return Response(
                {"error": f"You can only unlock {self.FREE_UNLOCK_LIMIT} technologies on the free plan."},
                status=400
            )

        with transaction.atomic():
            Assignment.objects.create(
                user=user,
                technology=technology,
                assigned_by=user,
                organization=user.organization,
                question=None
            )

            UserTechnologyProgress.objects.get_or_create(
                user=user,
                technology=technology,
                defaults={'progress': 0, 'completed': 0, 'total': technology.questions.count()}
            )

        return Response({"message": "Technology unlocked successfully!"}, status=200)


class LockTechnologyAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        user = request.user
        technology_id = request.data.get('technology_id')

        if not technology_id:
            return Response({"error": "technology_id is required."}, status=400)

        if not getattr(user, 'is_individual', False):
            return Response({"error": "Only individual candidates can lock technologies."}, status=403)

        technology = get_object_or_404(Technology, id=technology_id)

        with transaction.atomic():
            # Only delete the assignment, NOT the progress
            Assignment.objects.filter(user=user, technology=technology, question__isnull=True).delete()
            # Progress preserved for next unlock
            # UserTechnologyProgress.objects.filter(user=user, technology=technology).delete()  # ← COMMENTED OUT

        return Response({"message": "Technology locked successfully! Progress saved for next time."}, status=200)
