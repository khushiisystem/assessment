from __future__ import annotations

import csv
import json
import logging
import os
import uuid
from datetime import timedelta
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from django.conf import settings
from django.contrib.auth import get_user_model, logout, update_session_auth_hash
from django.contrib.auth.hashers import check_password
from django.core.files.storage import default_storage
from django.db import transaction
from django.db.models import Avg, Count, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.utils.text import slugify
from rest_framework import permissions, serializers, status
from rest_framework.authentication import SessionAuthentication, TokenAuthentication
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response
from rest_framework_simplejwt.authentication import JWTAuthentication
import json
from rest_framework.views import APIView
import boto3
from AI_assessment.models import AIAssessment, CandidateAIAssessment
from rest_framework.permissions import IsAuthenticated
from django.db.models import Count, Q

from AI_assessment.utils import send_ai_assessment_notification
from core.models import (
    Assessment,
    AssessmentQuestion,
    CandidateAssessment,
    Category,
    Feedback,
    OTPVerification,
    ProctoringIncident,
    Question,
    Response as CandidateResponse,
    SQLDataset,
    SQLQuestion,
    SQLTestCase,
    TestCase,
    User,
)
from .utils import (
    evaluate_answer,
    execute_code_with_judge0,
    export_candidates_to_excel,
    export_questions_to_excel,
    export_results_to_excel,
    generate_password,
    send_assignment_notification,
    send_candidate_credentials,
    send_assessment_completion_email,
    send_cheating_alert_email,
)
from .utils_sql_judge0 import (
    build_sqlite_script,
    is_select_only,
    parse_rows,
    rowset_equal,
    submit_to_judge0_sql,
)

from .serializers import (
    AssessmentAssignmentSerializer,
    AssessmentUnassignSerializer,
    AssessmentDuplicateSerializer,
    AssessmentSerializer,
    BulkUploadSerializer,
    CandidateAssessmentSerializer,
    CandidateBulkDeleteSerializer,
    AssessmentBulkDeleteSerializer,
    CandidateCreateSerializer,
    CandidateImportSerializer,
    CandidateProfileSerializer,
    CandidateQuickAssignSerializer,
    CandidateRegistrationSerializer,
    CandidateSerializer,
    # CandidateWebhookSerializer,
    ChangePasswordSerializer,
    ForgotPasswordSerializer,
    OTPVerificationSerializer,
    ProctoringIncidentCreateSerializer,
    ProctoringIncidentSerializer,
    QuestionBulkDeleteSerializer,
    QuestionCreateSerializer,
    QuestionImportSerializer,
    QuestionSerializer,
    ResetPasswordOTPSerializer,
    ResetPasswordSerializer,
    RunCodeSerializer,
    SQLDatasetSerializer,
    SQLGradeSerializer,
    SQLRunSerializer,
    SaveAnswerSerializer,
    CategorySerializer,
    CandidateResponseSerializer,
)

from AI_assessment.serializers import (
    AIAssessmentSerializer as AIAssessmentListSerializer,
)
from datetime import datetime, timedelta
from django.utils import timezone

UserModel = get_user_model()

logger = logging.getLogger(__name__)

def _is_empty_value(value):
    return value is None or (isinstance(value, str) and not value.strip())


def _load_rows_from_upload(uploaded_file):
    name = (uploaded_file.name or "").lower()
    if name.endswith(".csv"):
        uploaded_file.seek(0)
        decoded = uploaded_file.read().decode("utf-8-sig")
        return [dict(row) for row in csv.DictReader(decoded.splitlines())]

    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(col).strip() if col is not None else "" for col in rows[0]]
    records = []
    for row in rows[1:]:
        record = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[idx] if idx < len(row) and row[idx] is not None else ""
        records.append(record)
    return records

def normalize_output(value: str) -> str:
    return (value or "").replace("\r\n", "\n").strip()

ADMIN_ROLES = {"super_admin", "org_admin", "manager"}


def is_admin(user: User) -> bool:
    return getattr(user, "role", None) in ADMIN_ROLES


def is_org_admin(user: User) -> bool:
    return getattr(user, "role", None) == "org_admin"


def is_manager(user: User) -> bool:
    return getattr(user, "role", None) == "manager"


def can_access_assessment(user, assessment) -> bool:
    """Object-level visibility for a regular Assessment, mirroring the list rules:
    super_admin -> any; org_admin -> same organization; manager -> only own."""
    role = getattr(user, "role", None)
    if role == "super_admin" or getattr(user, "is_superuser", False):
        return True
    if role == "org_admin":
        org_id = getattr(user, "organization_id", None)
        return org_id is not None and getattr(assessment, "organization_id", None) == org_id
    if role == "manager":
        return getattr(assessment, "created_by_id", None) == user.id
    return False


def is_candidate(user: User) -> bool:
    return getattr(user, "role", None) == "candidate"


class AdminPermission(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and is_admin(request.user))


class SuperAdminPermission(permissions.BasePermission):
    """Only platform super admins (direct candidate add + bulk uploads)."""
    def has_permission(self, request, view):
        u = request.user
        return bool(u and u.is_authenticated and (u.is_superuser or getattr(u, "role", None) == "super_admin"))


def _zecdata_org():
    """The dedicated organization that owns Super-Admin-added candidates."""
    from organization.models import Organization
    org, _ = Organization.objects.get_or_create(
        name="Zecdata",
        defaults={"organization_type": "Company", "status": "Active", "is_active": True},
    )
    return org


class CandidatePermission(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and is_candidate(request.user))


# class IsAdminOrReadOnly(permissions.BasePermission):
#     """
#     Allow only authenticated users to read data,
#     and only admins (staff) to modify data.
#     """
#     def has_permission(self, request, view):
#         if not request.user or not request.user.is_authenticated:
#             return False
#         if request.method in permissions.SAFE_METHODS:
#             return True  
#         return request.user.is_staff



class StandardResultsSetPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = "page_size"
    max_page_size = 200


def send_otp_sms(phone: str, otp_code: str) -> bool:
    logger.info("Sending OTP SMS", extra={"phone": phone, "otp_code": otp_code})
    return True


def send_otp_email(email: str, otp_code: str, purpose: str = "registration") -> bool:
    from django.core.mail import send_mail
    from django.template.loader import render_to_string

    subject = f"OTP for {purpose.title()}"
    message = f"Your OTP for {purpose} is: {otp_code}\\n\\nThis OTP is valid for 10 minutes."
    try:
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [email],
            fail_silently=False,
            html_message=render_to_string(
                'emails/sys_otp.html',
                {
                    'otp_code': otp_code,
                    'purpose': purpose,
                },
            ),
        )
        return True
    except Exception as exc:
        logger.exception("Failed to send OTP email", extra={"email": email})
        return False


# def _store_candidate_resume(file_obj, candidate_email: str) -> Optional[str]:
#     if not file_obj:
#         return None
#     ext = os.path.splitext(file_obj.name)[1]
#     key = f"{candidate_email.lower()}/resumes/{uuid.uuid4().hex}{ext}"
#     saved_path = default_storage.save(key, file_obj)
#     return default_storage.url(saved_path)

def _split_name(full_name: str) -> tuple[str, str]:
    parts = (full_name or "").strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def _make_unique_username(base: str) -> str:
    candidate = base
    counter = 1
    while UserModel.objects.filter(username=candidate).exists():
        candidate = f"{base}{counter}"
        counter += 1
    return candidate


def _derive_username(email: str, name: str, phone: str) -> str:
    if email:
        local = email.split("@")[0]
        base = slugify(local)[:30] or None
        if base:
            return _make_unique_username(base)
    first, last = _split_name(name)
    base = slugify((first or "user") + (last[:1] or ""))[:20]
    tail = (phone[-4:] if phone else get_random_string(4).lower())
    return _make_unique_username(f"{base}{tail}")


class APIRootView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, *args, **kwargs):
        return Response(
            {
                "message": "Training & Certification API",
                "user": request.user.email,
                "role": request.user.role,
            }
        )


class CandidateRegisterAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request, *args, **kwargs):
        logger.info("Candidate registration attempt started")

        serializer = CandidateRegistrationSerializer(data=request.data)

        if not serializer.is_valid():
            logger.warning(
                "Candidate registration validation failed",
                extra={"errors": serializer.errors}
            )
            serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        try:
            resume_file = data["resume"]
            logger.info("Resume received", extra={"resume_filename": resume_file.name})

            file_ext = os.path.splitext(resume_file.name)[1]
            temp_key = f"temp_resumes/{uuid.uuid4().hex}{file_ext}"
            temp_path = default_storage.save(temp_key, resume_file)
            temp_url = default_storage.url(temp_path)

            logger.info("Resume stored temporarily", extra={"path": temp_path})
            
            temp_data = {
                "first_name": data["first_name"],
                "last_name": data["last_name"],
                "email": data["email"],
                "phone": data["phone"],
                "profile": data.get("profile", ""),
                "organization_id": data.get("organization_id"),
                "resume_temp_s3_path": temp_path,
                "resume_temp_s3_url": temp_url,
                "resume_original_name": resume_file.name,
            }

            otp_code = OTPVerification.generate_otp()

            otp_record = OTPVerification.objects.create(
                phone=data["phone"],
                email=data["email"],
                otp_code=otp_code,
                otp_type="registration",
                temp_data=temp_data,
                expires_at=timezone.now() + timedelta(minutes=1),
            )

            logger.info(
                "OTP generated",
                extra={"otp_id": otp_record.id, "email": data["email"]}
            )

            sms_sent = send_otp_sms(data["phone"], otp_code)
            email_sent = send_otp_email(data["email"], otp_code, "registration")

            if not (sms_sent or email_sent):
                logger.error("Failed to send OTP", extra={"email": data["email"]})
                return Response(
                    {"detail": "Failed to send OTP. Try again later."},
                    status=status.HTTP_503_SERVICE_UNAVAILABLE,
                )

            logger.info("OTP sent successfully", extra={"otp_id": otp_record.id})

            return Response({
                "otp_id": otp_record.id,
                "resume_temp_url": temp_url,
                "message": "OTP sent successfully.",
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            logger.exception("Error during candidate registration")
            return Response(
                {"detail": "Something went wrong"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class VerifyRegistrationOTPAPI(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request, otp_id):
        logger.info("OTP verification attempt", extra={"otp_id": otp_id})
        
        # Validate OTP record exists
        try:
            otp_record = OTPVerification.objects.get(
                id=otp_id,
                otp_type='registration',
                is_verified=False
            )
        except OTPVerification.DoesNotExist:
            logger.warning("Invalid or expired OTP session", extra={"otp_id": otp_id})
            return Response({
                "status": False,
                "message": "Invalid or expired OTP session."
            }, status=status.HTTP_400_BAD_REQUEST)

        # Check expiry
        if otp_record.is_expired():
            logger.warning("OTP has expired", extra={"otp_id": otp_id})
            return Response({
                "status": False,
                "message": "OTP has expired. Please register again."
            }, status=status.HTTP_400_BAD_REQUEST)

        # Validate OTP Code
        serializer = OTPVerificationSerializer(data=request.data)
        if not serializer.is_valid():
            logger.warning("OTP validation failed", extra={"otp_id": otp_id, "errors": serializer.errors})
            return Response({
                "status": False,
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        entered_otp = serializer.validated_data["otp_code"]

        # Check OTP match
        if entered_otp != otp_record.otp_code:
            logger.warning("Invalid OTP entered", extra={"otp_id": otp_id})
            return Response({
                "status": False,
                "message": "Invalid OTP. Please try again."
            }, status=status.HTTP_400_BAD_REQUEST)

        # Get temp data from OTP
        temp_data = otp_record.temp_data

        try:
            # Generate username + password
            username = _derive_username(
                temp_data["email"],
                f"{temp_data['first_name']} {temp_data['last_name']}",
                temp_data["phone"]
            )

            raw_password = get_random_string(12)

            # Create user
            user = User.objects.create_user(
                username=username,
                email=temp_data['email'],
                first_name=temp_data['first_name'],
                last_name=temp_data['last_name'],
                password=raw_password,
                phone=temp_data['phone'],
                profile=temp_data.get('profile', ''),
                role="candidate"
            )

            logger.info("User created successfully", extra={"user_id": user.id, "email": user.email})

        # ---------- Move RESUME from TEMP → PERMANENT S3 ----------
        
            if "resume_temp_s3_path" in temp_data:
                temp_s3_path = temp_data["resume_temp_s3_path"]

                # Initialize S3
                s3_client = boto3.client(
                    "s3",
                    aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                    aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                    region_name=settings.AWS_S3_REGION_NAME,
                )

                # Permanent S3 key
                file_extension = os.path.splitext(temp_data["resume_original_name"])[1].lower()
                candidate_email = user.email.lower().replace(" ", "")
                unique_filename = f"resume_{uuid.uuid4().hex[:8]}{file_extension}"

                permanent_s3_key = f"{candidate_email}/resumes/{unique_filename}"

                # Copy file data
                if default_storage.exists(temp_s3_path):
                    temp_file = default_storage.open(temp_s3_path)

                    s3_client.upload_fileobj(
                        temp_file,
                        settings.AWS_STORAGE_BUCKET_NAME,
                        permanent_s3_key,
                        ExtraArgs={"ContentType": "application/pdf"}
                    )

                    temp_file.close()
                    default_storage.delete(temp_s3_path)

                    # Set final URL
                    final_url = (
                        f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3."
                        f"{settings.AWS_S3_REGION_NAME}.amazonaws.com/{permanent_s3_key}"
                    )

                    user.resume_s3_url = final_url
                    user.save()

                    logger.info("Resume moved to permanent storage", extra={"user_id": user.id})

        except Exception as e:
            logger.exception("Error during user creation and resume processing", extra={"otp_id": otp_id})
            return Response(
                {"detail": "Registration failed. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # Mark OTP verified
        otp_record.is_verified = True
        otp_record.save()

        # Send email with credentials
        try:
            send_candidate_credentials(user, raw_password)
            logger.info("Credentials email sent", extra={"user_id": user.id})
        except Exception as e:
            logger.exception("Failed to send credentials email", extra={"user_id": user.id})

        logger.info("Registration completed successfully", extra={"user_id": user.id, "otp_id": otp_id})
        
        return Response({
            "status": True,
            "message": "Registration successful!",
            "username": username,
            "password_sent_to": user.email
        }, status=status.HTTP_200_OK)



class ResendRegistrationOTPAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request, otp_id: int, *args, **kwargs):
        logger.info("OTP resend request", extra={"otp_id": otp_id})
        
        try:
            otp_record = get_object_or_404(
                OTPVerification,
                id=otp_id,
                otp_type="registration",
                is_verified=False,
            )
        except Exception as e:
            logger.warning("Invalid OTP record for resend", extra={"otp_id": otp_id})
            return Response(
                {"detail": "Invalid OTP session."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        new_code = OTPVerification.generate_otp()
        otp_record.otp_code = new_code
        otp_record.expires_at = timezone.now() + timedelta(minutes=1)
        otp_record.save(update_fields=["otp_code", "expires_at"])

        logger.info("New OTP generated for resend", extra={"otp_id": otp_id})

        sms_sent = send_otp_sms(otp_record.phone, new_code)
        email_sent = send_otp_email(otp_record.email, new_code, "registration")

        if not (sms_sent or email_sent):
            logger.error("Failed to resend OTP", extra={"otp_id": otp_id, "email": otp_record.email})
            return Response(
                {"detail": "Unable to send OTP."},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        logger.info("OTP resent successfully", extra={"otp_id": otp_id})
        return Response({"message": "OTP resent."})


class AssessmentStatusView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, *args, **kwargs):
        logger.info("Assessment status check", extra={"user_id": request.user.id, "user_role": getattr(request.user, 'role', None)})
        
        if not is_candidate(request.user):
            logger.info("Non-candidate user accessing assessment status", extra={"user_id": request.user.id})
            return Response({"in_progress": False})

        try:
            candidate_assessment = (
                CandidateAssessment.objects.filter(
                    candidate=request.user, status="in_progress"
                )
                .select_related("assessment")
                .first()
            )

            if not candidate_assessment:
                logger.info("No in-progress assessment found", extra={"user_id": request.user.id})
                return Response({"in_progress": False})

            logger.info("In-progress assessment found", extra={
                "user_id": request.user.id,
                "assessment_id": candidate_assessment.assessment_id
            })
            
            return Response(
                {
                    "in_progress": True,
                    "assessment_id": candidate_assessment.assessment_id,
                    "assessment_title": candidate_assessment.assessment.title,
                }
            )
        except Exception as e:
            logger.exception("Error checking assessment status", extra={"user_id": request.user.id})
            return Response(
                {"detail": "Failed to check assessment status"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class AdminDashboardView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, *args, **kwargs):
        logger.info("Admin dashboard data requested", extra={"user_id": request.user.id})
        
        try:
            from organization.context import current_organization_id, current_user_is_super_admin
            is_super = current_user_is_super_admin.get()
            org_id = current_organization_id.get()
            
            logger.error(f"DASHBOARD DEBUG: user={request.user.email}, role={request.user.role}, is_super_context={is_super}, org_id_context={org_id}, is_superuser_attr={request.user.is_superuser}")

            if not is_super:
                # If they are NOT a super admin, strictly filter by their organization.
                # If they have no organization, they see nothing.
                candidates_qs = UserModel.objects.filter(role="candidate")
                if org_id:
                    candidates_qs = candidates_qs.filter(organization_id=org_id)
                else:
                    candidates_qs = candidates_qs.none()
                
                total_candidates = candidates_qs.count()
                
                # Assessments: Show only those created by this admin
                assessments_qs = Assessment.objects.filter(created_by=request.user)
                total_assessments = assessments_qs.count() 
                
                # Questions: Show only those created by this admin
                questions_qs = Question.objects.filter(created_by=request.user)
                total_questions = questions_qs.count()
                
                active_assessments = assessments_qs.filter(is_active=True).count()
                
                try:
                    total_assessments += AIAssessment.objects.filter(created_by=request.user).count()
                except Exception as e:
                    logger.warning("Failed to count AI assessments", extra={"error": str(e)})

                completed_regular = CandidateAssessment.objects.filter(status="completed", candidate__organization_id=org_id).count()
                try:
                    completed_ai = CandidateAIAssessment.objects.filter(status="completed", candidate__organization_id=org_id).count()
                except Exception as e:
                    logger.warning("Failed to count AI assessment completions", extra={"error": str(e)})
                    completed_ai = 0

                recent_assessments = list(
                    assessments_qs.order_by("-created_at")[:5].values(
                        "id", "title", "start_date", "end_date", "is_active"
                    )
                )
                
                recent_candidates = list(
                    candidates_qs.order_by("-date_joined")[:5]
                    .values("id", "first_name", "last_name", "email", "date_joined")
                )
            else:
                total_candidates = UserModel.objects.filter(role="candidate").count()
                total_assessments = Assessment.objects.count()
                total_questions = Question.objects.count()
                active_assessments = Assessment.objects.filter(is_active=True).count()

                try:
                    total_assessments += AIAssessment.objects.count()
                except Exception as e:
                    logger.warning("Failed to count AI assessments", extra={"error": str(e)})

                completed_regular = CandidateAssessment.objects.filter(status="completed").count()
                try:
                    completed_ai = CandidateAIAssessment.objects.filter(status="completed").count()
                except Exception as e:
                    logger.warning("Failed to count AI assessment completions", extra={"error": str(e)})
                    completed_ai = 0

                recent_assessments = list(
                    Assessment.objects.order_by("-created_at")[:5].values(
                        "id", "title", "start_date", "end_date", "is_active"
                    )
                )
                recent_candidates = list(
                    UserModel.objects.filter(role="candidate")
                    .order_by("-date_joined")[:5]
                    .values("id", "first_name", "last_name", "email", "date_joined")
                )

            # --- Org-scoped aggregates for the dashboard KPIs (no schema change) ---
            # Reuse the same scope the view already applies: super admins see the
            # whole platform; org admins see their whole organization.
            from django.db.models import Avg
            cand_filter = {} if is_super else {"candidate__organization_id": org_id}

            # Completion rate = completed (regular + AI) / total assigned (regular + AI).
            try:
                total_assigned = (
                    CandidateAssessment.objects.filter(**cand_filter).count()
                    + CandidateAIAssessment.objects.filter(**cand_filter).count()
                )
                completed_total = completed_regular + completed_ai
                completion_rate = round(completed_total / total_assigned * 100, 1) if total_assigned else 0.0
            except Exception as e:
                logger.warning("Failed to compute completion rate", extra={"error": str(e)})
                completion_rate = 0.0

            # Average pass score across completed regular assessments.
            try:
                avg_reg = CandidateAssessment.objects.filter(
                    status="completed", **cand_filter
                ).aggregate(a=Avg("percentage"))["a"]
                average_pass_score = round(avg_reg or 0, 1)
            except Exception as e:
                logger.warning("Failed to compute average pass score", extra={"error": str(e)})
                average_pass_score = 0.0

            # Integrity flags = proctoring incidents recorded for in-scope candidates.
            try:
                from core.models import ProctoringIncident
                integrity_flags_count = ProctoringIncident.objects.filter(**cand_filter).count()
            except Exception as e:
                logger.warning("Failed to count proctoring incidents", extra={"error": str(e)})
                integrity_flags_count = 0

            # Org's candidate invite limit (for Plan Usage). None = unlimited / platform.
            candidate_limit = None
            if not is_super and org_id:
                try:
                    from organization.models import Organization
                    _org = Organization.objects.filter(id=org_id).first()
                    candidate_limit = _org.candidate_limit if _org else None
                except Exception:
                    candidate_limit = None

            # --- Org-admin only: Workforce Funnel + Organization Health ---
            # Computed from EXISTING data (no schema change / no migration). Left as
            # None for super admins and other roles so nothing else changes.
            workforce_funnel = None
            org_health = None
            if not is_super and org_id:
                from django.db.models import Avg
                from django.utils import timezone
                from datetime import timedelta

                # Workforce Funnel — distinct people who reached each stage.
                try:
                    from learning.models import UserTechnologyProgress
                    ltp = UserTechnologyProgress.objects.filter(
                        user__organization_id=org_id, user__role="candidate"
                    )

                    invited = total_candidates  # every candidate account is an invite
                    registered = UserModel.objects.filter(
                        role="candidate", organization_id=org_id, last_login__isnull=False
                    ).count()
                    course_started = ltp.filter(progress__gt=0).values("user").distinct().count()
                    course_completed = ltp.filter(progress__gte=100).values("user").distinct().count()
                    assessment_done = (
                        CandidateAssessment.objects.filter(
                            status="completed", candidate__organization_id=org_id
                        ).values("candidate").distinct().count()
                    )
                    try:
                        ai_done = (
                            CandidateAIAssessment.objects.filter(
                                status="completed", candidate__organization_id=org_id
                            ).values("candidate").distinct().count()
                        )
                    except Exception:
                        ai_done = 0

                    workforce_funnel = [
                        {"key": "invited", "label": "Invited", "value": invited},
                        {"key": "registered", "label": "Registered", "value": registered},
                        {"key": "course_started", "label": "Course Started", "value": course_started},
                        {"key": "course_completed", "label": "Course Completed", "value": course_completed},
                        {"key": "assessment_done", "label": "Assessment Done", "value": assessment_done},
                        {"key": "ai_assessment", "label": "AI Assessment", "value": ai_done},
                    ]
                except Exception as e:
                    logger.warning("Failed to compute workforce funnel", extra={"error": str(e)})
                    workforce_funnel = None

                # Organization Health — category scores on a 0-100 scale.
                try:
                    from learning.models import UserTechnologyProgress

                    # Learning = avg course progress (already 0-100).
                    learning = round(
                        UserTechnologyProgress.objects.filter(
                            user__organization_id=org_id, user__role="candidate"
                        ).aggregate(a=Avg("progress"))["a"] or 0, 1
                    )

                    # Assessment = avg pass score on completed regular assessments (0-100).
                    assessment = round(average_pass_score, 1)

                    # AI Assessment = avg overall_score (0-10) scaled to 0-100.
                    try:
                        ai_raw = CandidateAIAssessment.objects.filter(
                            status="completed", candidate__organization_id=org_id
                        ).aggregate(a=Avg("overall_score"))["a"] or 0
                        ai_assessment = round(ai_raw * 10, 1)
                    except Exception:
                        ai_assessment = 0.0

                    # Engagement = % of candidates active in the last 30 days.
                    active_cutoff = timezone.now() - timedelta(days=30)
                    active_recent = UserModel.objects.filter(
                        role="candidate", organization_id=org_id, last_login__gte=active_cutoff
                    ).count()
                    engagement = round(active_recent / total_candidates * 100, 1) if total_candidates else 0.0

                    categories = [learning, assessment, ai_assessment, engagement]
                    overall = round(sum(categories) / len(categories))

                    org_health = {
                        "overall_score": overall,
                        "learning": learning,
                        "assessment": assessment,
                        "ai_assessment": ai_assessment,
                        "engagement": engagement,
                    }
                except Exception as e:
                    logger.warning("Failed to compute organization health", extra={"error": str(e)})
                    org_health = None

            # --- Org-admin only: Top Performers + Attention Required ---
            # Computed from EXISTING data (no schema change). None for super admins.
            top_performers = None
            attention_required = None
            if not is_super and org_id:
                from django.db.models import Avg, Q, F
                from django.utils import timezone
                from datetime import timedelta

                # Top Performers — highest average score on completed assessments.
                try:
                    rows = (
                        CandidateAssessment.objects.filter(
                            status="completed", candidate__organization_id=org_id
                        )
                        .values(
                            "candidate__id",
                            "candidate__first_name",
                            "candidate__last_name",
                            "candidate__email",
                        )
                        .annotate(score=Avg("percentage"))
                        .order_by("-score")[:5]
                    )
                    top_performers = [
                        {
                            "id": r["candidate__id"],
                            "name": (
                                f"{r['candidate__first_name'] or ''} {r['candidate__last_name'] or ''}".strip()
                                or r["candidate__email"]
                            ),
                            "score": round(r["score"] or 0),
                        }
                        for r in rows
                    ]
                except Exception as e:
                    logger.warning("Failed to compute top performers", extra={"error": str(e)})
                    top_performers = None

                # Attention Required — actionable counts the admin can follow up on.
                try:
                    from learning.models import UserTechnologyProgress

                    active_cutoff = timezone.now() - timedelta(days=30)
                    cand_qs = UserModel.objects.filter(role="candidate", organization_id=org_id)

                    # Inactive = no login in the last 30 days (or never logged in).
                    inactive = cand_qs.filter(
                        Q(last_login__lt=active_cutoff) | Q(last_login__isnull=True)
                    ).count()

                    # Not started learning = candidates with no course progress yet.
                    started_ids = (
                        UserTechnologyProgress.objects.filter(
                            user__organization_id=org_id, user__role="candidate", progress__gt=0
                        )
                        .values_list("user_id", flat=True)
                        .distinct()
                    )
                    not_started = cand_qs.exclude(id__in=started_ids).count()

                    # Failed = completed assessments scored below their passing mark.
                    failed = CandidateAssessment.objects.filter(
                        status="completed",
                        candidate__organization_id=org_id,
                        assessment__passing_percentage__gt=0,
                        percentage__lt=F("assessment__passing_percentage"),
                    ).count()

                    # AI assessments pending = assigned AI assessments not yet completed.
                    try:
                        ai_pending = (
                            CandidateAIAssessment.objects.filter(candidate__organization_id=org_id)
                            .exclude(status="completed")
                            .count()
                        )
                    except Exception:
                        ai_pending = 0

                    attention_required = [
                        {"key": "inactive", "label": "inactive employees", "count": inactive},
                        {"key": "not_started", "label": "not started learning", "count": not_started},
                        {"key": "failed", "label": "failed assessments", "count": failed},
                        {"key": "ai_pending", "label": "AI assessments pending", "count": ai_pending},
                    ]
                except Exception as e:
                    logger.warning("Failed to compute attention required", extra={"error": str(e)})
                    attention_required = None

            dashboard_data = {
                "total_candidates": total_candidates,
                "total_assessments": total_assessments,
                "active_assessments": active_assessments,
                "total_questions": total_questions,
                "completed_assessments": completed_regular + completed_ai,
                "completion_rate": completion_rate,
                "average_pass_score": average_pass_score,
                "integrity_flags_count": integrity_flags_count,
                "candidate_limit": candidate_limit,
                "recent_assessments": recent_assessments,
                "recent_candidates": recent_candidates,
                "workforce_funnel": workforce_funnel,
                "org_health": org_health,
                "top_performers": top_performers,
                "attention_required": attention_required,
            }
            
            logger.info("Admin dashboard data retrieved successfully", extra={
                "user_id": request.user.id,
                "total_candidates": total_candidates,
                "total_assessments": total_assessments
            })

            # TEMP DIAGNOSTIC — confirm the org-admin branch ran and the new
            # widgets are populated. Remove once verified.
            logger.error(
                f"DASHBOARD WIDGETS DEBUG: is_super={is_super}, org_id={org_id}, "
                f"workforce_funnel={'set' if workforce_funnel else 'None'}, "
                f"org_health={'set' if org_health else 'None'}, "
                f"funnel_values={[s['value'] for s in workforce_funnel] if workforce_funnel else None}, "
                f"org_health={org_health}"
            )
            
            return Response(dashboard_data)
            
        except Exception as e:
            logger.exception("Error retrieving admin dashboard data", extra={"user_id": request.user.id})
            return Response(
                {"detail": "Failed to retrieve dashboard data"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CandidateDashboardView(APIView):
    permission_classes = [CandidatePermission]

    def get(self, request, *args, **kwargs):
        logger.info("Candidate dashboard data requested", extra={"user_id": request.user.id})
        
        try:
            candidate_assessments = CandidateAssessment.objects.filter(candidate=request.user)
            assigned = candidate_assessments.filter(status="assigned")
            in_progress = candidate_assessments.filter(status="in_progress")
            completed = candidate_assessments.filter(status="completed")

            upcoming = (
                Assessment.objects.filter(is_active=True, start_date__gt=timezone.now())
                .exclude(id__in=candidate_assessments.values_list("assessment_id", flat=True))
                .order_by("start_date")[:5]
            )

            avg_score = completed.aggregate(avg=Avg("percentage"))["avg"] or 0

            ai_data = {"assigned": 0, "in_progress": 0, "completed": 0, "average_score": 0}
            try:
                ai_assessments = CandidateAIAssessment.objects.filter(candidate=request.user)
                ai_data["assigned"] = ai_assessments.filter(status="assigned").count()
                ai_data["in_progress"] = ai_assessments.filter(status="in_progress").count()
                completed_ai = ai_assessments.filter(status="completed")
                ai_data["completed"] = completed_ai.count()
                ai_data["average_score"] = completed_ai.aggregate(avg=Avg("overall_score"))["avg"] or 0
            except Exception as e:
                logger.warning("Failed to retrieve AI assessment data", extra={"user_id": request.user.id, "error": str(e)})

            dashboard_data = {
                "assigned": assigned.count(),
                "in_progress": in_progress.count(),
                "completed": completed.count(),
                "average_score": round(avg_score, 2) if avg_score else 0,
                "upcoming_assessments": list(
                    upcoming.values("id", "title", "start_date", "end_date", "duration")
                ),
                "ai_summary": ai_data,
            }
            
            logger.info("Candidate dashboard data retrieved successfully", extra={
                "user_id": request.user.id,
                "assigned_count": assigned.count(),
                "completed_count": completed.count()
            })
            
            return Response(dashboard_data)
            
        except Exception as e:
            logger.exception("Error retrieving candidate dashboard data", extra={"user_id": request.user.id})
            return Response(
                {"detail": "Failed to retrieve dashboard data"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



from django.db.models.functions import Concat
from django.db.models import Value as V, Q

class CandidateListView(APIView, StandardResultsSetPagination):
    permission_classes = [AdminPermission]

    def get(self, request, *args, **kwargs):
        logger.error(f"CANDIDATE LIST DEBUG: user={request.user.email}, role={request.user.role}, is_superuser={request.user.is_superuser}")
        logger.info("Candidate list requested", extra={"user_id": request.user.id})
        
        try:
            queryset = (
                UserModel.objects.filter(role="candidate")
                .order_by("-date_joined")
                .prefetch_related("assignments__technology")
            )

            from organization.context import current_organization_id, current_user_is_super_admin
            is_super_ctx = current_user_is_super_admin.get()
            org_id_ctx = current_organization_id.get()
            
            logger.error(f"VIEW FORENSIC: user={request.user.email}, id={request.user.id}, role={getattr(request.user, 'role', 'N/A')}, is_staff={request.user.is_staff}, is_superuser={request.user.is_superuser}, is_super_ctx={is_super_ctx}, org_id_ctx={org_id_ctx}")

            if not is_super_ctx:
                if org_id_ctx:
                    # Scope to the organization (tenant boundary) so the admin sees
                    # ALL their org's candidates — including invited ones and those
                    # created by other admins — not just ones they personally created.
                    queryset = queryset.filter(organization_id=org_id_ctx)
                    logger.error(f"VIEW FORENSIC: Applied filter org={org_id_ctx}")
                else:
                    queryset = queryset.none()
                    logger.error("VIEW FORENSIC: No organization ID in context, returning empty")
            else:
                logger.error("VIEW FORENSIC: Bypassing isolation because is_super_ctx is TRUE")

            # Annotate full name for search
            queryset = queryset.annotate(
                full_name_search=Concat('first_name', V(' '), 'last_name')
            )

            # --- Query Params ---
            search = request.query_params.get("search", "").strip()
            exact = request.query_params.get("exact", "").strip()
            profile = request.query_params.get("profile", "").strip()
            technology = request.query_params.get("technology", "").strip()
            joined_from = request.query_params.get("joined_from")
            joined_to = request.query_params.get("joined_to")
            email_domains = request.query_params.get("email_domains", "").strip()

            logger.info("Candidate search filters applied", extra={
                "user_id": request.user.id,
                "search": search,
                "profile": profile,
                "technology": technology
            })

            # ---------- PARTIAL SEARCH ----------
            if search:
                queryset = queryset.filter(
                    Q(username__icontains=search)
                    | Q(email__icontains=search)
                    | Q(first_name__icontains=search)
                    | Q(last_name__icontains=search)
                    | Q(full_name_search__icontains=search)
                )

            # ---------- EXACT SEARCH ----------
            if exact:
                queryset = queryset.filter(
                    Q(username__iexact=exact)
                    | Q(email__iexact=exact)
                    | Q(first_name__iexact=exact)
                    | Q(last_name__iexact=exact)
                    | Q(full_name_search__iexact=exact)
                )
            exclude_domains = request.query_params.getlist("exclude_email_domains")
            if exclude_domains:
                for domain in exclude_domains:
                    queryset = queryset.exclude(email__iendswith=f"@{domain}")

            if email_domains:
                domains = [d.strip() for d in email_domains.split(",") if d.strip()]
                domain_query = Q()
                for domain in domains:
                    domain_query |= Q(email__iendswith=f"@{domain}")
                queryset = queryset.filter(domain_query)

            # ---------- PROFILE FILTER ----------
            if profile:
                queryset = queryset.filter(profile__icontains=profile)

            # ---------- TECHNOLOGY FILTER ----------
            if technology:
                queryset = queryset.filter(
                    assignments__technology__name__icontains=technology
                ).distinct()  

            # ---------- DATE FILTERS ----------

            if joined_from:
                try:
                    joined_from_date = datetime.strptime(joined_from, "%Y-%m-%d")
                    queryset = queryset.filter(date_joined__gte=joined_from_date)
                except Exception as e:
                    logger.warning("Invalid joined_from date format", extra={"user_id": request.user.id, "joined_from": joined_from})

            if joined_to:
                try:
                    joined_to_date = datetime.strptime(joined_to, "%Y-%m-%d") + timedelta(days=1)
                    queryset = queryset.filter(date_joined__lt=joined_to_date)
                except Exception as e:
                    logger.warning("Invalid joined_to date format", extra={"user_id": request.user.id, "joined_to": joined_to})

            # ---------- PAGINATION ----------
            page = self.paginate_queryset(queryset, request, view=self)
            serializer = CandidateSerializer(page, many=True)

            # ---------- EXTRA DATA ----------
            available_assessments = list(
                Assessment.objects.filter(end_date__gt=timezone.now())
                .order_by("-created_at")
                .values("id", "title", "end_date")
            )

            try:
                ai_assessments = list(
                    AIAssessment.objects.filter(end_date__gt=timezone.now())
                    .order_by("-created_at")
                    .values("id", "title", "end_date")
                )
            except Exception as e:
                logger.warning("Failed to retrieve AI assessments", extra={"user_id": request.user.id, "error": str(e)})
                ai_assessments = []
                
            response_data = {
                "candidates": serializer.data,
                "available_assessments": available_assessments,
                "ai_assessments": ai_assessments,
            }
            
            logger.info("Candidate list retrieved successfully", extra={
                "user_id": request.user.id,
                "total_candidates": queryset.count(),
                "filtered_candidates": len(serializer.data) if serializer.data else 0
            })
            
            return self.get_paginated_response(response_data)
            
        except Exception as e:
            logger.exception("Error retrieving candidate list", extra={"user_id": request.user.id})
            return Response(
                {"detail": "Failed to retrieve candidate list"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



class CandidateCreateView(APIView):
    # Direct candidate add is Super-Admin only; org admins must invite instead.
    permission_classes = [SuperAdminPermission]

    def post(self, request, *args, **kwargs):
        logger.info("Candidate creation attempt", extra={"user_id": request.user.id})
        
        try:
            serializer = CandidateCreateSerializer(data=request.data)
            if not serializer.is_valid():
                logger.warning("Candidate creation validation failed", extra={
                    "user_id": request.user.id,
                    "errors": serializer.errors
                })
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            # Enforce the organization's candidate invite limit (null = unlimited).
            org = getattr(request.user, "organization", None)
            if org and org.candidate_limit is not None:
                current = UserModel.objects.filter(organization=org, role="candidate").count()
                if current >= org.candidate_limit:
                    logger.warning(
                        "Candidate limit reached",
                        extra={"organization_id": org.id, "limit": org.candidate_limit},
                    )
                    return Response(
                        {"detail": f"Candidate limit reached ({org.candidate_limit}). Ask your administrator to raise it."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            resume_file = request.FILES.get("resume")
            if resume_file:
                logger.info("Resume file received for candidate creation", extra={
                    "user_id": request.user.id,
                    "resume_filename": resume_file.name 
                })
            with transaction.atomic():
                # Super-Admin-added candidates belong to the dedicated Zecdata org.
                user = serializer.save(
                    organization=_zecdata_org(),
                    created_by=request.user
                )
                
                logger.info("Candidate created successfully", extra={
                    "user_id": user.id,
                    "email": user.email,
                    "organization_id": user.organization_id
                })

                # ---------- SAME LOGIC AS VerifyRegistrationOTPAPI ----------
                try:
                    if resume_file:

                        # Initialize S3
                        s3_client = boto3.client(
                            "s3",
                            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                            region_name=settings.AWS_S3_REGION_NAME,
                        )

                        # Permanent S3 key
                        file_extension = os.path.splitext(resume_file.name)[1].lower()
                        candidate_email = user.email.lower().replace(" ", "")
                        unique_filename = f"resume_{uuid.uuid4().hex[:8]}{file_extension}"

                        permanent_s3_key = f"{candidate_email}/resumes/{unique_filename}"

                        # Upload directly from InMemoryUploadedFile
                        s3_client.upload_fileobj(
                            resume_file,
                            settings.AWS_STORAGE_BUCKET_NAME,
                            permanent_s3_key,
                            ExtraArgs={"ContentType": "application/pdf"},
                        )

                        # Final public S3 URL
                        final_url = (
                            f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3."
                            f"{settings.AWS_S3_REGION_NAME}.amazonaws.com/{permanent_s3_key}"
                        )

                        # Save in DB
                        user.resume_s3_url = final_url
                        user.save(update_fields=["resume_s3_url"])
                        
                        logger.info("Resume uploaded successfully for candidate", extra={
                            "user_id": request.user.id,
                            "candidate_id": user.id
                        })

                except Exception as e:
                    logger.exception("Error uploading resume to S3 for candidate", extra={
                        "user_id": request.user.id,
                        "candidate_id": user.id
                    })
                # -------------------------------------------------------------

            # Send credentials
            raw_password = getattr(user, "_raw_password", None)
            if raw_password:
                try:
                    send_candidate_credentials(user, raw_password)
                    logger.info("Credentials email sent to new candidate", extra={
                        "user_id": request.user.id,
                        "candidate_id": user.id
                    })
                except Exception as e:
                    logger.exception("Failed to send credentials email to candidate", extra={
                        "user_id": request.user.id,
                        "candidate_id": user.id
                    })

            logger.info("Candidate created successfully", extra={
                "user_id": request.user.id,
                "candidate_id": user.id,
                "candidate_email": user.email
            })
            
            return Response(CandidateSerializer(user).data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.exception("Error during candidate creation", extra={"user_id": request.user.id})
            return Response(
                {"detail": "Failed to create candidate"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CandidateImportView(APIView):
    # Bulk candidate upload is Super-Admin only.
    permission_classes = [SuperAdminPermission]

    def post(self, request, *args, **kwargs):
        logger.info("Candidate import attempt", extra={"user_id": request.user.id})
        
        try:
            serializer = CandidateImportSerializer(data=request.data)
            if not serializer.is_valid():
                logger.warning("Candidate import validation failed", extra={
                    "user_id": request.user.id,
                    "errors": serializer.errors
                })
                serializer.is_valid(raise_exception=True)

            upload = serializer.validated_data["file"]
            logger.info("Import file received", extra={
    "user_id": request.user.id,
    "upload_filename": upload.name
})
            
            try:
                rows = _load_rows_from_upload(upload)
                logger.info("File loaded successfully", extra={
                    "user_id": request.user.id,
                    "total_rows": len(rows)
                })
            except Exception as exc:
                logger.error("Failed to read import file", extra={
                    "user_id": request.user.id,
                    "error": str(exc),
                    "filename": upload.name
                })
                return Response(
                    {"detail": f"Unable to read file: {exc}"}, status=status.HTTP_400_BAD_REQUEST
                )

            # Enforce the organization's candidate invite limit before importing.
            org = getattr(request.user, "organization", None)
            if org and org.candidate_limit is not None:
                current = UserModel.objects.filter(organization=org, role="candidate").count()
                if current + len(rows) > org.candidate_limit:
                    return Response(
                        {
                            "detail": (
                                f"Import would exceed the candidate limit ({org.candidate_limit}). "
                                f"Current: {current}, in file: {len(rows)}. Ask your administrator to raise it."
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            success, failed, errors = 0, 0, []

            for index, row in enumerate(rows):
                try:
                    email = str(row.get("email", "") or "").strip().lower()
                    # Reject duplicate emails up front with a clear message.
                    if email and UserModel.objects.filter(email__iexact=email).exists():
                        raise ValueError(f"Email '{email}' already exists")
                    raw_username = str(row.get("username") or "")
                    if raw_username:
                        username = _make_unique_username(slugify(raw_username))
                    elif email:
                        username = _make_unique_username(email.split("@")[0])
                    else:
                        username = _make_unique_username(f"candidate{get_random_string(4).lower()}")
                    if UserModel.objects.filter(username=username).exists():
                        raise ValueError(f"Username '{username}' already exists")

                    password = generate_password()
                    user = UserModel.objects.create_user(
                        username=username,
                        email=email,
                        first_name=row.get("first_name", ""),
                        last_name=row.get("last_name", ""),
                        phone=str(row.get("phone", "")),
                        role="candidate",
                        password=password,
                        organization=_zecdata_org(),
                        created_by=request.user,
                    )

                    if email:
                        try:
                            send_candidate_credentials(user, password)
                        except Exception as e:
                            logger.warning("Failed to send credentials for imported candidate", extra={
                                "user_id": request.user.id,
                                "candidate_id": user.id,
                                "candidate_email": email
                            })
                    success += 1
                except Exception as exc:
                    failed += 1
                    errors.append(f"Row {index + 2}: {exc}")

            logger.info("Candidate import completed", extra={
                "user_id": request.user.id,
                "imported": success,
                "failed": failed,
                "total_processed": success + failed
            })
            
            return Response(
                {
                    "imported": success,
                    "failed": failed,
                    "errors": errors[:25],
                }
            )
            
        except Exception as e:
            logger.exception("Error during candidate import", extra={"user_id": request.user.id})
            return Response(
                {"detail": "Failed to import candidates"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ResumePresignedURLView(APIView):
    """Generate a presigned download URL for a candidate's resume stored in S3."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk: int, *args, **kwargs):
        logger.info("Resume presigned URL requested", extra={
            "user_id": request.user.id,
            "candidate_id": pk
        })
        
        try:
            candidate = get_object_or_404(UserModel, pk=pk)

            # Authorization: only the candidate themselves, a super admin, or an
            # admin of the SAME organization may fetch a resume. Prevents IDOR
            # (User is not tenant-scoped, so we must check explicitly).
            u = request.user
            is_super = bool(u.is_superuser or getattr(u, "role", None) == "super_admin")
            is_owner = candidate.id == u.id
            same_org_admin = (
                getattr(u, "role", None) in {"org_admin", "manager"}
                and candidate.organization_id is not None
                and candidate.organization_id == getattr(u, "organization_id", None)
            )
            if not (is_super or is_owner or same_org_admin):
                logger.warning(
                    "Unauthorized resume access blocked",
                    extra={"user_id": u.id, "candidate_id": pk},
                )
                return Response({"detail": "Not authorized to access this resume."}, status=status.HTTP_403_FORBIDDEN)

            if not candidate.resume_s3_url:
                logger.warning("No resume found for candidate", extra={
                    "user_id": request.user.id,
                    "candidate_id": pk,
                    "candidate_email": candidate.email
                })
                return Response({"detail": "No resume uploaded."}, status=status.HTTP_404_NOT_FOUND)

        
            import re
            from urllib.parse import urlparse
            parsed = urlparse(candidate.resume_s3_url)
            # Extract bucket name and region from URL: https://bucket.s3.region.amazonaws.com/key
            hostname = parsed.hostname or ""
            match = re.match(r"^(.+)\.s3[.-]([a-z0-9-]+)\.amazonaws\.com$", hostname)
            if match:
                bucket_name = match.group(1)
                region_name = match.group(2)
            else:
                bucket_name = getattr(settings, "AWS_STORAGE_BUCKET_NAME", None)
                region_name = getattr(settings, "AWS_S3_REGION_NAME", "ap-south-1")
            s3_key = parsed.path.lstrip("/")

            s3_client = boto3.client(
                "s3",
                aws_access_key_id=getattr(settings, "AWS_ACCESS_KEY_ID", None),
                aws_secret_access_key=getattr(settings, "AWS_SECRET_ACCESS_KEY", None),
                region_name=region_name,
            )

            presigned_url = s3_client.generate_presigned_url(
                "get_object",
                Params={
                    "Bucket": bucket_name,
                    "Key": s3_key,
                },
                ExpiresIn=300,  # 5 minutes
            )

            logger.info("Resume presigned URL generated successfully", extra={
                "user_id": request.user.id,
                "candidate_id": pk,
                "candidate_email": candidate.email
            })
            
            return Response({"url": presigned_url})
        except Exception as e:
            logger.exception("Error generating resume presigned URL", extra={
                "user_id": request.user.id,
                "candidate_id": pk
            })
            return Response(
                {"detail": f"Failed to generate download URL: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class CandidateDetailView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, pk: int, *args, **kwargs):
        logger.info("Candidate detail requested", extra={
            "user_id": request.user.id,
            "candidate_id": pk
        })
        
        try:
            from django.db.models import Case, When, IntegerField, Value
            from learning.models import UserTechnologyProgress, Assignment, Technology

            candidate = get_object_or_404(UserModel, pk=pk, role="candidate")
            
            # --- FIX: Get ALL technologies assigned to this candidate ---
            # Method 1: From Assignment table (most reliable)
            assigned_assignments = Assignment.objects.filter(
                user=candidate
            ).select_related('technology')
            
            # Method 2: From UserTechnologyProgress
            learning_progress = TenantQuerySet(UserTechnologyProgress, using='default').filter(
                user=candidate
            ).select_related('technology').annotate(
                status_order=Case(
                    When(progress__gt=0, progress__lt=100, then=Value(0)),
                    When(progress=0, then=Value(1)),
                    When(progress__gte=100, then=Value(2)),
                    default=Value(3),
                    output_field=IntegerField(),
                )
            ).order_by('status_order', '-progress')

            # --- FIX: Create a combined list of ALL assigned technologies ---
            learning_assignments = []
            seen_techs = set()
            
            # First, process all assigned assignments directly
            for assignment in assigned_assignments:
                tech = assignment.technology
                seen_techs.add(tech.id)
                
                # Find progress if exists
                progress = UserTechnologyProgress.objects.filter(
                    user=candidate,
                    technology=tech
                ).first()
                
                if progress:
                    if progress.progress >= 100:
                        status_label = "Completed"
                    elif progress.progress > 0:
                        status_label = "In Progress"
                    else:
                        status_label = "Not Started"
                    
                    learning_assignments.append({
                        "technology_id": str(tech.id),
                        "technology_name": tech.name,
                        "progress": progress.progress,
                        "completed": progress.completed,
                        "total": progress.total,
                        "status": status_label,
                        "assigned_at": assignment.assigned_at if assignment else None,
                        "due_at": assignment.due_at if assignment else None,
                        "last_active": progress.updated_at if hasattr(progress, 'updated_at') else None,
                        "notes": getattr(assignment, 'notes', 'NA'),
                        "assignment_id": assignment.id if assignment else None,
                    })
                else:
                    # Assignment exists but no progress - treat as "Not Started"
                    learning_assignments.append({
                        "technology_id": str(tech.id),
                        "technology_name": tech.name,
                        "progress": 0,
                        "completed": 0,
                        "total": 0,
                        "status": "Assigned",
                        "assigned_at": assignment.assigned_at if assignment else None,
                        "due_at": assignment.due_at if assignment else None,
                        "last_active": None,
                        "notes": getattr(assignment, 'notes', 'NA'),
                        "assignment_id": assignment.id if assignment else None,
                    })
            
            # Then, add any progress records without assignments (shouldn't happen normally)
            for progress in learning_progress:
                tech = progress.technology
                if tech.id not in seen_techs:
                    seen_techs.add(tech.id)
                    
                    if progress.progress >= 100:
                        status_label = "Completed"
                    elif progress.progress > 0:
                        status_label = "In Progress"
                    else:
                        status_label = "Not Started"
                    
                    learning_assignments.append({
                        "technology_id": str(tech.id),
                        "technology_name": tech.name,
                        "progress": progress.progress,
                        "completed": progress.completed,
                        "total": progress.total,
                        "status": status_label,
                        "assigned_at": None,
                        "due_at": None,
                        "last_active": progress.updated_at if hasattr(progress, 'updated_at') else None,
                        "notes": 'NA',
                        "assignment_id": None,
                    })

            # --- FIX: Ensure unlocked technologies are properly populated ---
            unlocked_tech_details = []
            if getattr(candidate, 'is_individual', False):
                # Get assigned technology IDs from Assignment table
                assigned_ids = TenantQuerySet(
                    Assignment, using='default'
                ).filter(user=candidate).values_list('technology_id', flat=True).distinct()
                
                techs = Technology.objects.all_for_super_admin().filter(
                    id__in=assigned_ids
                ).values('id', 'name')
                unlocked_tech_details = [
                    {"id": str(t['id']), "name": t['name']} for t in techs
                ]
            else:
                # For regular candidates, get all assigned technologies
                techs = Technology.objects.filter(
                    id__in=Assignment.objects.filter(user=candidate).values_list('technology_id', flat=True)
                ).values('id', 'name')
                unlocked_tech_details = [
                    {"id": str(t['id']), "name": t['name']} for t in techs
                ]

            logger.info(f"Found {len(learning_assignments)} learning assignments for candidate {pk}")
            logger.info(f"Unlocked technologies: {unlocked_tech_details}")

            return Response({
                "candidate": CandidateSerializer(candidate).data,
                "assessments": CandidateAssessmentSerializer(
                    CandidateAssessment.objects.filter(candidate=candidate), many=True
                ).data,
                "ai_assessments": [
                    {
                        "id": ca.id,
                        "ai_assessment": AIAssessmentListSerializer(ca.ai_assessment).data,
                        "status": ca.status,
                        "assigned_date": ca.assigned_date,
                        "end_time": ca.end_time,
                        "overall_score": ca.overall_score,
                    }
                    for ca in CandidateAIAssessment.objects.filter(candidate=candidate)
                ],
                "learning_assignments": learning_assignments,  # This is what populates the "Courses" section
                "unlocked_technologies": unlocked_tech_details,
            })
            
        except Exception as e:
            logger.exception("Error retrieving candidate detail", extra={
                "user_id": request.user.id,
                "candidate_id": pk
            })
            return Response(
                {"detail": "Failed to retrieve candidate details"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    permission_classes = [AdminPermission]

    def get(self, request, pk: int, *args, **kwargs):
        logger.info("Candidate detail requested", extra={
            "user_id": request.user.id,
            "candidate_id": pk
        })
        
        try:
            from django.db.models import Case, When, IntegerField, Value
            from learning.models import UserTechnologyProgress, Assignment

            candidate = get_object_or_404(UserModel, pk=pk, role="candidate")
            assigned_assessments = CandidateAssessment.objects.filter(candidate=candidate)

            # Get AI assessments
            from AI_assessment.models import CandidateAIAssessment
            ai_assessments = CandidateAIAssessment.objects.filter(
                candidate=candidate
            ).select_related('ai_assessment').order_by('-assigned_date')

            # Get learning assignments sorted by status priority
                        # Get learning assignments sorted by status priority
            # Get learning assignments sorted by status priority
            from organization.managers import TenantQuerySet
            learning_progress = TenantQuerySet(UserTechnologyProgress, using='default').filter(
                user=candidate
            ).select_related('technology').annotate(
                status_order=Case(
                    When(progress__gt=0, progress__lt=100, then=Value(0)),
                    When(progress=0, then=Value(1)),
                    When(progress__gte=100, then=Value(2)),
                    default=Value(3),
                    output_field=IntegerField(),
                )
            ).order_by('status_order', '-progress')

            learning_assignments = []
            for progress in learning_progress:
                # 🔥 FIXED: Specific technology ke liye assignment find karo
                tech_assignment = Assignment.objects.filter(
                    user=candidate, technology=progress.technology
                ).first()

                if progress.progress >= 100:
                    status_label = "Completed"
                elif progress.progress > 0:
                    status_label = "In Progress"
                else:
                    status_label = "Not Started"

                learning_assignments.append({
                    "technology_id": str(progress.technology.id),
                    "technology_name": progress.technology.name,
                    "progress": progress.progress,
                    "completed": progress.completed,
                    "total": progress.total,
                    "status": status_label,
                    "assigned_at": tech_assignment.assigned_at if tech_assignment else None,
                    "due_at": tech_assignment.due_at if tech_assignment else None,
                    "last_active": progress.updated_at if hasattr(progress, 'updated_at') else None,
                    "notes": getattr(tech_assignment, 'notes', 'NA'),
                    "assignment_id": tech_assignment.id if tech_assignment else None,
                })

            logger.info("Candidate detail retrieved successfully", extra={
                "user_id": request.user.id,
                "candidate_id": pk,
                "candidate_email": candidate.email,
                "regular_assessments_count": assigned_assessments.count(),
                "ai_assessments_count": ai_assessments.count(),
                "learning_assignments_count": len(learning_assignments)
            })
            # Unlocked technologies with names for individual candidates
            unlocked_tech_details = []
            if getattr(candidate, 'is_individual', False):
                from learning.models import Technology, Assignment
                # Get assigned technology IDs from Assignment table
                assigned_ids = TenantQuerySet(
                Assignment, using='default'
            ).filter(user=candidate).values_list('technology_id', flat=True).distinct()
                
                techs = Technology.objects.all_for_super_admin().filter(
                    id__in=assigned_ids
                ).values('id', 'name')
                unlocked_tech_details = [
                    {"id": str(t['id']), "name": t['name']} for t in techs
                ]
            
            return Response({
                "candidate": CandidateSerializer(candidate).data,
                "assessments": CandidateAssessmentSerializer(
                    assigned_assessments, many=True
                ).data,
                "ai_assessments": [
                    {
                        "id": ca.id,
                        "ai_assessment": AIAssessmentListSerializer(ca.ai_assessment).data,
                        "status": ca.status,
                        "assigned_date": ca.assigned_date,
                        "end_time": ca.end_time,
                        "overall_score": ca.overall_score,
                    }
                    for ca in ai_assessments
                ],
                "learning_assignments": learning_assignments,
                "unlocked_technologies": unlocked_tech_details,
            })
            
        except Exception as e:
            logger.exception("Error retrieving candidate detail", extra={
                "user_id": request.user.id,
                "candidate_id": pk
            })
            return Response(
                {"detail": "Failed to retrieve candidate details"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CandidateDeleteView(APIView):
    permission_classes = [AdminPermission]

    def delete(self, request, pk: int, *args, **kwargs):
        logger.info("Candidate deletion attempt", extra={
            "user_id": request.user.id,
            "candidate_id": pk
        })
        
        try:
            candidate = get_object_or_404(UserModel, pk=pk, role="candidate")
            candidate_email = candidate.email
            candidate.delete()
            
            logger.info("Candidate deleted successfully", extra={
                "user_id": request.user.id,
                "candidate_id": pk,
                "candidate_email": candidate_email
            })
            
            return Response(status=status.HTTP_204_NO_CONTENT)
            
        except Exception as e:
            logger.exception("Error during candidate deletion", extra={
                "user_id": request.user.id,
                "candidate_id": pk
            })
            return Response(
                {"detail": "Failed to delete candidate"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CandidateBulkDeleteView(APIView):
    permission_classes = [AdminPermission]

    def post(self, request, *args, **kwargs):
        logger.info("Candidate bulk delete attempt", extra={"user_id": request.user.id})
        
        try:
            serializer = CandidateBulkDeleteSerializer(data=request.data)
            if not serializer.is_valid():
                logger.warning("Candidate bulk delete validation failed", extra={
                    "user_id": request.user.id,
                    "errors": serializer.errors
                })
                serializer.is_valid(raise_exception=True)
                
            candidate_ids = serializer.validated_data["candidate_ids"]
            logger.info("Bulk delete parameters validated", extra={
                "user_id": request.user.id,
                "candidate_count": len(candidate_ids)
            })
            
            deleted, _ = UserModel.objects.filter(
                id__in=candidate_ids, 
                role="candidate",
                organization=getattr(request.user, 'organization', None)
            ).delete()
            
            logger.info("Candidate bulk delete completed", extra={
                "user_id": request.user.id,
                "requested_count": len(candidate_ids),
                "deleted_count": deleted
            })
            
            return Response({"deleted": deleted})
            
        except Exception as e:
            logger.exception("Error during candidate bulk delete", extra={"user_id": request.user.id})
            return Response(
                {"detail": "Failed to bulk delete candidates"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class AssessmentBulkDeleteView(APIView):
    """Bulk delete regular assessments (admin only)."""
    permission_classes = [AdminPermission]

    def post(self, request, *args, **kwargs):
        logger.info("Assessment bulk delete attempt", extra={"user_id": request.user.id})
        
        try:
            serializer = AssessmentBulkDeleteSerializer(data=request.data)
            if not serializer.is_valid():
                logger.warning("Assessment bulk delete validation failed", extra={
                    "user_id": request.user.id,
                    "errors": serializer.errors
                })
                serializer.is_valid(raise_exception=True)
                
            assessment_ids = serializer.validated_data["assessment_ids"]
            logger.info("Assessment bulk delete parameters validated", extra={
                "user_id": request.user.id,
                "assessment_count": len(assessment_ids)
            })
            
            delete_qs = Assessment.objects.filter(
                id__in=assessment_ids,
                organization=getattr(request.user, 'organization', None)
            )
            # A manager may only delete assessments they created; org_admin /
            # super_admin may delete any within their organization scope.
            if is_manager(request.user):
                delete_qs = delete_qs.filter(created_by=request.user)
            deleted, _ = delete_qs.delete()
            
            logger.info("Assessment bulk delete completed", extra={
                "user_id": request.user.id,
                "requested_count": len(assessment_ids),
                "deleted_count": deleted
            })
            
            return Response({"deleted": deleted}, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.exception("Error during assessment bulk delete", extra={"user_id": request.user.id})
            return Response(
                {"detail": "Failed to bulk delete assessments"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CandidateQuickAssignView(APIView):
    permission_classes = [AdminPermission]

    def post(self, request, candidate_id: int, *args, **kwargs):
        logger.info("Candidate quick assign attempt", extra={
            "user_id": request.user.id,
            "candidate_id": candidate_id
        })
        
        try:
            serializer = CandidateQuickAssignSerializer(data=request.data)
            if not serializer.is_valid():
                logger.warning("Candidate quick assign validation failed", extra={
                    "user_id": request.user.id,
                    "candidate_id": candidate_id,
                    "errors": serializer.errors
                })
                serializer.is_valid(raise_exception=True)
                
            candidate = get_object_or_404(UserModel, pk=candidate_id, role="candidate")
            selection = serializer.validated_data.get("selection")
            assessment_token = serializer.validated_data.get("assessment_id")
            ai_token = serializer.validated_data.get("ai_assessment_id")

            regular_id = None
            ai_id = None

            token = selection or assessment_token or ai_token
            if token and isinstance(token, str):
                if token.startswith("regular_"):
                    regular_id = token.replace("regular_", "")
                elif token.startswith("ai_"):
                    ai_id = token.replace("ai_", "")

            if assessment_token and not regular_id:
                regular_id = assessment_token
            if ai_token and not ai_id:
                ai_id = ai_token

            if regular_id:
                assessment = get_object_or_404(Assessment, pk=int(regular_id))
                existing = CandidateAssessment.objects.filter(
                    candidate=candidate, assessment=assessment
                ).first()
                if existing:
                    logger.warning("Assessment already assigned to candidate", extra={
                        "user_id": request.user.id,
                        "candidate_id": candidate_id,
                        "assessment_id": assessment.id
                    })
                    return Response(
                        {"detail": "Assessment already assigned to this candidate."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                candidate_assessment = CandidateAssessment.objects.create(
                    candidate=candidate, assessment=assessment, assigned_by=request.user
                )
                send_assignment_notification(candidate_assessment)
                
                logger.info("Regular assessment assigned to candidate", extra={
                    "user_id": request.user.id,
                    "candidate_id": candidate_id,
                    "assessment_id": assessment.id,
                    "assessment_title": assessment.title
                })
                
                return Response(
                    {"detail": "Assessment assigned successfully."},
                    status=status.HTTP_201_CREATED,
                )

            if ai_id:
                try:
                    ai_assessment = get_object_or_404(AIAssessment, pk=int(ai_id))
                    existing = CandidateAIAssessment.objects.filter(
                        candidate=candidate, ai_assessment=ai_assessment
                    ).first()
                    if existing:
                        logger.warning("AI assessment already assigned to candidate", extra={
                            "user_id": request.user.id,
                            "candidate_id": candidate_id,
                            "ai_assessment_id": ai_assessment.id
                        })
                        return Response(
                            {"detail": "AI assessment already assigned to this candidate."},
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                    candidate_ai = CandidateAIAssessment.objects.create(
                        candidate=candidate, ai_assessment=ai_assessment, assigned_by=request.user
                    )
                    send_ai_assessment_notification(candidate_ai)
                    logger.info("AI assessment assigned to candidate", extra={
                        "user_id": request.user.id,
                        "candidate_id": candidate_id,
                        "ai_assessment_id": ai_assessment.id,
                        "ai_assessment_title": ai_assessment.title
                    })
                    
                    return Response(
                        {"detail": "AI assessment assigned successfully."},
                        status=status.HTTP_201_CREATED,
                    )
                except Exception as e:
                    logger.exception("Error assigning AI assessment", extra={
                        "user_id": request.user.id,
                        "candidate_id": candidate_id,
                        "ai_assessment_id": ai_id
                    })
                    return Response(
                        {"detail": "Failed to assign AI assessment."},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    )

            logger.warning("No valid assessment provided for assignment", extra={
                "user_id": request.user.id,
                "candidate_id": candidate_id
            })
            return Response(
                {"detail": "No valid assessment provided."},
                status=status.HTTP_400_BAD_REQUEST,
            )
            
        except Exception as e:
            logger.exception("Error during candidate quick assign", extra={
                "user_id": request.user.id,
                "candidate_id": candidate_id
            })
            return Response(
                {"detail": "Failed to assign assessment"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class QuestionListView(APIView, StandardResultsSetPagination):
    permission_classes = [AdminPermission]

    def get(self, request, *args, **kwargs):
        logger.info("Question list requested", extra={"user_id": request.user.id})
        
        try:
            queryset = (
                Question.objects
                .select_related("category")
                .all()
                .order_by("-created_at")
            )

            # ------------------ QUERY PARAMS ------------------
            search = request.query_params.get("search", "").strip()
            category = request.query_params.get("category")
            difficulty = request.query_params.get("difficulty")
            question_type = request.query_params.get("question_type")

            logger.info("Question filters applied", extra={
                "user_id": request.user.id,
                "search": search,
                "category": category,
                "difficulty": difficulty,
                "question_type": question_type
            })

            # ------------------ SEARCH ------------------
            if search:
                category_match = Category.objects.filter(name__iexact=search).first()
                if category_match:
                    queryset = queryset.filter(category=category_match)
                else:
                    queryset = queryset.filter(
                        Q(title__icontains=search) |
                        Q(description__icontains=search) |
                        Q(tags__icontains=search) |
                        Q(category__name__icontains=search)
                    )

            # ------------------ FILTERS ------------------
            if category:
                if str(category).isdigit():
                    queryset = queryset.filter(category_id=category)
                else:
                    queryset = queryset.filter(category__name__iexact=category)

            if difficulty:
                queryset = queryset.filter(difficulty=difficulty)

            if question_type:
                queryset = queryset.filter(question_type=question_type)

            # ------------------ PAGINATION ------------------
            page = self.paginate_queryset(queryset, request, view=self)
            serializer = QuestionSerializer(page, many=True)

            categories = list(Category.objects.values("id", "name"))

            logger.info("Question list retrieved successfully", extra={
                "user_id": request.user.id,
                "total_questions": queryset.count(),
                "page_size": len(page) if page else 0
            })
            
            return self.get_paginated_response(
                {
                    "questions": serializer.data,
                    "categories": categories,
                }
            )
            
        except Exception as e:
            logger.exception("Error retrieving question list", extra={"user_id": request.user.id})
            return Response(
                {"detail": "Failed to retrieve questions"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class QuestionDetailView(APIView):
    """
    Get detailed information about a specific question.
    Includes all question data, testcases, and SQL details if applicable.
    """
    permission_classes = [AdminPermission]

    def get(self, request, question_id: int, *args, **kwargs):
        logger.info("Question detail retrieval attempt", extra={"user_id": request.user.id})
        question = get_object_or_404(Question, pk=question_id)
        serializer = QuestionSerializer(question)
        return Response(serializer.data)


class QuestionCreateView(APIView):
    permission_classes = [AdminPermission]

    def post(self, request, *args, **kwargs):
        logger.info("Question creation attempt", extra={"user_id": request.user.id})
        
        try:
            serializer = QuestionCreateSerializer(data=request.data)
            if not serializer.is_valid():
                logger.warning("Question creation validation failed", extra={
                    "user_id": request.user.id,
                    "errors": serializer.errors
                })
                serializer.is_valid(raise_exception=True)

            # Save question + SQL metadata (inside serializer.create)
            question = serializer.save(
                created_by=request.user,
                organization=getattr(request.user, 'organization', None)
            )
            
            logger.info("Question created successfully", extra={
                "user_id": request.user.id,
                "question_id": question.id,
                "question_title": question.title[:50] + "..." if len(question.title) > 50 else question.title,
                "question_type": question.question_type
            })
            
            return Response(QuestionSerializer(question).data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.exception("Error during question creation", extra={"user_id": request.user.id})
            return Response(
                {"detail": "Failed to create question"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class QuestionEditView(APIView):
    permission_classes = [AdminPermission]

    def put(self, request, question_id: int, *args, **kwargs):
        logger.info("Question update attempt", extra={
            "user_id": request.user.id,
            "question_id": question_id
        })
        
        try:
            question = get_object_or_404(Question, pk=question_id)
            old_title = question.title
            
            serializer = QuestionCreateSerializer(instance=question, data=request.data)
            if not serializer.is_valid():
                logger.warning("Question update validation failed", extra={
                    "user_id": request.user.id,
                    "question_id": question_id,
                    "errors": serializer.errors
                })
                serializer.is_valid(raise_exception=True)
                
            serializer.save()
            
            logger.info("Question updated successfully", extra={
                "user_id": request.user.id,
                "question_id": question_id,
                "old_title": old_title[:50] + "..." if len(old_title) > 50 else old_title,
                "new_title": question.title[:50] + "..." if len(question.title) > 50 else question.title
            })
            
            return Response(QuestionSerializer(question).data)
            
        except Exception as e:
            logger.exception("Error during question update", extra={
                "user_id": request.user.id,
                "question_id": question_id
            })
            return Response(
                {"detail": "Failed to update question"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def patch(self, request, question_id: int, *args, **kwargs):
        logger.info("Question partial update attempt", extra={
            "user_id": request.user.id,
            "question_id": question_id
        })
        
        try:
            question = get_object_or_404(Question, pk=question_id)
            old_title = question.title
            
            serializer = QuestionCreateSerializer(
                instance=question, data=request.data, partial=True
            )
            if not serializer.is_valid():
                logger.warning("Question partial update validation failed", extra={
                    "user_id": request.user.id,
                    "question_id": question_id,
                    "errors": serializer.errors
                })
                serializer.is_valid(raise_exception=True)
                
            serializer.save()
            
            logger.info("Question partially updated successfully", extra={
                "user_id": request.user.id,
                "question_id": question_id,
                "old_title": old_title[:50] + "..." if len(old_title) > 50 else old_title,
                "new_title": question.title[:50] + "..." if len(question.title) > 50 else question.title
            })
            
            return Response(QuestionSerializer(question).data)
            
        except Exception as e:
            logger.exception("Error during question partial update", extra={
                "user_id": request.user.id,
                "question_id": question_id
            })
            return Response(
                {"detail": "Failed to update question"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class QuestionDeleteView(APIView):
    permission_classes = [AdminPermission]

    def delete(self, request, question_id: int, *args, **kwargs):
        logger.info("Question deletion attempt", extra={
            "user_id": request.user.id,
            "question_id": question_id
        })
        
        try:
            question = get_object_or_404(Question, pk=question_id)
            question_title = question.title
            question.delete()
            
            logger.info("Question deleted successfully", extra={
                "user_id": request.user.id,
                "question_id": question_id,
                "question_title": question_title[:50] + "..." if len(question_title) > 50 else question_title
            })
            
            return Response(status=status.HTTP_204_NO_CONTENT)
            
        except Exception as e:
            logger.exception("Error during question deletion", extra={
                "user_id": request.user.id,
                "question_id": question_id
            })
            return Response(
                {"detail": "Failed to delete question"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class QuestionBulkDeleteView(APIView):
    permission_classes = [AdminPermission]

    def post(self, request, *args, **kwargs):
        logger.info("Question bulk delete attempt", extra={"user_id": request.user.id})
        
        try:
            serializer = QuestionBulkDeleteSerializer(data=request.data)
            if not serializer.is_valid():
                logger.warning("Question bulk delete validation failed", extra={
                    "user_id": request.user.id,
                    "errors": serializer.errors
                })
                serializer.is_valid(raise_exception=True)
                
            question_ids = serializer.validated_data["question_ids"]
            logger.info("Question bulk delete parameters validated", extra={
                "user_id": request.user.id,
                "question_count": len(question_ids)
            })
            
            deleted, _ = Question.objects.filter(id__in=question_ids).delete()
            
            logger.info("Question bulk delete completed", extra={
                "user_id": request.user.id,
                "requested_count": len(question_ids),
                "deleted_count": deleted
            })
            
            return Response({"deleted": deleted})
            
        except Exception as e:
            logger.exception("Error during question bulk delete", extra={"user_id": request.user.id})
            return Response(
                {"detail": "Failed to bulk delete questions"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class QuestionImportView(APIView):
    authentication_classes = [JWTAuthentication, TokenAuthentication, SessionAuthentication]
    # Bulk question upload is Super-Admin only.
    permission_classes = [SuperAdminPermission]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request, *args, **kwargs):
        logger.info("Question import started", extra={"user": request.user.id})

        serializer = QuestionImportSerializer(data=request.data)

        if not serializer.is_valid():
            logger.warning(
                "Question import validation failed",
                extra={"errors": serializer.errors}
            )
            serializer.is_valid(raise_exception=True)

        upload = serializer.validated_data["file"]

        try:
            rows = _load_rows_from_upload(upload)
            logger.info("File loaded successfully", extra={"rows": len(rows)})
        except Exception as exc:
            logger.error("File parsing failed", extra={"error": str(exc)})
            return Response(
                {"detail": f"Unable to parse file: {exc}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created, failed, errors = 0, 0, []

        def _get_value(row_data, key, default=""):
            value = row_data.get(key, default)
            if _is_empty_value(value):
                return default
            return value

        def _get_string(row_data, key, default=""):
            value = _get_value(row_data, key, default)
            if isinstance(value, str):
                return value.strip()
            return value

        def _parse_float(value, default=0.0):
            if _is_empty_value(value):
                return default
            try:
                return float(value)
            except (TypeError, ValueError):
                return default

        def _parse_bool(value, default=False):
            if isinstance(value, bool):
                return value
            if _is_empty_value(value):
                return default
            return str(value).strip().lower() in {"1", "true", "yes", "y"}

        def _parse_json_list(value, field_name):
            if _is_empty_value(value):
                return []
            if isinstance(value, list):
                return value
            if isinstance(value, str):
                value = value.strip()
                if not value:
                    return []
                try:
                    data = json.loads(value)
                except json.JSONDecodeError as exc:
                    raise ValueError(f"{field_name} must contain valid JSON. {exc}")
                if isinstance(data, list):
                    return data
            raise ValueError(f"{field_name} must be a JSON array of objects.")

        for index, row in enumerate(rows):
            row_number = index + 2

            try:
                with transaction.atomic():
                    logger.info("Processing row", extra={"row": row_number})

                    qtype = _get_string(row, "question_type", "mcq_single") or "mcq_single"

                    category_name = _get_string(row, "category")
                    if not category_name:
                        raise ValueError("Category is required.")

                    category, _ = Category.objects.get_or_create(name=category_name)

                    marks_value = _parse_float(_get_value(row, "marks", 1), 1)

                    question = Question.objects.create(
                        title=_get_string(row, "title", ""),
                        question_type=qtype,
                        category=category,
                        difficulty=_get_string(row, "difficulty", "medium") or "medium",
                        marks=int(marks_value) if marks_value else 1,
                        description=_get_string(row, "description", ""),
                        sample_input=_get_string(row, "sample_input", ""),
                        sample_output=_get_string(row, "sample_output", ""),
                        option1=_get_string(row, "option1", ""),
                        option2=_get_string(row, "option2", ""),
                        option3=_get_string(row, "option3", ""),
                        option4=_get_string(row, "option4", ""),
                        option5=_get_string(row, "option5", ""),
                        correct_answer=_get_string(row, "correct_answer", ""),
                        tags=_get_string(row, "tags", ""),
                        created_by=request.user,
                        organization=getattr(request.user, 'organization', None),
                    )

                    logger.info(
                        "Question created",
                        extra={"question_id": question.id, "row": row_number}
                    )

                    if qtype == "coding":
                        coding_testcases = _parse_json_list(_get_value(row, "coding_testcases", ""), "coding_testcases")
                        if not coding_testcases:
                            raise ValueError("coding_testcases must contain at least one test case for coding questions.")

                        for tc in coding_testcases:
                            if not isinstance(tc, dict):
                                raise ValueError("Each coding test case must be a JSON object.")
                            TestCase.objects.create(
                                question=question,
                                input_data=tc.get("input_data", ""),
                                expected_output=tc.get("expected_output", ""),
                                points=_parse_float(tc.get("points", 1.0), 1.0),
                                is_hidden=_parse_bool(tc.get("is_hidden", True), True),
                            )

                        logger.info("Coding testcases added", extra={"row": row_number})

                    if qtype == "sql":
                        dataset_name = _get_string(row, "sql_dataset", "")
                        if not dataset_name:
                            raise ValueError("sql_dataset is required for SQL questions.")

                        try:
                            dataset = SQLDataset.objects.get(name=dataset_name)
                        except SQLDataset.DoesNotExist:
                            raise ValueError(f"SQL dataset '{dataset_name}' not found. "
                                "Create this dataset first or use an existing SQL dataset name."
                            )

                        reference_solution = _get_string(row, "sql_reference_solution", "")
                        if not reference_solution:
                            raise ValueError("sql_reference_solution is required for SQL questions.")

                        SQLQuestion.objects.create(
                            question=question,
                            dataset=dataset,
                            reference_solution=reference_solution,
                            strict_column_order=_parse_bool(_get_value(row, "sql_strict_column_order", False), False),
                            float_tolerance=_parse_float(_get_value(row, "sql_float_tolerance", 0.0), 0.0),
                            max_rows=int(_parse_float(_get_value(row, "sql_max_rows", 5000), 5000)),
                        )

                        logger.info("SQL question configured", extra={"row": row_number})
                        sql_testcases = _parse_json_list(_get_value(row, "sql_testcases", ""), "sql_testcases")
                        for tc in sql_testcases:
                            if not isinstance(tc, dict):
                                raise ValueError("Each SQL test case must be a JSON object.")
                            SQLTestCase.objects.create(
                                question=question,
                                setup_sql=tc.get("setup_sql", ""),
                                points=_parse_float(tc.get("points", 1.0), 1.0),
                                is_hidden=_parse_bool(tc.get("is_hidden", True), True),
                            )

                        logger.info("SQL testcases added", extra={"row": row_number})
                created += 1
            except Exception as exc:
                failed += 1
                error_msg = f"Row {row_number}: {exc}"
                errors.append(error_msg)

                logger.error(
                    "Row processing failed",
                    extra={"row": row_number, "error": str(exc)}
                )

        # logger.info(
        #     "Question import completed",
        #     extra={"created": created, "failed": failed}
        #  )
        logger.info(
            "Question import completed - Created: %d, Failed: %d",
            created,
            failed
        )
        

        return Response({
            "created": created,
            "failed": failed,
            "errors": errors[:25],
        })

class AssessmentListView(APIView, StandardResultsSetPagination):
    permission_classes = [AdminPermission]

    def get(self, request, *args, **kwargs):
        logger.info(
            "Assessment list API called",
            extra={"user_id": request.user.id}
        )

        queryset = (
            Assessment.objects.all()
            .select_related("created_by")
            .prefetch_related("categories", "questions")
            .order_by("-created_at")
        )

        # Visibility:
        #   super_admin -> every assessment
        #   org_admin   -> all assessments across their organization
        #   manager     -> only the assessments they created
        role = getattr(request.user, "role", None)
        if role == "super_admin" or request.user.is_superuser:
            pass
        elif role == "org_admin":
            org_id = getattr(request.user, "organization_id", None)
            queryset = queryset.filter(organization_id=org_id) if org_id else queryset.none()
        elif role == "manager":
            queryset = queryset.filter(created_by=request.user)
        else:
            queryset = queryset.none()

        logger.info(
            "Initial assessment queryset loaded",
            extra={"total_assessments": queryset.count()}
        )

        now = timezone.now()

        # --- Status filter (active/upcoming/completed) ---
        status_filter = request.query_params.get("status")

        if status_filter == "active":
            queryset = queryset.filter(start_date__lte=now, end_date__gte=now)
            logger.info("Applied active status filter")

        elif status_filter == "upcoming":
            queryset = queryset.filter(start_date__gt=now)
            logger.info("Applied upcoming status filter")

        elif status_filter == "completed":
            queryset = queryset.filter(end_date__lt=now)
            logger.info("Applied completed status filter")

        # --- Search filter ---
        query = request.query_params.get("search", "").strip()

        if query:
            queryset = queryset.filter(title__icontains=query)

            logger.info(
                "Applied search filter",
                extra={"search_query": query}
            )

        # --- Date filters ---
        start_date_after = request.query_params.get("start_date_after", "").strip()
        start_date_before = request.query_params.get("start_date_before", "").strip()

        if start_date_after:
            try:
                from_date = datetime.strptime(start_date_after, "%Y-%m-%d")
                from_date = timezone.make_aware(from_date) if timezone.is_naive(from_date) else from_date
                queryset = queryset.filter(start_date__gte=from_date)

                logger.info(
                    "Applied start_date_after filter",
                    extra={"start_date_after": str(from_date)}
                )

            except Exception as e:
                logger.error(
                    "Error parsing start_date_after",
                    extra={
                        "value": start_date_after,
                        "error": str(e),
                    }
                )

        if start_date_before:
            try:
                to_date = datetime.strptime(start_date_before, "%Y-%m-%d")
                # Set to end of day for inclusive filtering
                to_date = datetime.combine(to_date.date(), datetime.max.time())
                to_date = timezone.make_aware(to_date) if timezone.is_naive(to_date) else to_date
                queryset = queryset.filter(start_date__lte=to_date)

                logger.info(
                    "Applied start_date_before filter",
                    extra={"start_date_before": str(to_date)}
                )

            except Exception as e:
                logger.error(
                    "Error parsing start_date_before",
                    extra={
                        "value": start_date_before,
                        "error": str(e),
                    }
                )

        logger.info(
            "Filtered assessments count",
            extra={"count": queryset.count()}
        )

        # --- Pagination ---
        page = self.paginate_queryset(queryset, request, view=self)

        if page is not None:
            logger.info("Pagination applied")

        instances = list(page) if page is not None else list(queryset)

        serializer = AssessmentSerializer(instances, many=True)

        # --- Candidate stats ---
        ids = [assessment.id for assessment in instances]

        logger.info(
            "Fetching candidate assessment stats",
            extra={"assessment_count": len(ids)}
        )

        stats_qs = (
            CandidateAssessment.objects.filter(assessment_id__in=ids)
            .values("assessment_id")
            .annotate(
                total=Count("id"),
                completed=Count("id", filter=Q(status="completed")),
                in_progress=Count("id", filter=Q(status="in_progress")),
                assigned=Count("id", filter=Q(status="assigned")),
            )
        )
        stats_map = {item["assessment_id"]: item for item in stats_qs}

        for obj, data in zip(instances, serializer.data):
            stats = stats_map.get(obj.id, {})
            data.update(
                {
                    "total_candidates": stats.get("total", 0),
                    "completed_count": stats.get("completed", 0),
                    "in_progress_count": stats.get("in_progress", 0),
                    "not_started_count": stats.get("assigned", 0),
                    "question_count": obj.questions.count(),
                    "category_names": [c.name for c in obj.categories.all()],
                }
            )

        logger.info(
            "Assessment list response prepared successfully",
            extra={"returned_count": len(serializer.data)}
        )

        # --- Response ---
        if page is not None:
            response = self.get_paginated_response(serializer.data)
            response.data["status"] = status_filter

            logger.info("Paginated response returned")

            return response

        logger.info("Non-paginated response returned")

        return Response({
            "status": status_filter,
            "results": serializer.data,
        })

class AssessmentCreateView(APIView):
    permission_classes = [AdminPermission]

    def post(self, request, *args, **kwargs):
        logger.info(
            "Assessment creation started",
            extra={"user_id": request.user.id}
        )

        serializer = AssessmentSerializer(
            data=request.data,
            context={"request": request},
        )

        if not serializer.is_valid():
            logger.warning(
                "Assessment creation validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        try:
            assessment = serializer.save(
                created_by=request.user,
                organization=getattr(request.user, 'organization', None)
            )

            # Attach selected questions chosen during the wizard. The create
            # serializer doesn't write the M2M-through field, so create the
            # AssessmentQuestion rows directly (same pattern as the update /
            # auto-fill paths). question_ids are tenant-scoped via Question's
            # manager, so only the org's own questions can be attached.
            question_ids = request.data.get("question_ids") or []
            attached = 0
            for qid in question_ids:
                q = Question.objects.filter(id=qid).first()
                if q:
                    _, created = AssessmentQuestion.objects.get_or_create(
                        assessment=assessment, question=q
                    )
                    attached += 1 if created else 0

            logger.info(
                "Assessment created successfully",
                extra={
                    "assessment_id": assessment.id,
                    "title": assessment.title,
                    "created_by": request.user.id,
                }
            )

            return Response(
                AssessmentSerializer(
                    assessment,
                    context={"request": request},
                ).data,
                status=status.HTTP_201_CREATED
            )

        except Exception as e:
            logger.exception(
                "Assessment creation failed",
                extra={"error": str(e)}
            )

            return Response(
                {"detail": "Failed to create assessment"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class AssessmentDetailView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, assessment_id: int, *args, **kwargs):

        logger.info(
            "Assessment detail API called",
            extra={
                "assessment_id": assessment_id,
                "user_id": request.user.id,
            }
        )

        try:
            assessment = get_object_or_404(Assessment, pk=assessment_id)

            if not can_access_assessment(request.user, assessment):
                return Response(
                    {"status": "error", "message": "You do not have access to this assessment."},
                    status=status.HTTP_403_FORBIDDEN,
                )

            logger.info(
                "Assessment fetched successfully",
                extra={
                    "assessment_id": assessment.id,
                    "title": assessment.title,
                }
            )

            questions = assessment.questions.all()

            logger.info(
                "Questions fetched for assessment",
                extra={
                    "assessment_id": assessment.id,
                    "question_count": questions.count(),
                }
            )

            # CandidateAssessment queryset
            candidate_qs = CandidateAssessment.objects.filter(
                assessment=assessment
            )

            stats = candidate_qs.aggregate(
                total_assigned=Count("id"),
                completed=Count("id", filter=Q(status="completed")),
                in_progress=Count("id", filter=Q(status="in_progress")),
                expired=Count("id", filter=Q(status="expired")),
            )

            logger.info(
                "Assessment stats calculated",
                extra={
                    "assessment_id": assessment.id,
                    "stats": stats,
                }
            )

            response_data = {
                "assessment": AssessmentSerializer(
                    assessment,
                    context={"stats": stats}
                ).data,
                "questions": QuestionSerializer(
                    questions,
                    many=True
                ).data,
            }

            logger.info(
                "Assessment detail response prepared successfully",
                extra={"assessment_id": assessment.id}
            )

            return Response(response_data)

        except Exception as e:
            logger.exception(
                "Failed to fetch assessment details",
                extra={
                    "assessment_id": assessment_id,
                    "error": str(e),
                }
            )

            return Response(
                {"detail": "Failed to fetch assessment details"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
     
class AssessmentAutofillQuestionsView(APIView):
    permission_classes = [AdminPermission]

    def post(self, request, assessment_id):

        logger.info(
            "Assessment autofill started",
            extra={
                "assessment_id": assessment_id,
                "user_id": request.user.id,
            }
        )

        assessment = get_object_or_404(Assessment, id=assessment_id)

        rules = request.data.get("rules", [])

        logger.info(
            "Autofill rules received",
            extra={"rules_count": len(rules)}
        )

        summary = {
            "added": 0,
            "skipped": 0,
            "errors": []
        }

        existing_ids = set(
            assessment.questions.values_list("id", flat=True)
        )

        for idx, rule in enumerate(rules, start=1):
            try:
                logger.info(
                    "Processing autofill rule",
                    extra={
                        "rule_number": idx,
                        "rule_data": rule,
                    }
                )

                category_id = rule.get("category_id")

                if not category_id:
                    raise ValueError("category_id missing")

                qs = Question.objects.filter(category_id=category_id)

                qtype = rule.get("question_type")

                if qtype and qtype != "any":
                    qs = qs.filter(question_type=qtype)

                difficulty = rule.get("difficulty")

                if difficulty and difficulty != "any":
                    qs = qs.filter(difficulty=difficulty)

                qs = qs.exclude(id__in=existing_ids)

                count = int(rule.get("count", 0))

                if count > 0:
                    qs = qs.order_by("?")[:count]
                else:
                    qs = qs.order_by("?")

                added_now = 0

                for q in qs:
                    aq, created = AssessmentQuestion.objects.get_or_create(
                        assessment=assessment,
                        question=q
                    )

                    if created:
                        summary["added"] += 1
                        added_now += 1
                        existing_ids.add(q.id)

                        logger.info(
                            "Question added to assessment",
                            extra={
                                "assessment_id": assessment.id,
                                "question_id": q.id,
                            }
                        )

                    else:
                        summary["skipped"] += 1

                if added_now == 0:
                    warning_msg = f"Rule {idx}: No available questions found"

                    summary["errors"].append(warning_msg)

                    logger.warning(
                        "No questions added for rule",
                        extra={"rule_number": idx}
                    )

            except Exception as e:
                error_msg = f"Rule {idx}: {str(e)}"

                summary["errors"].append(error_msg)

                logger.error(
                    "Error processing autofill rule",
                    extra={
                        "rule_number": idx,
                        "error": str(e),
                    }
                )

        logger.info(
            "Assessment autofill completed",
            extra={
                "assessment_id": assessment.id,
                "added": summary["added"],
                "skipped": summary["skipped"],
                "errors_count": len(summary["errors"]),
            }
        )

        return Response(summary, status=200)


class AssessmentUpdateView(APIView):
    permission_classes = [AdminPermission]

    def put(self, request, assessment_id: int, *args, **kwargs):

        logger.info(
            "Assessment full update started",
            extra={
                "assessment_id": assessment_id,
                "user_id": request.user.id,
            }
        )

        assessment = get_object_or_404(Assessment, pk=assessment_id)

        if not can_access_assessment(request.user, assessment):
            return Response(
                {"status": "error", "message": "You do not have access to this assessment."},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = AssessmentSerializer(
            instance=assessment,
            data=request.data,
            context={"request": request},
        )

        if not serializer.is_valid():
            logger.warning(
                "Assessment PUT validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        serializer.save()

        logger.info(
            "Assessment updated successfully",
            extra={
                "assessment_id": assessment.id,
                "title": assessment.title,
            }
        )

        return Response(
            AssessmentSerializer(
                assessment,
                context={"request": request},
            ).data
        )

    def patch(self, request, assessment_id: int, *args, **kwargs):

        logger.info(
            "Assessment partial update started",
            extra={
                "assessment_id": assessment_id,
                "user_id": request.user.id,
            }
        )

        assessment = get_object_or_404(Assessment, pk=assessment_id)

        serializer = AssessmentSerializer(
            instance=assessment,
            data=request.data,
            partial=True,
            context={"request": request},
        )

        if not serializer.is_valid():
            logger.warning(
                "Assessment PATCH validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        serializer.save()

        logger.info(
            "Assessment partially updated successfully",
            extra={
                "assessment_id": assessment.id,
                "title": assessment.title,
            }
        )

        return Response(
            AssessmentSerializer(
                assessment,
                context={"request": request},
            ).data
        )


class AssessmentCandidatesView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, assessment_id: int, *args, **kwargs):

        logger.info(
            "Assessment candidates API called",
            extra={
                "assessment_id": assessment_id,
                "user_id": request.user.id,
            }
        )

        assessment = get_object_or_404(Assessment, pk=assessment_id)

        logger.info(
            "Assessment fetched successfully",
            extra={
                "assessment_id": assessment.id,
                "title": assessment.title,
            }
        )

        assignments = CandidateAssessment.objects.filter(
            assessment=assessment
        )

        logger.info(
            "Candidate assignments fetched",
            extra={
                "assessment_id": assessment.id,
                "assignments_count": assignments.count(),
            }
        )

        response_data = {
            "assessment": AssessmentSerializer(assessment).data,
            "assignments": CandidateAssessmentSerializer(assignments, many=True).data,
        }

        logger.info(
            "Assessment candidates response prepared successfully",
            extra={"assessment_id": assessment.id}
        )

        return Response(response_data)

class AssessmentCandidatesStatusView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, assessment_id: int, *args, **kwargs):

        logger.info(
            "Assessment candidates status API called",
            extra={
                "assessment_id": assessment_id,
                "user_id": request.user.id,
            }
        )

        assessment = get_object_or_404(Assessment, pk=assessment_id)

        if not can_access_assessment(request.user, assessment):
            return Response(
                {"status": "error", "message": "You do not have access to this assessment."},
                status=status.HTTP_403_FORBIDDEN,
            )

        logger.info(
            "Assessment fetched successfully",
            extra={
                "assessment_id": assessment.id,
                "title": assessment.title,
            }
        )

        status_filter = request.query_params.get("status", "").strip().lower()

        logger.info(
            "Status filter received",
            extra={"status_filter": status_filter}
        )

        valid_statuses = {"assigned", "in_progress", "completed", "expired"}

        assignments = CandidateAssessment.objects.filter(
            assessment=assessment
        ).select_related("candidate", "assigned_by")

        logger.info(
            "Candidate assignments fetched",
            extra={
                "assessment_id": assessment.id,
                "total_assignments": assignments.count(),
            }
        )

        # Calculate summary counts
        summary = {
            # "total_assigned": assignments.filter(status='assigned').count(),
            "total_in_progress": assignments.filter(status='in_progress').count(),
            "total_completed": assignments.filter(status='completed').count(),
            "total_expired": assignments.filter(status='expired').count(),
            "total_candidates": assignments.count()
        }

        logger.info(
            "Assessment summary calculated",
            extra={"summary": summary}
        )

        if status_filter:

            if status_filter not in valid_statuses:

                logger.warning(
                    "Invalid status filter received",
                    extra={"status_filter": status_filter}
                )

                return Response(
                    {
                        "detail": "Invalid status filter. Use assigned, in_progress, completed, or expired."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if status_filter != "assigned":
                assignments = assignments.filter(status=status_filter)

                logger.info(
                    "Assignments filtered by status",
                    extra={
                        "status_filter": status_filter,
                        "filtered_count": assignments.count(),
                    }
                )

        # Apply pagination
        paginator = DashboardPagination()
        paginated_assignments = paginator.paginate_queryset(assignments, request)

        logger.info(
            "Pagination applied on assignments",
            extra={
                "paginated_count": len(paginated_assignments)
            }
        )
        # Group assignments by status from paginated results
        grouped = {
            "assigned": [],
            "in_progress": [],
            "completed": [],
            "expired": []
        }

        assigned_bucket = []

        for assignment in paginated_assignments:
            candidate_data = {
                "candidate_assessment_id": assignment.id,
                "id": assignment.candidate.id,
                "username": assignment.candidate.username,
                "email": assignment.candidate.email,
                "first_name": assignment.candidate.first_name,
                "last_name": assignment.candidate.last_name,
                "phone": assignment.candidate.phone,
                "profile": assignment.candidate.profile,
                "date_joined": assignment.candidate.date_joined,
                "assigned_date": assignment.assigned_date,
                "start_time": assignment.start_time,
                "end_time": assignment.end_time,
                "status": assignment.status,
                "score": float(assignment.score) if assignment.score else 0.0,
                "total_marks": float(assignment.total_marks) if assignment.total_marks else 0.0,
                "percentage": float(assignment.percentage) if assignment.percentage else 0.0,
            }
            grouped[assignment.status].append(candidate_data)

            if (
                status_filter == "assigned"
                or assignment.status == "assigned"
            ):
                assigned_bucket.append(candidate_data)

        logger.info(
            "Assignments grouped successfully",
            extra={
                "assigned_count": len(assigned_bucket),
                "in_progress_count": len(grouped["in_progress"]),
                "completed_count": len(grouped["completed"]),
                "expired_count": len(grouped["expired"]),
            }
        )

        response_data = {
            "summary": summary,
            "assigned": assigned_bucket,
            "in_progress": grouped["in_progress"],
            "completed": grouped["completed"],
            "expired": grouped["expired"],
        }

        logger.info(
            "Assessment candidates status response prepared successfully",
            extra={"assessment_id": assessment.id}
        )

        # Return paginated response with grouped data
        return paginator.get_paginated_response(response_data)
class AssessmentCandidatesWithScoreView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, assessment_id: int, *args, **kwargs):
        assessment = get_object_or_404(Assessment, pk=assessment_id)

        if not can_access_assessment(request.user, assessment):
            return Response(
                {"status": "error", "message": "You do not have access to this assessment."},
                status=status.HTTP_403_FORBIDDEN,
            )

        status_filter = request.query_params.get("status", "").strip().lower()
        page = int(request.query_params.get("page", 1))
        page_size = int(request.query_params.get("page_size", 10))

        qs = CandidateAssessment.objects.filter(
            assessment=assessment
        ).select_related("candidate")

        if status_filter in {"assigned", "in_progress", "completed", "expired"}:
            qs = qs.filter(status=status_filter)

        total_count = qs.count()
        qs = qs[(page - 1) * page_size: page * page_size]

        total_marks_default = assessment.questions.aggregate(total=Sum("marks")).get("total") or 0

        data = []
        for a in qs:
            total_marks = float(a.total_marks) if a.total_marks else float(total_marks_default)
            score = float(a.score) if a.score else 0.0
            percentage = round(float(a.percentage), 2) if a.percentage else 0.0

            data.append({
                "candidate_assessment_id": a.id,
                "id": a.candidate.id,
                "first_name": a.candidate.first_name,
                "last_name": a.candidate.last_name,
                "email": a.candidate.email,
                "phone": a.candidate.phone,
                "profile": a.candidate.profile,
                "status": a.status,
                "assigned_at": a.assigned_date,
                "score": score,
                "total_marks": total_marks,
                "percentage": percentage,
            })

        return Response({
            "total_count": total_count,
            "results": data,
        })


class AssessmentAssignView(APIView):
    permission_classes = [AdminPermission]

    def post(self, request, assessment_id: int, *args, **kwargs):

        logger.info(
            "Assessment assignment started",
            extra={
                "assessment_id": assessment_id,
                "assigned_by": request.user.id,
            }
        )

        serializer = AssessmentAssignmentSerializer(data=request.data)

        if not serializer.is_valid():

            logger.warning(
                "Assessment assignment validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        assessment = get_object_or_404(Assessment, pk=assessment_id)

        if not can_access_assessment(request.user, assessment):
            return Response(
                {"status": "error", "message": "You do not have access to this assessment."},
                status=status.HTTP_403_FORBIDDEN,
            )

        logger.info(
            "Assessment fetched successfully",
            extra={
                "assessment_id": assessment.id,
                "title": assessment.title,
            }
        )

        created = 0
        newly_assigned_candidates = []
        candidate_ids = serializer.validated_data["candidate_ids"]

        logger.info(
            "Candidate assignment request received",
            extra={"candidate_count": len(candidate_ids)}
        )

        for candidate_id in candidate_ids:

            try:
                candidate = UserModel.objects.filter(id=candidate_id, role="candidate").first()
                if not candidate:

                    logger.warning(
                        "Candidate not found or invalid role",
                        extra={"candidate_id": candidate_id}
                    )

                    continue

                obj, created_flag = CandidateAssessment.objects.get_or_create(
                    candidate=candidate,
                    assessment=assessment,
                    defaults={"assigned_by": request.user},
                )

                if created_flag:

                    created += 1

                    newly_assigned_candidates.append(candidate_id)

                    logger.info(
                        "Candidate assigned successfully",
                        extra={
                            "candidate_id": candidate.id,
                            "assessment_id": assessment.id,
                        }
                    )

                    send_assignment_notification(obj)

                    logger.info(
                        "Assignment notification sent",
                        extra={
                            "candidate_id": candidate.id,
                            "assessment_id": assessment.id,
                        }
                    )

                else:

                    logger.info(
                        "Candidate already assigned",
                        extra={
                            "candidate_id": candidate.id,
                            "assessment_id": assessment.id,
                        }
                    )

            except Exception as e:

                logger.error(
                    "Error assigning candidate",
                    extra={
                        "candidate_id": candidate_id,
                        "assessment_id": assessment.id,
                        "error": str(e),
                    }
                )

        logger.info(
            "Assessment assignment completed",
            extra={
                "assessment_id": assessment.id,
                "assigned_count": created,
            }
        )

        return Response({
            "assigned_count": created,
            "newly_assigned_candidate_ids": newly_assigned_candidates
        })


class AssessmentUnassignView(APIView):
    permission_classes = [AdminPermission]

    def post(self, request, assessment_id: int, *args, **kwargs):

        logger.info(
            "Assessment unassignment started",
            extra={
                "assessment_id": assessment_id,
                "requested_by": request.user.id,
            }
        )

        serializer = AssessmentUnassignSerializer(data=request.data)
        if not serializer.is_valid():

            logger.warning(
                "Assessment unassignment validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        assessment = get_object_or_404(Assessment, pk=assessment_id)

        if not can_access_assessment(request.user, assessment):
            return Response(
                {"status": "error", "message": "You do not have access to this assessment."},
                status=status.HTTP_403_FORBIDDEN,
            )

        logger.info(
            "Assessment fetched successfully for unassignment",
            extra={
                "assessment_id": assessment.id,
                "title": assessment.title,
            }
        )

        ca_ids = serializer.validated_data[
            "candidate_assessment_ids"
        ]

        logger.info(
            "Candidate assessments received for unassignment",
            extra={"candidate_assessment_count": len(ca_ids)}
        )

        try:
            with transaction.atomic():

                assignments = CandidateAssessment.objects.filter(
                    id__in=ca_ids, assessment=assessment
                )

                logger.info(
                    "Assignments fetched for deletion",
                    extra={
                        "assignments_count": assignments.count()
                    }
                )

                candidate_ids = list(assignments.values_list("candidate_id", flat=True))
                # Clean up related data
                response_deleted, _ = CandidateResponse.objects.filter(
                    candidate_id__in=candidate_ids,
                    assessment=assessment
                ).delete()

                logger.info(
                    "Candidate responses deleted",
                    extra={"deleted_count": response_deleted}
                )

                incident_deleted, _ = ProctoringIncident.objects.filter(
                    candidate_id__in=candidate_ids,
                    assessment=assessment
                ).delete()

                logger.info(
                    "Proctoring incidents deleted",
                    extra={"deleted_count": incident_deleted}
                )

                feedback_deleted, _ = Feedback.objects.filter(
                    candidate_id__in=candidate_ids,
                    assessment=assessment
                ).delete()

                logger.info(
                    "Feedback records deleted",
                    extra={"deleted_count": feedback_deleted}
                )

                deleted, _ = assignments.delete()

                logger.info(
                    "Candidate assignments deleted successfully",
                    extra={"deleted_count": deleted}
                )

        except Exception as e:

            logger.exception(
                "Error during assessment unassignment",
                extra={
                    "assessment_id": assessment.id,
                    "error": str(e),
                }
            )

            return Response(
                {"detail": "Failed to unassign candidates"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        logger.info(
            "Assessment unassignment completed successfully",
            extra={
                "assessment_id": assessment.id,
                "unassigned_count": deleted,
            }
        )

        return Response({"unassigned_count": deleted})


class AssessmentDuplicateView(APIView):
    permission_classes = [AdminPermission]

    def post(self, request, assessment_id: int, *args, **kwargs):

        logger.info(
            "Assessment duplication started",
            extra={
                "assessment_id": assessment_id,
                "requested_by": request.user.id,
            }
        )

        serializer = AssessmentDuplicateSerializer(data=request.data)

        if not serializer.is_valid():

            logger.warning(
                "Assessment duplication validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        assessment = get_object_or_404(Assessment, pk=assessment_id)

        logger.info(
            "Original assessment fetched successfully",
            extra={
                "assessment_id": assessment.id,
                "title": assessment.title,
            }
        )

        new_title = serializer.validated_data.get("title") or f"{assessment.title} (Copy)"

        logger.info(
            "New assessment title prepared",
            extra={"new_title": new_title}
        )

        try:
            with transaction.atomic():

                new_assessment = Assessment.objects.create(
                    title=new_title,
                    description=assessment.description,
                    created_by=request.user,
                    is_active=assessment.is_active,
                    duration=assessment.duration,
                    start_date=assessment.start_date,
                    end_date=assessment.end_date,
                    shuffle_questions=assessment.shuffle_questions,
                    shuffle_options=assessment.shuffle_options,
                    instructions=assessment.instructions,
                )

                logger.info(
                    "Duplicate assessment created",
                    extra={
                        "new_assessment_id": new_assessment.id,
                        "new_title": new_assessment.title,
                    }
                )

                new_assessment.categories.set(assessment.categories.all())

                logger.info(
                    "Categories copied successfully",
                    extra={
                        "new_assessment_id": new_assessment.id,
                        "category_count": assessment.categories.count(),
                    }
                )

                new_assessment.questions.set(assessment.questions.all())

                logger.info(
                    "Questions copied successfully",
                    extra={
                        "new_assessment_id": new_assessment.id,
                        "question_count": assessment.questions.count(),
                    }
                )

        except Exception as e:

            logger.exception(
                "Assessment duplication failed",
                extra={
                    "assessment_id": assessment.id,
                    "error": str(e),
                }
            )

            return Response(
                {"detail": "Failed to duplicate assessment"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        logger.info(
            "Assessment duplicated successfully",
            extra={
                "original_assessment_id": assessment.id,
                "new_assessment_id": new_assessment.id,
            }
        )

        return Response(
            {
                "message": "Assessment duplicated.",
                "assessment_id": new_assessment.id
            },
            status=status.HTTP_201_CREATED,
        )

# Pagination

class DashboardPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = "page_size"
    max_page_size = 200


class ResultsDashboardView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, *args, **kwargs):

        logger.info(
            "Results dashboard API called",
            extra={"user_id": request.user.id}
        )

        # ------------------------ FILTERS ------------------------
        status_filter = request.GET.get("status")  # completed, assigned, expired
        assessment_type = request.GET.get("assessment_type")  # regular, ai, all
        assigned_from = request.GET.get("assigned_from")  # YYYY-MM-DD format
        assigned_to = request.GET.get("assigned_to")      # YYYY-MM-DD format

        logger.info(
            "Dashboard filters received",
            extra={
                "status_filter": status_filter,
                "assessment_type": assessment_type,
                "assigned_from": assigned_from,
                "assigned_to": assigned_to,
            }
        )

        # ------------------------ PAGINATION ---------------------
        paginator = DashboardPagination()

        # FETCH REGULAR + AI CANDIDATE ASSESSMENTS
        # Scope to the requesting admin's organization for org_admin role
        user = request.user
        regular_qs = CandidateAssessment.objects.select_related("candidate", "assessment")
        ai_qs = CandidateAIAssessment.objects.select_related("candidate", "ai_assessment")

        if getattr(user, "role", None) == "org_admin" and user.organization_id:
            regular_qs = regular_qs.filter(candidate__organization=user.organization)
            ai_qs = ai_qs.filter(candidate__organization=user.organization)

        logger.info(
            "Initial querysets loaded",
            extra={
                "regular_count": regular_qs.count(),
                "ai_count": ai_qs.count(),
            }
        )

        # ------------------------ SEARCH ------------------------
        search = request.GET.get("search", "").strip()

        if search:

            logger.info(
                "Applying search filter",
                extra={"search": search}
            )

            regular_qs = regular_qs.filter(
                Q(assessment__title__icontains=search)
                | Q(candidate__username__icontains=search)
                | Q(candidate__email__icontains=search)
                | Q(candidate__first_name__icontains=search)
                | Q(candidate__last_name__icontains=search)
            )

            ai_qs = ai_qs.filter(
                Q(ai_assessment__title__icontains=search)
                | Q(candidate__username__icontains=search)
                | Q(candidate__email__icontains=search)
                | Q(candidate__first_name__icontains=search)
                | Q(candidate__last_name__icontains=search)
            )

        # ------------------------ DATE FILTERS ------------------------
        if assigned_from:
            try:
                from_date = datetime.strptime(assigned_from, "%Y-%m-%d")
                # Set to start of day
                from_date = datetime.combine(from_date.date(), datetime.min.time())
                # Make timezone aware if using timezone.now()
                from_date = timezone.make_aware(from_date) if timezone.is_naive(from_date) else from_date
                
                regular_qs = regular_qs.filter(assigned_date__gte=from_date)
                ai_qs = ai_qs.filter(assigned_date__gte=from_date)

                logger.info(
                    "Applied assigned_from filter",
                    extra={"from_date": str(from_date)}
                )

            except Exception as e:

                logger.error(
                    "Error parsing assigned_from",
                    extra={
                        "assigned_from": assigned_from,
                        "error": str(e),
                    }
                )

        if assigned_to:
            try:
                to_date = datetime.strptime(assigned_to, "%Y-%m-%d")
                # Set to end of day (add 1 day)
                to_date = datetime.combine(to_date.date(), datetime.max.time())
                # Make timezone aware if using timezone.now()
                to_date = timezone.make_aware(to_date) if timezone.is_naive(to_date) else to_date
                
                regular_qs = regular_qs.filter(assigned_date__lte=to_date)
                ai_qs = ai_qs.filter(assigned_date__lte=to_date)

                logger.info(
                    "Applied assigned_to filter",
                    extra={"to_date": str(to_date)}
                )

            except Exception as e:

                logger.error(
                    "Error parsing assigned_to",
                    extra={
                        "assigned_to": assigned_to,
                        "error": str(e),
                    }
                )

        logger.info(
            "Counts after date filters",
            extra={
                "regular_count": regular_qs.count(),
                "ai_count": ai_qs.count(),
            }
        )

        # ----- Status filter -----
        if status_filter in ["completed", "assigned", "expired", "in_progress"]:
            regular_qs = regular_qs.filter(status=status_filter)
            ai_qs = ai_qs.filter(status=status_filter)

            logger.info(
                "Applied status filter",
                extra={
                    "status_filter": status_filter,
                    "regular_count": regular_qs.count(),
                    "ai_count": ai_qs.count(),
                }
            )

        # ----- Assessment type filter -----
        if assessment_type == "regular":
            ai_qs = ai_qs.none()

            logger.info(
                "Applied regular assessment type filter"
            )

        elif assessment_type == "ai":
            regular_qs = regular_qs.none()

            logger.info(
                "Applied AI assessment type filter"
            )

        # ----- Build combined list -----
        combined_list = []

        # REGULAR ASSESSMENTS
        for ca in regular_qs:
            combined_list.append({
                "assessment_id": ca.assessment.id,
                "candidate_assessment_id": ca.id,
                "assessment_type": "regular",
                "assessment_title": ca.assessment.title,
                "candidate_id": ca.candidate.id,
                "candidate_name": ca.candidate.get_full_name() or ca.candidate.username,
                "candidate_email": ca.candidate.email,
                "status": ca.status,
                "score": f"{ca.percentage:.1f}%" if ca.status == "completed" else "-",
                "completed_date": ca.end_time,
                "assigned_date":  ca.assigned_date,
            })

        logger.info(
            "Regular assessments added to response",
            extra={"count": len(combined_list)}
        )

        # AI ASSESSMENTS
        for ca in ai_qs:
            combined_list.append({
                "assessment_id": ca.ai_assessment.id,
                "candidate_ai_assessment_id": ca.id,
                "assessment_type": "ai",
                "assessment_title": ca.ai_assessment.title,
                "candidate_id": ca.candidate.id,
                "candidate_name": ca.candidate.get_full_name() or ca.candidate.username,
                "candidate_email": ca.candidate.email,
                "status": ca.status,
                "score": f"{ca.overall_score:.1f}/10" if ca.status == "completed" else "-",
                "completed_date": ca.end_time,
                "assigned_date":  ca.assigned_date,
            })

        logger.info(
            "AI assessments added to response",
            extra={"total_combined_count": len(combined_list)}
        )

        # ----- Sort by completed date -----
        combined_list = sorted(
            combined_list,
            key=lambda x: (x["status"] != "completed", x["completed_date"] or timezone.now()),
            reverse=True,
        )

        logger.info("Combined list sorted successfully")

        # ----- Paginate -----
        paginated = paginator.paginate_queryset(combined_list, request)


        logger.info(
            "Pagination applied",
            extra={"paginated_count": len(paginated)}
        )
        # -------------------- METRICS DATA --------------------
        from organization.context import current_organization_id, current_user_is_super_admin
        is_super = current_user_is_super_admin.get()
        org_id = current_organization_id.get()

        if not is_super:
            if org_id:
                # Filter by both organization and current admin ownership
                total_assessments = (
                    Assessment.objects.filter(created_by=request.user).count() 
                    + AIAssessment.objects.filter(created_by=request.user).count()
                )
                total_candidates = User.objects.filter(
                    role="candidate", 
                    organization_id=org_id,
                    created_by=request.user
                ).count()
                
                total_completed = (
                    CandidateAssessment.objects.filter(status="completed", candidate__created_by=request.user).count()
                    + CandidateAIAssessment.objects.filter(status="completed", candidate__created_by=request.user).count()
                )

                # Combined avg score for candidates created by this admin
                regular_stats = CandidateAssessment.objects.filter(
                    status="completed", 
                    candidate__created_by=request.user
                ).aggregate(
                    reg_avg=Avg("percentage"),
                    reg_count=Count("id"),
                )

                ai_stats = CandidateAIAssessment.objects.filter(
                    status="completed", 
                    candidate__created_by=request.user
                ).aggregate(
                    ai_avg=Avg("overall_score"),
                    ai_count=Count("id"),
                )
            else:
                total_assessments = 0
                total_candidates = 0
                total_completed = 0
                regular_stats = {"reg_avg": 0, "reg_count": 0}
                ai_stats = {"ai_avg": 0, "ai_count": 0}
        else:
            total_assessments = Assessment.objects.count() + AIAssessment.objects.count()
            total_candidates = User.objects.filter(role="candidate").count()
            total_completed = (
                CandidateAssessment.objects.filter(status="completed").count()
                + CandidateAIAssessment.objects.filter(status="completed").count()
            )

            # Combined avg score
            regular_stats = CandidateAssessment.objects.filter(status="completed").aggregate(
                reg_avg=Avg("percentage"),
                reg_count=Count("id"),
            )

            ai_stats = CandidateAIAssessment.objects.filter(status="completed").aggregate(
                ai_avg=Avg("overall_score"),
                ai_count=Count("id"),
            )

        reg_avg = regular_stats["reg_avg"] or 0
        ai_avg = (ai_stats["ai_avg"] or 0) * 10  # convert 0–10 → 0–100

        total_count = (regular_stats["reg_count"] or 0) + (ai_stats["ai_count"] or 0)

        if total_count > 0:
            avg_score = (
                (reg_avg * (regular_stats["reg_count"] or 0))
                + (ai_avg * (ai_stats["ai_count"] or 0))
            ) / total_count
        else:
            avg_score = 0

        logger.info(
            "Average score calculated",
            extra={"avg_score": round(avg_score, 2)}
        )

        response_data = {
            "metrics": {
                "total_assessments": total_assessments,
                "total_candidates": total_candidates,
                "total_completed": total_completed,
                "avg_score": round(avg_score, 2),
            },
            "assessment_candidates": paginated,
        }

        logger.info(
            "Results dashboard response prepared successfully"
        )

        # FINAL JSON RESPONSE
        return paginator.get_paginated_response(
            response_data
        )

class AssessmentResultsView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, assessment_id: int, *args, **kwargs):

        logger.info(
            "Assessment results API called",
            extra={
                "assessment_id": assessment_id,
                "user_id": request.user.id,
            }
        )

        assessment = get_object_or_404(Assessment, pk=assessment_id)

        logger.info(
            "Assessment fetched successfully",
            extra={
                "assessment_id": assessment.id,
                "title": assessment.title,
            }
        )

        results = CandidateAssessment.objects.filter(
            assessment=assessment, status="completed"
        )

        logger.info(
            "Completed assessment results fetched",
            extra={
                "assessment_id": assessment.id,
                "results_count": results.count(),
            }
        )

        serializer = CandidateAssessmentSerializer(results, many=True)

        response_data = {
            "assessment": AssessmentSerializer(assessment).data,
            "results": serializer.data,
        }

        logger.info(
            "Assessment results response prepared successfully",
            extra={"assessment_id": assessment.id}
        )

        return Response(response_data)


class CandidateResultsView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, candidate_id: int, *args, **kwargs):

        logger.info(
            "Candidate results API called",
            extra={
                "candidate_id": candidate_id,
                "user_id": request.user.id,
            }
        )

        candidate = get_object_or_404(UserModel, pk=candidate_id, role="candidate")


        logger.info(
            "Candidate fetched successfully",
            extra={
                "candidate_id": candidate.id,
                "candidate_email": candidate.email,
            }
        )

        results = CandidateAssessment.objects.filter(candidate=candidate)

        logger.info(
            "Candidate assessment results fetched",
            extra={
                "candidate_id": candidate.id,
                "results_count": results.count(),
            }
        )

        response_data = {
                "candidate": CandidateSerializer(candidate).data,
                "results": CandidateAssessmentSerializer(results, many=True).data,

        }

        logger.info(
            "Candidate results response prepared successfully",
            extra={"candidate_id": candidate.id}
        )

        return Response(response_data)


class AdminCandidateAssessmentResultView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, candidate_assessment_id: int, *args, **kwargs):

        logger.info(
            "Admin candidate assessment result API called",
            extra={
                "candidate_assessment_id": candidate_assessment_id,
                "user_id": request.user.id,
            }
        )

        candidate_assessment = get_object_or_404(
            CandidateAssessment, pk=candidate_assessment_id
        )

        logger.info(
            "Candidate assessment fetched successfully",
            extra={
                "candidate_assessment_id": candidate_assessment.id,
                "candidate_id": candidate_assessment.candidate.id,
                "assessment_id": candidate_assessment.assessment.id,
            }
        )

        assessment = candidate_assessment.assessment
        candidate = candidate_assessment.candidate

        # All questions in the assessment
        all_questions = assessment.questions.all()

        logger.info(
            "Assessment questions fetched",
            extra={
                "assessment_id": assessment.id,
                "question_count": all_questions.count(),
            }
        )

        # Candidate responses
        responses_qs = CandidateResponse.objects.filter(
            candidate=candidate,
            assessment=assessment,
        ).select_related("question")

        logger.info(
            "Candidate responses fetched",
            extra={
                "candidate_id": candidate.id,
                "response_count": responses_qs.count(),
            }
        )

        # Map responses by question_id for easy lookup
        responses_dict = {r.question_id: r for r in responses_qs}

        # Prepare response data

        response_data = []
        for question in all_questions:
            response = responses_dict.get(question.id)
            if response:
                serialized = CandidateResponseSerializer(response).data


                logger.info(
                    "Serialized attempted question response",
                    extra={
                        "question_id": question.id,
                        "candidate_id": candidate.id,
                    }
                )

            else:
                logger.info(
                    "Question not attempted",
                    extra={
                        "question_id": question.id,
                        "candidate_id": candidate.id,
                    }
                )

                serialized = {
                    "question_id": question.id,
                    "question_title": question.title,
                    "question_description": question.description,
                    "question_marks": question.marks,
                    "question_difficulty": question.difficulty,
                    "question_type": question.question_type,
                    "answer": None,
                    "answer_text": "",
                    "question_options": CandidateResponseSerializer.get_question_options(
                        CandidateResponseSerializer(), type('obj', (object,), {'question': question})
                    ),
                    "correct_answer": question.correct_answer or "",
                    "correct_answer_text": question.correct_answer or "",
                    "is_correct": False,
                    "marks_obtained": 0,
                }
            response_data.append(serialized)

        logger.info(
            "Response data prepared",
            extra={
                "candidate_assessment_id":
                    candidate_assessment.id,
                "response_count": len(response_data),
            }
        )

        # FETCH FEEDBACK (same as candidate view)
        feedback = Feedback.objects.filter(
            candidate=candidate,
            assessment=assessment
        ).first()

        if feedback:

            logger.info(
                "Feedback fetched successfully",
                extra={
                    "candidate_id": candidate.id,
                    "assessment_id": assessment.id,
                }
            )

        else:

            logger.info(
                "No feedback found",
                extra={
                    "candidate_id": candidate.id,
                    "assessment_id": assessment.id,
                }
            )

        # Proctoring incidents
        incidents = ProctoringIncident.objects.filter(
            candidate=candidate,
            assessment=assessment,
        ).order_by("-timestamp")

        logger.info(
            "Proctoring incidents fetched",
            extra={
                "candidate_id": candidate.id,
                "incident_count": incidents.count(),
            }
        )

        incident_summary = {}

        for incident in incidents:

            label = incident.get_incident_type_display()

            incident_summary[label] = (
                incident_summary.get(label, 0) + 1
            )

        logger.info(
            "Incident summary prepared",
            extra={
                "candidate_id": candidate.id,
                "incident_summary": incident_summary,
            }
        )

        response_payload = {
                "candidate_assessment": CandidateAssessmentSerializer(
                    candidate_assessment
                ).data,
                "responses": response_data,
                "proctoring_incidents": ProctoringIncidentSerializer(incidents, many=True).data,
                "incident_summary": incident_summary,

                "feedback": {
                    "rating": feedback.rating,
                    "comments": feedback.comments,
                } if feedback else None,
        }

        logger.info(
            "Admin candidate assessment result response prepared successfully",
            extra={
                "candidate_assessment_id":
                    candidate_assessment.id
            }
        )

        return Response(response_payload)

class CandidateSubmissionsView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, assessment_id: int, candidate_id: int, *args, **kwargs):

        logger.info(
            "Candidate submissions API called",
            extra={
                "assessment_id": assessment_id,
                "candidate_id": candidate_id,
                "user_id": request.user.id,
            }
        )

        assessment = get_object_or_404(
            Assessment,
            pk=assessment_id
        )

        logger.info(
            "Assessment fetched successfully",
            extra={
                "assessment_id": assessment.id,
                "assessment_title": assessment.title,
            }
        )

        candidate = get_object_or_404(
            UserModel,
            pk=candidate_id
        )

        logger.info(
            "Candidate fetched successfully",
            extra={
                "candidate_id": candidate.id,
                "candidate_email": candidate.email,
            }
        )

        candidate_assessment = CandidateAssessment.objects.filter(
            candidate=candidate, assessment=assessment
        ).first()

        if candidate_assessment:

            logger.info(
                "Candidate assessment found",
                extra={
                    "candidate_assessment_id":
                        candidate_assessment.id,
                    "status":
                        candidate_assessment.status,
                }
            )

        else:

            logger.warning(
                "Candidate assessment not found",
                extra={
                    "candidate_id": candidate.id,
                    "assessment_id": assessment.id,
                }
            )

        responses = CandidateResponse.objects.filter(
            candidate=candidate, assessment=assessment
        ).select_related("question")

        logger.info(
            "Candidate responses fetched",
            extra={
                "candidate_id": candidate.id,
                "response_count": responses.count(),
            }
        )

        response_data = CandidateResponseSerializer(responses, many=True).data

        logger.info(
            "Candidate responses serialized successfully",
            extra={
                "candidate_id": candidate.id,
                "serialized_count": len(response_data),
            }
        )

        response_payload = {
            "assessment": AssessmentSerializer(
                assessment
            ).data,

            "candidate": CandidateSerializer(
                candidate
            ).data,

            "candidate_assessment":
                CandidateAssessmentSerializer(
                    candidate_assessment
                ).data
                if candidate_assessment
                else None,

            "responses": response_data,
        }

        logger.info(
            "Candidate submissions response prepared successfully",
            extra={
                "assessment_id": assessment.id,
                "candidate_id": candidate.id,
            }
        )

        return Response(response_payload)

class SaveAnswerView(APIView):
    permission_classes = [CandidatePermission]

    def post(self, request, *args, **kwargs):

        logger.info(
            "Save answer API called",
            extra={"candidate_id": request.user.id}
        )

        serializer = SaveAnswerSerializer(data=request.data)

        if not serializer.is_valid():

            logger.warning(
                "Save answer validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        question = get_object_or_404(
            Question,
            pk=serializer.validated_data["question_id"]
        )

        logger.info(
            "Question fetched successfully",
            extra={
                "question_id": question.id,
                "question_type": question.question_type,
            }
        )

        assessment_id = serializer.validated_data.get(
            "assessment_id"
        )

        if assessment_id:

            assessment = get_object_or_404(
                Assessment,
                pk=assessment_id
            )

            logger.info(
                "Assessment fetched using assessment_id",
                extra={
                    "assessment_id": assessment.id,
                    "candidate_id": request.user.id,
                }
            )

        else:
            candidate_assessment = (
                CandidateAssessment.objects.filter(
                    candidate=request.user,
                    status__in=[
                        "in_progress",
                        "assigned"
                    ]
                )
                .order_by("-start_time")
                .first()
            )

            if not candidate_assessment:

                logger.warning(
                    "Assessment context missing",
                    extra={"candidate_id": request.user.id}
                )

                return Response(
                    {"detail": "Assessment context missing."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            assessment = (candidate_assessment.assessment)

        updates = {}

        if "answer" in serializer.validated_data:

            updates["answer"] = (
                serializer.validated_data["answer"]
                or ""
            ).rstrip("\r\n")

        if "code_language" in serializer.validated_data:

            updates["code_language"] = (
                serializer.validated_data[
                    "code_language"
                ]
                or ""
            )

        # =====================================================
        # SQL QUESTION
        # =====================================================

        if (
            question.question_type == "sql"
            and "answer" in updates
        ):

            import json
            from .utils_sql_judge0 import (
                build_sqlite_script,
                is_select_only,
                parse_rows,
                rowset_equal,
                submit_to_judge0_sql,
            )

            query = (updates["answer"].strip())
            separator = ( "\n\n---[OUTPUT]---\n")
            summary_separator = ("\n\n---[SUMMARY]---\n")
            if not is_select_only(query):
                updates["answer"] = (
                    query
                    + separator
                    + (
                        "ERROR: Only SELECT "
                        "queries are allowed."
                    )
                )

                updates["answer"] += (
                    summary_separator
                    + json.dumps({
                        "passed_count": 0,
                        "total_cases": 0,
                        "earned_points": 0,
                        "total_points": 0,
                        "expected_output": "",
                    })
                )

                updates["marks_obtained"] = 0
                updates["is_correct"] = False

            else:

                try:

                    sqlmeta = (
                        SQLQuestion.objects
                        .select_related("dataset")
                        .filter(question=question)
                        .first()
                    )

                    if (
                        not sqlmeta
                        or not sqlmeta.dataset
                    ):

                        updates["answer"] = (
                            query
                            + separator
                            + (
                                "ERROR: SQL metadata "
                                "or dataset missing."
                            )
                        )

                        updates["marks_obtained"] = 0
                        updates["is_correct"] = False

                    else:

                        schema = (
                            sqlmeta.dataset.schema_ddl
                            or ""
                        )

                        seed = (
                            sqlmeta.dataset.seed_sql
                            or ""
                        )

                        max_rows = (
                            sqlmeta.max_rows
                            or 5000
                        )

                        preview_script = (
                            build_sqlite_script(
                                schema,
                                seed,
                                "",
                                query,
                                max_rows,
                            )
                        )

                        preview_result = (
                            submit_to_judge0_sql(
                                preview_script
                            )
                        )

                        preview_err = (
                            preview_result.get(
                                "stderr"
                            )
                        )

                        if preview_err:

                            updates["answer"] = (
                                query
                                + separator
                                + (
                                    f"ERROR: "
                                    f"{preview_err[:500]}"
                                )
                            )

                            testcases_sql = list(
                                SQLTestCase.objects
                                .filter(
                                    question=question
                                )
                            )

                            total_pts = sum(
                                float(tc.points or 1)
                                for tc in testcases_sql
                            )

                            updates["answer"] += (
                                summary_separator
                                + json.dumps({
                                    "passed_count": 0,
                                    "total_cases":
                                        len(
                                            testcases_sql
                                        ),
                                    "earned_points": 0,
                                    "total_points":
                                        total_pts,
                                    "expected_output":
                                        "",
                                })
                            )

                            updates[
                                "marks_obtained"
                            ] = 0

                            updates[
                                "is_correct"
                            ] = False

                        else:

                            (
                                _,
                                preview_rows
                            ) = parse_rows(
                                preview_result.get(
                                    "stdout"
                                ) or ""
                            )

                            preview_output = (
                                json.dumps(
                                    preview_rows[:10],
                                    indent=2
                                )
                            )

                            updates["answer"] = (
                                query
                                + separator
                                + preview_output
                            )

                            testcases_sql = list(
                                SQLTestCase.objects
                                .filter(
                                    question=question
                                )
                            )

                            total_pts = 0.0
                            earned_pts = 0.0
                            passed = 0
                            expected_output = ""

                            for tc in testcases_sql:

                                pts = float(
                                    tc.points or 1
                                )

                                total_pts += pts

                                setup_sql = (
                                    getattr(
                                        tc,
                                        "setup_sql",
                                        ""
                                    )
                                    or ""
                                )

                                ref_script = (
                                    build_sqlite_script(
                                        schema,
                                        seed,
                                        setup_sql,
                                        sqlmeta.reference_solution,
                                        max_rows,
                                    )
                                )

                                cand_script = (
                                    build_sqlite_script(
                                        schema,
                                        seed,
                                        setup_sql,
                                        query,
                                        max_rows,
                                    )
                                )

                                ref_result = (
                                    submit_to_judge0_sql(
                                        ref_script
                                    )
                                )

                                cand_result = (
                                    submit_to_judge0_sql(
                                        cand_script
                                    )
                                )

                                if (
                                    ref_result.get(
                                        "stderr"
                                    )
                                    or cand_result.get(
                                        "stderr"
                                    )
                                ):
                                    continue

                                (
                                    _,
                                    ref_rows
                                ) = parse_rows(
                                    ref_result.get(
                                        "stdout"
                                    )
                                    or ""
                                )

                                (
                                    _,
                                    cand_rows
                                ) = parse_rows(
                                    cand_result.get(
                                        "stdout"
                                    )
                                    or ""
                                )

                                if (
                                    not expected_output
                                    and ref_rows
                                ):

                                    expected_output = (
                                        json.dumps(
                                            ref_rows[:5],
                                            indent=2
                                        )
                                    )

                                if rowset_equal(
                                    ref_rows,
                                    cand_rows,
                                    strict_order=bool(
                                        sqlmeta.strict_column_order
                                    ),
                                    float_tol=float(
                                        sqlmeta.float_tolerance
                                        or 0.0
                                    ),
                                ):

                                    passed += 1
                                    earned_pts += pts

                            updates["answer"] += (
                                summary_separator
                                + json.dumps({
                                    "passed_count":
                                        passed,
                                    "total_cases":
                                        len(
                                            testcases_sql
                                        ),
                                    "earned_points":
                                        earned_pts,
                                    "total_points":
                                        total_pts,
                                    "expected_output":
                                        expected_output,
                                })
                            )

                            updates[
                                "marks_obtained"
                            ] = earned_pts

                            updates[
                                "is_correct"
                            ] = (
                                earned_pts
                                >= total_pts
                                and total_pts > 0
                            )

                except Exception as e:

                    logger.exception(
                        "SQL grading error"
                    )

                    updates["answer"] = (
                        query
                        + separator
                        + f"ERROR: {str(e)}"
                    )

                    updates["marks_obtained"] = 0
                    updates["is_correct"] = False

        # =====================================================
        # CODING QUESTION
        # =====================================================

        if (
            question.question_type == "coding"
            and "answer" in updates
        ):

            import json

            code = updates["answer"]

            language = (
                updates.get(
                    "code_language",
                    "python"
                )
                .strip()
                .lower()
            )

            separator = (
                "\n\n---[RESULTS]---\n"
            )

            summary_separator = (
                "\n\n---[SUMMARY]---\n"
            )

            testcases = list(
                question.testcases.all()
                .order_by("id")
            )

            if not testcases:

                testcases = [
                    TestCase(
                        question=question,
                        input_data=(
                            question.sample_input
                            or ""
                        ),
                        expected_output=(
                            question.sample_output
                            or ""
                        ),
                        points=(
                            question.marks
                            or 1
                        ),
                        is_hidden=False,
                    )
                ]

            execution_results = []

            total_pts = 0.0
            earned_pts = 0.0
            passed_count = 0

            # Question marks auto divide karo testcases mein
            total_question_marks = float(question.marks or len(testcases))
            per_tc_marks = round(total_question_marks / len(testcases), 1)

            for tc in testcases:

                pts = per_tc_marks

                total_pts += pts

                result = (
                    execute_code_with_judge0(
                        code,
                        language,
                        tc.input_data or ""
                    )
                )

                stdout = normalize_output(
                    result.get(
                        "stdout",
                        ""
                    )
                )

                expected = normalize_output(
                    tc.expected_output
                    or ""
                )

                passed = (
                    result.get("status")
                    == "Accepted"
                    and stdout == expected
                )

                if passed:

                    passed_count += 1
                    earned_pts += pts

                execution_results.append({

                    "input":
                        (
                            tc.input_data
                            if not tc.is_hidden
                            else "(hidden)"
                        ),

                    "expected_output":
                        (
                            tc.expected_output
                            if not tc.is_hidden
                            else "(hidden)"
                        ),

                    "stdout": stdout,

                    "stderr":
                        result.get(
                            "stderr"
                        ),

                    "compile_output":
                        result.get(
                            "compile_output"
                        ),

                    "status":
                        result.get(
                            "status"
                        ),

                    "passed": passed,

                    "time":
                        result.get("time"),

                    "memory":
                        result.get("memory"),

                    "points": pts,

                    "received":
                        pts if passed else 0.0,
                })
            # Loop ke baad ek baar calculate
            earned_pts = round((passed_count / len(testcases)) * total_question_marks, 1)
            total_pts = total_question_marks

            summary = {
                "passed_count":
                    passed_count,
                "total_cases":
                    len(testcases),
                "earned_points":
                    earned_pts,
                "total_points": round(total_pts, 1),
            }

            updates["marks_obtained"] = earned_pts

            updates["is_correct"] = (
                earned_pts >= total_pts
                and total_pts > 0
            )

            updates["answer"] = (
                code
                + separator
                + json.dumps(
                    execution_results,
                    indent=2
                )
                + summary_separator
                + json.dumps(
                    summary,
                    indent=2
                )
            )

        # =====================================================
        # SAVE RESPONSE
        # =====================================================

        response_obj, created = (
            CandidateResponse.objects
            .get_or_create(
                candidate=request.user,
                assessment=assessment,
                question=question,
                defaults=updates
                if updates else {},
            )
        )

        if updates and not created:

            for key, value in updates.items():

                setattr(
                    response_obj,
                    key,
                    value
                )

            response_obj.save()

        if question.question_type in [
            "mcq_single",
            "mcq_multiple",
            "true_false",
            "fill_blank",
        ]:

            response_obj.evaluate_mcq()

            ca = (
                CandidateAssessment.objects
                .filter(
                    candidate=request.user,
                    assessment=assessment,
                )
                .first()
            )

            if ca:
                ca.calculate_score()

        else:

            response_obj.save(
                update_fields=[
                    "responded_at",
                    "answer",
                    "marks_obtained",
                    "is_correct",
                    "code_language",
                ]
            )

            if question.question_type in [
                "sql",
                "coding",
            ]:

                ca = (
                    CandidateAssessment.objects
                    .filter(
                        candidate=request.user,
                        assessment=assessment,
                    )
                    .first()
                )

                if ca:
                    ca.calculate_score()

        candidate_assessment = (
            CandidateAssessment.objects
            .filter(
                candidate=request.user,
                assessment=assessment,
            )
            .first()
        )

        totals = (
            {
                "score":
                    candidate_assessment.score,

                "total_marks":
                    candidate_assessment.total_marks,

                "percentage":
                    round(
                        candidate_assessment.percentage,
                        2
                    ),
            }
            if candidate_assessment
            else {}
        )

        response_payload = {
            "status": "success",
            "question_id": question.id,
            "answer": response_obj.answer,
            "marks_obtained":
                response_obj.marks_obtained,
            "current_total": totals,
        }

        return Response(response_payload)

class RunCodeView(APIView):
    permission_classes = [IsAuthenticated]

    MAX_OUTPUT_SIZE = 10000

    def truncate_output(self, value):
        """
        Prevent massive payloads from freezing frontend.
        """
        if not value:
            return ""
        value = str(value)
        if len(value) > self.MAX_OUTPUT_SIZE:
            return (
                value[:self.MAX_OUTPUT_SIZE]
                + "\n\n[Output truncated]"
            )
        return value

    def post(self, request, *args, **kwargs):
        serializer = RunCodeSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        question_id = serializer.validated_data.get("question_id")
        source_code = (
            serializer.validated_data.get("source_code") or 
            serializer.validated_data.get("code", "")
        )
          
        language = serializer.validated_data.get("language","")

        override_stdin = serializer.validated_data.get("stdin", None)
        use_custom_input = serializer.validated_data.get("use_custom_input", False)
        assessment_id = serializer.validated_data.get("assessment_id", None)
        logger.info("RunCodeView request: " "question_id=%s assessment_id=%s " "language=%s", question_id,assessment_id,language,)
        
        if not question_id:
            return Response(
                {"detail": "question_id is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate source code
        if not source_code.strip():
            return Response(
                {
                    "detail":
                    "Source code cannot be empty"
                },
                status=status.HTTP_400_BAD_REQUEST
            )    
        # Validate language
        normalized_language = (
            language.strip().lower()
        )
        allowed_languages = (
            settings.JUDGE0_LANGUAGE_MAPPING.keys()
        )
        if (
            normalized_language
            not in allowed_languages
        ):
            logger.warning(
                "Unsupported language: %s",
                language
            )
            return Response(
                {
                    "detail":
                    f"Unsupported language: "
                    f"{language}"
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        # Load question
        try:
            question = Question.objects.get(pk=question_id)
        except Question.DoesNotExist:
            logger.error("Question %s not found",question_id)
            return Response(
                {"detail": f"Question "f"{question_id} "f"not found"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        # Validate coding question
        if question.question_type != "coding":
            logger.error("Question %s is not coding " "(type=%s)", question_id, question.question_type, )
            return Response(
                {"detail": f"Question " f"{question_id} " f"is not coding type"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        testcases = list(question.testcases.all().order_by("id"))
        # Fallback testcase
        if not testcases:
            fallback_input = (
                question.sample_input or ""
            )
            fallback_expected = (
                question.sample_output or ""
            )
            testcases = [ TestCase(
                question=question,
                input_data=fallback_input,
                expected_output=fallback_expected,
                points=question.marks or 1.0,
                    is_hidden=False,
                )
            ]
            
        first_input = (
            override_stdin
            if (
                use_custom_input
                and override_stdin is not None
            )
            else (
                testcases[0].input_data or ""
            )
        )

        logger.info(
            "Executing initial Judge0 run"
        )

        initial_result = execute_code_with_judge0(
            source_code,
            normalized_language,
            first_input,
        )

        initial_status = (
            initial_result.get("status")
        )

        fatal_errors = [
            "Compilation Error",
            "Runtime Error",
            "Segmentation Fault",
            "Floating Point Exception",
            "Time Limit Exceeded",
            "Memory Limit Exceeded",
            "Internal Error",
            "Exec Format Error",
        ]

        # -----------------------------------
        # Fail fast for fatal errors
        # -----------------------------------

        if initial_status in fatal_errors:

            logger.warning(
                "Execution failed early: %s",
                initial_status
            )

            return Response({
                "status": "success",
                "data": {
                    "results": [{
                        "input": "",
                        "expected_output": "",
                        "stdout": self.truncate_output(
                            initial_result.get(
                                "stdout",
                                ""
                            )
                        ),
                        "stderr": self.truncate_output(
                            initial_result.get(
                                "stderr",
                                ""
                            )
                        ),
                        "compile_output":
                        self.truncate_output(
                            initial_result.get(
                                "compile_output",
                                ""
                            )
                        ),
                        "status": initial_status,
                        "passed": False,
                        "time": initial_result.get(
                            "time"
                        ),
                        "memory": initial_result.get(
                            "memory"
                        ),
                        "is_hidden": False,
                        "points": 0,
                    }],
                    "summary": {
                        "total_points": 0,
                        "earned_points": 0,
                        "passed_count": 0,
                        "total_cases": 0,
                    }
                }
            })

        results = []
        total_points = 0.0
        earned_points = 0.0
        total_question_marks = float(question.marks or len(testcases))
        #per_tc_marks = total_question_marks / len(testcases)
        per_tc_marks = round(total_question_marks / len(testcases), 1)

        for index, tc in enumerate(testcases):
            stdin = (
                override_stdin
                if (use_custom_input and override_stdin is not None)
                else (tc.input_data or "")
            )
            # Reuse first execution
            if index == 0:
                result = initial_result
            else:
                logger.info("Executing testcase #%s", index + 1)
                result = execute_code_with_judge0(source_code,normalized_language,stdin,)

            stdout = (
                (
                    result.get("stdout")
                    or ""
                )
                .replace("\r\n", "\n")
                .strip()
            )

            expected = (
                (
                    tc.expected_output
                    or ""
                )
                .replace("\r\n", "\n")
                .strip()
            )

            stderr = self.truncate_output(
                result.get("stderr", "")
            )

            compile_output = (
                self.truncate_output(
                    result.get(
                        "compile_output",
                        ""
                    )
                )
            )

            stdout = self.truncate_output(
                stdout
            )

            execution_status = (
                result.get("status", "")
            )

            passed = (
                execution_status
                == "Accepted"
                and stdout == expected
            )

            points = per_tc_marks

            
            if passed:
                earned_points += points
            total_points += points

            results.append({
                "input": (tc.input_data if not tc.is_hidden else "(hidden)"),
                "expected_output": (tc.expected_output if not tc.is_hidden else "(hidden)"),
                "stdout": stdout,
                "stderr": stderr,
                "compile_output":compile_output,
                "status":execution_status,
                "passed": passed,
                "time": result.get("time"),
                "memory": result.get("memory"),
                "is_hidden": tc.is_hidden,
                "points":points,
                "received":points if passed else 0.0,
            })
        passed_count_final = sum(1 for r in results if r["passed"])
        earned_points = round((passed_count_final / len(testcases)) * total_question_marks, 1)

        summary = {
            "total_points": round(total_question_marks, 1),
            "earned_points": earned_points,
            "passed_count": passed_count_final,
            "total_cases": len(results),
        }

        # Auto-save coding response and marks with retry logic
        if assessment_id:
            import time
            from django.db import (OperationalError)
            max_retries = 5
            saved = False
            
            for attempt in range(max_retries):
                try:
                    response_obj, _ = (CandidateResponse.objects.get_or_create(
                        candidate=request.user,
                        assessment_id=assessment_id,
                        question=question,
                        defaults={"answer":source_code,"code_language":normalized_language,}
                    ) ) 
                    response_obj.answer = (source_code)
                    response_obj.code_language = (normalized_language)
                    response_obj.marks_obtained = (earned_points)
                    response_obj.is_correct = (earned_points >= total_points and total_points > 0)
                    response_obj.answer = response_obj.answer + f"\n\n---[SUMMARY]---\n{json.dumps(summary)}"
                    response_obj.save()
                    saved = True
                    break
                    
                except OperationalError as oe:
                    logger.warning(
                        "DB OperationalError "
                        "attempt=%s/%s error=%s",
                        attempt + 1,
                        max_retries,
                        oe,
                    )

                    if (
                        "locked"
                        in str(oe).lower()
                        and attempt
                        < max_retries - 1
                    ):

                        time.sleep(
                            0.05 * (attempt + 1)
                        )

                        continue

                    logger.exception(
                        "Final DB lock failure"
                    )

                    break

                except Exception:

                    logger.exception(
                        "Unexpected autosave error"
                    )

                    break

            if saved:
                try:
                    candidate_assessment = (CandidateAssessment.objects.filter(
                        candidate=request.user, assessment_id=assessment_id,
                    ).first()
                    )
                    if candidate_assessment:
                        candidate_assessment.calculate_score()
                except Exception:
                    logger.exception("Failed to recalculate "
                        "assessment score"
                    )
            else:
                logger.warning("Auto-save failed " "for user=%s question=%s",
                             request.user.id, question_id)

        return Response({
            "status": "success",
            "data": {
                "results": results,
                "summary": summary
            }
        })

class ProctoringIncidentView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, *args, **kwargs):
        serializer = ProctoringIncidentCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        incident = serializer.save()
        try:
            send_cheating_alert_email(incident)
        except Exception as exc:
            logger.exception("Failed to send proctoring alert email", extra={"incident_id": incident.id})
        return Response(ProctoringIncidentSerializer(incident).data, status=status.HTTP_201_CREATED)


class SQLRunView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, *args, **kwargs):

        logger.info(
            "SQL run API called",
            extra={"candidate_id": request.user.id}
        )

        serializer = SQLRunSerializer(data=request.data)

        if not serializer.is_valid():

            logger.warning(
                "SQL run validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        question = get_object_or_404(
            Question,
            pk=serializer.validated_data["question_id"]
        )

        logger.info(
            "Question fetched successfully",
            extra={
                "question_id": question.id,
                "question_type": question.question_type,
            }
        )

        if question.question_type != "sql":

            logger.warning(
                "Non-SQL question used in SQLRunView",
                extra={"question_id": question.id}
            )

            return Response(
                {"detail": "Question is not SQL."}, status=status.HTTP_400_BAD_REQUEST
            )

        sqlmeta = (
            SQLQuestion.objects.select_related("dataset")
            .filter(question=question)
            .first()
        )

        if not sqlmeta or not sqlmeta.dataset:

            logger.warning(
                "SQL metadata missing",
                extra={"question_id": question.id}
            )

            return Response(
                {"detail": "SQL metadata missing."},
                status=status.HTTP_400_BAD_REQUEST
            )

        query = serializer.validated_data["query"]

        logger.info(
            "SQL query received",
            extra={"question_id": question.id}
        )

        if not is_select_only(query):

            logger.warning(
                "Non-SELECT query attempted",
                extra={
                    "candidate_id": request.user.id,
                    "question_id": question.id,
                }
            )

            return Response(
                {"detail": "Only SELECT queries allowed."},
                status=status.HTTP_400_BAD_REQUEST
            )

        script = build_sqlite_script(
            sqlmeta.dataset.schema_ddl or "",
            sqlmeta.dataset.seed_sql or "",
            "",
            query,
            sqlmeta.max_rows or 5000,
        )

        logger.info(
            "SQLite execution script prepared",
            extra={"question_id": question.id}
        )

        result = submit_to_judge0_sql(script)

        stderr = result.get("stderr")

        if stderr:

            logger.error(
                "SQL execution failed",
                extra={
                    "question_id": question.id,
                    "error": stderr[:300],
                }
            )

            return Response({"error": stderr[:1000]}, status=status.HTTP_400_BAD_REQUEST)
        _, rows = parse_rows(result.get("stdout") or "")

        logger.info(
            "SQL query executed successfully",
            extra={
                "question_id": question.id,
                "row_count": len(rows),
            }
        )

        expected_rows = []
        try:
            testcases = list(SQLTestCase.objects.filter(question=question))
            if testcases and sqlmeta.reference_solution:
                ref_script = build_sqlite_script(
                    sqlmeta.dataset.schema_ddl or "",
                    sqlmeta.dataset.seed_sql or "",
                    "",
                    sqlmeta.reference_solution,
                    sqlmeta.max_rows or 5000,
                )
                ref_result = submit_to_judge0_sql(ref_script)
                if not ref_result.get("stderr"):
                    _, expected_rows = parse_rows(ref_result.get("stdout") or "")
                    expected_rows = expected_rows[:50]
        except Exception:
            expected_rows = []
            logger.exception("Failed to evaluate reference solution", extra={"question_id": question.id})
        return Response({
            "rows": rows[:50],
            "truncated": len(rows) > 50,
            "expected_rows": expected_rows,
        })


class SQLGradeView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, *args, **kwargs):

        logger.info(
            "SQL grade API called",
            extra={"candidate_id": request.user.id}
        )

        serializer = SQLGradeSerializer(data=request.data)

        if not serializer.is_valid():

            logger.warning(
                "SQL grade validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        question = get_object_or_404(
            Question,
            pk=serializer.validated_data["question_id"]
        )

        logger.info(
            "SQL grading question fetched",
            extra={"question_id": question.id}
        )

        if question.question_type != "sql":

            logger.warning(
                "Non-SQL question used in SQLGradeView",
                extra={"question_id": question.id}
            )

            return Response(
                {"detail": "Question is not SQL."}, status=status.HTTP_400_BAD_REQUEST
            )

        sqlmeta = (
            SQLQuestion.objects.select_related("dataset")
            .filter(question=question)
            .first()
        )

        if not sqlmeta or not sqlmeta.dataset:

            logger.warning(
                "SQL metadata missing during grading",
                extra={"question_id": question.id}
            )

            return Response({"detail": "SQL metadata missing."}, status=status.HTTP_400_BAD_REQUEST)

        query = serializer.validated_data["query"]
        if not is_select_only(query):

            logger.warning(
                "Invalid SQL query attempted",
                extra={
                    "candidate_id": request.user.id,
                    "question_id": question.id,
                }
            )

            return Response({"detail": "Only SELECT queries allowed."}, status=status.HTTP_400_BAD_REQUEST)

        assessment = get_object_or_404(Assessment, pk=serializer.validated_data["assessment_id"])

        logger.info(
            "Assessment fetched for SQL grading",
            extra={"assessment_id": assessment.id}
        )

        candidate_assessment = CandidateAssessment.objects.filter(
            candidate=request.user, assessment=assessment
        ).first()

        testcases = list(SQLTestCase.objects.filter(question=question))

        logger.info(
            "SQL testcases fetched",
            extra={
                "question_id": question.id,
                "testcase_count": len(testcases),
            }
        )

        if not testcases:
            logger.warning(
                "No SQL testcases found, using placeholder",
                extra={"question_id": question.id}
            )

            class Placeholder:
                id = 0
                setup_sql = ""
                points = float(question.marks or 1.0)
                is_hidden = False

            testcases = [Placeholder()]

        schema = sqlmeta.dataset.schema_ddl or ""
        seed = sqlmeta.dataset.seed_sql or ""
        max_rows = sqlmeta.max_rows or 5000

        total_points = 0.0
        earned_points = 0.0
        per_case = []

        for tc in testcases:

            logger.info(
                "Running SQL testcase",
                extra={
                    "testcase_id": getattr(tc, "id", 0),
                    "question_id": question.id,
                }
            )

            points = float(getattr(tc, "points", 1.0))
            total_points += points

            setup_sql = getattr(tc, "setup_sql", "") or ""

            ref_script = build_sqlite_script(
                schema,
                seed,
                setup_sql,
                sqlmeta.reference_solution,
                max_rows
            )

            cand_script = build_sqlite_script(
                schema,
                seed,
                setup_sql,
                query,
                max_rows
            )

            ref_result = submit_to_judge0_sql(ref_script)
            cand_result = submit_to_judge0_sql(cand_script)

            ref_err = ref_result.get("stderr")
            cand_err = cand_result.get("stderr")

            if ref_err or cand_err:

                logger.error(
                    "SQL testcase execution failed",
                    extra={
                        "testcase_id": getattr(tc, "id", 0),
                        "error": (ref_err or cand_err)[:300],
                    }
                )

                per_case.append({
                    "id": getattr(tc, "id", 0),
                    "passed": False,
                    "points": points,
                    "received": 0,
                    "hidden": getattr(tc, "is_hidden", False),
                    "error": (ref_err or cand_err)[:300],
                })

                continue

            _, ref_rows = parse_rows(
                ref_result.get("stdout") or ""
            )

            _, cand_rows = parse_rows(
                cand_result.get("stdout") or ""
            )

            passed = rowset_equal(
                ref_rows,
                cand_rows,
                strict_order=bool(
                    sqlmeta.strict_column_order
                ),
                float_tol=float(
                    sqlmeta.float_tolerance or 0.0
                ),
            )

            received = points if passed else 0.0

            if passed:
                earned_points += received

            logger.info(
                "SQL testcase evaluated",
                extra={
                    "testcase_id": getattr(tc, "id", 0),
                    "passed": passed,
                    "received": received,
                }
            )

            per_case.append({
                "id": getattr(tc, "id", 0),
                "passed": passed,
                "points": points,
                "received": received,
                "hidden": getattr(tc, "is_hidden", False),
            })

        response_obj, _ = CandidateResponse.objects.get_or_create(
            candidate=request.user,
            assessment=assessment,
            question=question
        )

        response_obj.answer = query
        response_obj.marks_obtained = earned_points

        response_obj.is_correct = bool(
            total_points and earned_points == total_points
        )

        response_obj.save()

        logger.info(
            "Candidate SQL response saved",
            extra={
                "response_id": response_obj.id,
                "earned_points": earned_points,
            }
        )

        if candidate_assessment:

            candidate_assessment.calculate_score()

            logger.info(
                "Candidate assessment score recalculated",
                extra={
                    "candidate_assessment_id":
                        candidate_assessment.id,
                    "score":
                        candidate_assessment.score,
                }
            )

        return Response(
            {
                "total_points": total_points,
                "earned_points": earned_points,
                "results": per_case,
            }
        )


class SQLDatasetCreateView(APIView):
    permission_classes = [AdminPermission]

    def post(self, request, *args, **kwargs):

        logger.info(
            "SQL dataset creation started",
            extra={"user_id": request.user.id}
        )

        serializer = SQLDatasetSerializer(data=request.data)

        if not serializer.is_valid():

            logger.warning(
                "SQL dataset validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        dataset = serializer.save()

        logger.info(
            "SQL dataset created successfully",
            extra={
                "dataset_id": dataset.id,
                "dataset_name": dataset.name,
            }
        )

        return Response(
            SQLDatasetSerializer(dataset).data,
            status=status.HTTP_201_CREATED
        )


class TestEmailView(APIView):
    permission_classes = [AdminPermission]

    def post(self, request, *args, **kwargs):

        logger.info(
            "Test email API called",
            extra={"user_id": request.user.id}
        )

        email = request.data.get("email") or request.user.email

        logger.info(
            "Preparing test email",
            extra={"email": email}
        )

        otp = OTPVerification.generate_otp()

        logger.info(
            "Generated OTP for test email",
            extra={"email": email}
        )

        send_otp_email(email, otp, "test")

        logger.info(
            "Test email sent successfully",
            extra={"email": email}
        )

        return Response({
            "message": f"Test email sent to {email}."
        })


class TestCodeExecutionView(APIView):
    permission_classes = [AdminPermission]

    def post(self, request, *args, **kwargs):

        logger.info(
            "Test code execution API called",
            extra={"user_id": request.user.id}
        )

        code = (
            request.data.get("code")
            or "print('hello world')"
        )

        language = (
            request.data.get("language")
            or "python"
        )

        logger.info(
            "Executing test code",
            extra={
                "language": language,
                "code_length": len(code),
            }
        )

        try:

            result = execute_code_with_judge0(
                code,
                language
            )

            logger.info(
                "Code execution completed successfully",
                extra={"language": language}
            )

            return Response(result)

        except Exception as e:

            logger.exception(
                "Code execution failed",
                extra={
                    "language": language,
                    "error": str(e),
                }
            )

            return Response(
                {"detail": "Code execution failed"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class DownloadCandidateTemplateXLSXView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, *args, **kwargs):

        logger.info(
            "Download candidate XLSX template API called",
            extra={"user_id": request.user.id}
        )

        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Candidates"

        headers = ["username", "email", "first_name", "last_name", "phone"]
        sheet.append(headers)
        sheet.append(["john_doe", "john@example.com", "John", "Doe", "1234567890"])
        sheet.append(["jane_smith", "jane@example.com", "Jane", "Smith", "0987654321"])


        logger.info(
            "Candidate XLSX template prepared successfully"
        )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="candidate_import_template.xlsx"'
        workbook.save(response)

        logger.info(
            "Candidate XLSX template download response ready"
        )

        return response


class DownloadCandidateTemplateCSVView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, *args, **kwargs):

        logger.info(
            "Download candidate CSV template API called",
            extra={"user_id": request.user.id}
        )

        headers = ["username", "email", "first_name", "last_name", "phone"]
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="candidate_import_template.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerow(["john_doe", "john@example.com", "John", "Doe", "1234567890"])
        writer.writerow(["jane_smith", "jane@example.com", "Jane", "Smith", "0987654321"])


        logger.info(
            "Candidate CSV template prepared successfully"
        )

        return response


class DownloadQuestionTemplateXLSXView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, *args, **kwargs):

        logger.info(
            "Download question template XLSX API called",
            extra={"user_id": request.user.id}
        )

        workbook = Workbook()

        logger.info("Workbook initialized successfully")

        sheet = workbook.active
        sheet.title = "Questions"

        logger.info(
            "Worksheet created successfully",
            extra={"sheet_name": sheet.title}
        )

        headers = [
            "title",
            "question_type",
            "category",
            "difficulty",
            "marks",
            "option1",
            "option2",
            "option3",
            "option4",
            "option5",
            "correct_answer",
            "tags",
            "description",
            "sample_input",
            "sample_output",
            "coding_testcases",
            "sql_dataset",
            "sql_reference_solution",
            "sql_strict_column_order",
            "sql_float_tolerance",
            "sql_max_rows",
            "sql_testcases",
        ]

        sheet.append(headers)

        logger.info(
            "Headers added to question template",
            extra={"header_count": len(headers)}
        )

        sample_rows = [
            {
                "title": "What is Python?",
                "question_type": "mcq_single",
                "category": "Programming",
                "difficulty": "easy",
                "marks": 1,
                "option1": "Interpreted",
                "option2": "Compiled",
                "option3": "Both",
                "option4": "None",
                "correct_answer": "A",
                "tags": "programming,basic",
                "description": "Basic Python question",
            },
            {
                "title": "Sum two numbers",
                "question_type": "coding",
                "category": "Programming",
                "difficulty": "medium",
                "marks": 5,
                "tags": "coding,basic",
                "description": "Write a program that prints the sum of two integers provided on a single line.",
                "sample_input": "2 3",
                "sample_output": "5",
                "coding_testcases": '[{"input_data": "2 3", "expected_output": "5", "points": 1, "is_hidden": false}, {"input_data": "10 5", "expected_output": "15", "points": 2, "is_hidden": true}]',

            },
            {
                "title": "Top customers by revenue",
                "question_type": "sql",
                "category": "Analytics",
                "difficulty": "hard",
                "marks": 5,
                "tags": "sql,analytics",
                "description": (
                    "Return the customers with "
                    "the highest total order amount."
                ),
                "sql_dataset": "SampleSalesDataset",
                "sql_reference_solution": (
                    "SELECT customer_id, "
                    "SUM(total_amount) AS total_spent "
                    "FROM orders "
                    "GROUP BY customer_id "
                    "ORDER BY total_spent DESC "
                    "LIMIT 5;"
                ),
                "sql_strict_column_order": False,
                "sql_float_tolerance": 0.01,
                "sql_max_rows": 5000,
                "sql_testcases": '[{"setup_sql": "-- optional setup before evaluation", "points": 1, "is_hidden": false}]',
            },
        ]

        logger.info(
            "Sample rows prepared",
            extra={"sample_row_count": len(sample_rows)}
        )

        for sample in sample_rows:
            sheet.append([sample.get(column, "") for column in headers])


        logger.info(
            "Sample rows added to worksheet successfully"
        )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="question_import_template.xlsx"'

        logger.info(
            "HTTP response prepared for XLSX download"
        )

        workbook.save(response)

        logger.info(
            "Question template XLSX generated successfully"
        )

        return response

class DownloadQuestionTemplateCSVView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, *args, **kwargs):

        logger.info(
            "Download question template CSV API called",
            extra={"user_id": request.user.id}
        )

        headers = [
            "title",
            "question_type",
            "category",
            "difficulty",
            "marks",
            "option1",
            "option2",
            "option3",
            "option4",
            "option5",
            "correct_answer",
            "tags",
            "description",
            "sample_input",
            "sample_output",
            "coding_testcases",
            "sql_dataset",
            "sql_reference_solution",
            "sql_strict_column_order",
            "sql_float_tolerance",
            "sql_max_rows",
            "sql_testcases",
        ]

        logger.info(
            "CSV headers prepared",
            extra={"header_count": len(headers)}
        )

        sample_rows = [
            {
                "title": "What is Python?",
                "question_type": "mcq_single",
                "category": "Programming",
                "difficulty": "easy",
                "marks": 1,
                "option1": "Interpreted",
                "option2": "Compiled",
                "option3": "Both",
                "option4": "None",
                "correct_answer": "A",
                "tags": "programming,basic",
                "description": "Basic Python question",
            },
            {
                "title": "Sum two numbers",
                "question_type": "coding",
                "category": "Programming",
                "difficulty": "medium",
                "marks": 5,
                "tags": "coding,basic",
                "description": "Write a program that prints the sum of two integers provided on a single line.",
                "sample_input": "2 3",
                "sample_output": "5",
                "coding_testcases": '[{"input_data": "2 3", "expected_output": "5", "points": 1, "is_hidden": false}, {"input_data": "10 5", "expected_output": "15", "points": 2, "is_hidden": true}]',
            },
            {
                "title": "Top customers by revenue",
                "question_type": "sql",
                "category": "Analytics",
                "difficulty": "hard",
                "marks": 5,
                "tags": "sql,analytics",
                "description": "Return the customers with the highest total order amount.",
                "sql_dataset": "SampleSalesDataset",
                "sql_reference_solution": "SELECT customer_id, SUM(total_amount) AS total_spent FROM orders GROUP BY customer_id ORDER BY total_spent DESC LIMIT 5;",
                "sql_strict_column_order": False,
                "sql_float_tolerance": 0.01,
                "sql_max_rows": 5000,
                "sql_testcases": '[{"setup_sql": "-- optional setup before evaluation", "points": 1, "is_hidden": false}]',
            },
        ]

        logger.info(
            "Sample rows prepared",
            extra={"sample_row_count": len(sample_rows)}
        )

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="question_import_template.csv"'


        logger.info(
            "CSV response initialized successfully"
        )

        writer = csv.writer(response)
        writer.writerow(headers)
        
        logger.info(
            "CSV headers written successfully"
        )

        for sample in sample_rows:
            writer.writerow([sample.get(column, "") for column in headers])


        logger.info(
            "Sample rows written to CSV successfully",
            extra={"written_rows": len(sample_rows)}
        )

        logger.info(
            "Question template CSV generated successfully"
        )

        return response


class ForgotPasswordView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request, *args, **kwargs):

        logger.info("Forgot password API called")

        serializer = ForgotPasswordSerializer(data=request.data)


        if not serializer.is_valid():

            logger.warning(
                "Forgot password validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        contact = serializer.validated_data["contact"]

        logger.info(
            "Password reset requested",
            extra={"contact": contact}
        )

        if "@" in contact:
            user = UserModel.objects.filter(email__iexact=contact, role="candidate").first()
        else:
            user = UserModel.objects.filter(phone=contact, role="candidate").first()

        if not user:

            logger.warning(
                "User not found for password reset",
                extra={"contact": contact}
            )

            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)


        logger.info(
            "User found for password reset",
            extra={
                "user_id": user.id,
                "email": user.email,
            }
        )

        otp_code = OTPVerification.generate_otp()
        otp_record = OTPVerification.objects.create(
            phone=user.phone,
            email=user.email,
            otp_code=otp_code,
            otp_type="password_reset",
            temp_data={"user_id": user.id},
            expires_at=timezone.now() + timedelta(minutes=10),
        )

        logger.info(
            "Password reset OTP created",
            extra={
                "otp_id": otp_record.id,
                "user_id": user.id,
            }
        )

        send_otp_sms(user.phone, otp_code)

        logger.info(
            "Password reset OTP SMS sent",
            extra={"user_id": user.id}
        )

        if user.email:
            send_otp_email(user.email, otp_code, "password reset")


            logger.info(
                "Password reset OTP email sent",
                extra={
                    "user_id": user.id,
                    "email": user.email,
                }
            )

        return Response({"otp_id": otp_record.id, "message": "OTP sent."})


class VerifyResetOTPView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request, otp_id: int, *args, **kwargs):

        logger.info(
            "Verify reset OTP API called",
            extra={"otp_id": otp_id}
        )

        otp_record = get_object_or_404(
            OTPVerification, pk=otp_id, otp_type="password_reset", is_verified=False
        )

        logger.info(
            "OTP record fetched successfully",
            extra={"otp_id": otp_record.id}
        )

        if otp_record.is_expired():

            logger.warning(
                "Password reset OTP expired",
                extra={"otp_id": otp_record.id}
            )

            return Response({"detail": "OTP expired."}, status=status.HTTP_400_BAD_REQUEST)

        serializer = ResetPasswordOTPSerializer(data=request.data)


        if not serializer.is_valid():

            logger.warning(
                "Reset OTP validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        if serializer.validated_data["otp_code"] != otp_record.otp_code:


            logger.warning(
                "Invalid password reset OTP entered",
                extra={"otp_id": otp_record.id}
            )

            return Response({"detail": "Invalid OTP."}, status=status.HTTP_400_BAD_REQUEST)

        otp_record.is_verified = True
        otp_record.save(update_fields=["is_verified"])

        logger.info(
            "Password reset OTP verified successfully",
            extra={"otp_id": otp_record.id}
        )

        return Response({"message": "OTP verified."})


class ResetPasswordView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request, otp_id: int, *args, **kwargs):

        logger.info(
            "Reset password API called",
            extra={"otp_id": otp_id}
        )

        otp_record = get_object_or_404(
            OTPVerification, pk=otp_id, otp_type="password_reset", is_verified=True
        )

        logger.info(
            "Verified OTP record fetched successfully",
            extra={"otp_id": otp_record.id}
        )

        serializer = ResetPasswordSerializer(data=request.data)

        if not serializer.is_valid():

            logger.warning(
                "Reset password validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        user_id = otp_record.temp_data.get("user_id") if otp_record.temp_data else None

        logger.info(
            "Fetching user for password reset",
            extra={"user_id": user_id}
        )

        user = get_object_or_404(UserModel, pk=user_id)
        user.set_password(serializer.validated_data["new_password"])
        user.save()

        logger.info(
            "User password reset successfully",
            extra={
                "user_id": user.id,
                "email": user.email,
            }
        )

        otp_record.delete()

        logger.info(
            "Password reset OTP deleted successfully",
            extra={"otp_id": otp_id}
        )

        return Response({"message": "Password reset successful."})


class CandidateProfileView(APIView):
    # This endpoint only ever reads/writes the CALLER's own record
    # (serializer instance = request.user), so every method is safe for any
    # authenticated user. GET and PUT/PATCH are open to all roles — manager,
    # org_admin, super_admin and candidate can all view and edit their own
    # profile (name, details, photo). Previously PUT/PATCH was candidate-only,
    # which 403'd managers and other admins on profile save.
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):

        logger.info(
            "Profile fetch API called",
            extra={"user_id": request.user.id, "role": getattr(request.user, "role", None)},
        )

        response_data = CandidateSerializer(
            request.user
        ).data

        logger.info(
            "Profile fetched successfully",
            extra={"user_id": request.user.id},
        )

        return Response(response_data)

    def put(self, request, *args, **kwargs):

        logger.info(
            "Candidate profile update API called",
            extra={"candidate_id": request.user.id}
        )

        serializer = CandidateProfileSerializer(
            instance=request.user,
            data=request.data,
            partial=True,
            context={"request": request}
        )

        if not serializer.is_valid():

            logger.warning(
                "Candidate profile validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        candidate = serializer.save()

        logger.info(
            "Candidate profile updated successfully",
            extra={"candidate_id": candidate.id}
        )

        resume_file = request.FILES.get("resume")

        # ---------------------------------------------------------
        # SAME LOGIC AS VerifyRegistrationOTPAPI FOR RESUME UPLOAD
        # ---------------------------------------------------------
        if resume_file:

            logger.info(
                "Resume upload initiated",
                extra={
                    "candidate_id": candidate.id,
                    "resume_filename": resume_file.name,
                }
            )

            try:
                # Initialize S3
                s3_client = boto3.client(
                    "s3",
                    aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                    aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                    region_name=settings.AWS_S3_REGION_NAME,
                )

                logger.info(
                    "S3 client initialized successfully",
                    extra={"candidate_id": candidate.id}
                )

                # Generate file details
                candidate_email = (candidate.email or candidate.username).lower().replace(" ", "")
                extension = os.path.splitext(resume_file.name)[1].lower()
                unique_filename = f"resume_{uuid.uuid4().hex[:8]}{extension}"

                # Final S3 path
                s3_key = f"{candidate_email}/resumes/{unique_filename}"


                logger.info(
                    "Generated S3 key for resume upload",
                    extra={
                        "candidate_id": candidate.id,
                        "s3_key": s3_key,
                    }
                )

                # Upload file
                s3_client.upload_fileobj(
                    resume_file,
                    settings.AWS_STORAGE_BUCKET_NAME,
                    s3_key,
                    ExtraArgs={"ContentType": resume_file.content_type}
                )

                logger.info(
                    "Resume uploaded to S3 successfully",
                    extra={"candidate_id": candidate.id}
                )

                # Generate final URL
                final_url = (
                    f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3."
                    f"{settings.AWS_S3_REGION_NAME}.amazonaws.com/{s3_key}"
                )

                # Save to user profile
                candidate.resume_s3_url = final_url
                candidate.save(update_fields=["resume_s3_url"])

                logger.info(
                    "Candidate resume URL updated successfully",
                    extra={
                        "candidate_id": candidate.id,
                        "resume_url": final_url,
                    }
                )

            except Exception as e:

                logger.exception(
                    "Error uploading candidate resume",
                    extra={
                        "candidate_id": candidate.id,
                        "error": str(e),
                    }
                )

        # ---------------------------------------------------------
        # Avatar upload is now handled in serializer
        # ---------------------------------------------------------

        response_data = CandidateSerializer(candidate).data

        logger.info(
            "Candidate profile response prepared successfully",
            extra={"candidate_id": candidate.id}
        )

        return Response(response_data)


class ChangePasswordView(APIView):
    permission_classes = [CandidatePermission]

    def post(self, request, *args, **kwargs):

        logger.info(
            "Change password API called",
            extra={"candidate_id": request.user.id}
        )

        serializer = ChangePasswordSerializer(data=request.data)

        if not serializer.is_valid():

            logger.warning(
                "Change password validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        if not request.user.check_password(
            serializer.validated_data["current_password"]
        ):

            logger.warning(
                "Incorrect current password entered",
                extra={"candidate_id": request.user.id}
            )

            return Response(
                {"detail": "Current password incorrect."},
                status=status.HTTP_400_BAD_REQUEST
            )

        request.user.set_password(
            serializer.validated_data["new_password"]
        )

        request.user.save()

        logger.info(
            "Password updated successfully",
            extra={"candidate_id": request.user.id}
        )

        update_session_auth_hash(
            request,
            request.user
        )

        logger.info(
            "Session auth hash updated successfully",
            extra={"candidate_id": request.user.id}
        )

        return Response({
            "message": "Password updated."
        })

class BulkUploadView(APIView):
    # Bulk upload is Super-Admin only.
    permission_classes = [SuperAdminPermission]

    def post(self, request, *args, **kwargs):

        logger.info(
            "Bulk upload API called",
            extra={"user_id": request.user.id}
        )

        serializer = BulkUploadSerializer(data=request.data)

        if not serializer.is_valid():

            logger.warning(
                "Bulk upload validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        model_type = serializer.validated_data["model_type"]
        upload = serializer.validated_data["file"]

        logger.info(
            "Bulk upload request received",
            extra={
                "model_type": model_type,
                "filename": upload.name,
            }
        )

        try:

            rows = _load_rows_from_upload(upload)

            logger.info(
                "Uploaded file parsed successfully",
                extra={
                    "model_type": model_type,
                    "total_rows": len(rows),
                }
            )

        except Exception as exc:

            logger.exception(
                "Failed to parse uploaded file",
                extra={
                    "filename": upload.name,
                    "error": str(exc),
                }
            )

            return Response(
                {"detail": f"Unable to parse file: {exc}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = 0
        errors = []

        if model_type == "candidate":
            logger.info("Starting bulk candidate upload")
            for index, row in enumerate(rows):
                try:
                    password = generate_password()
                    user = UserModel.objects.create_user(
                        username=row["username"],
                        email=row.get("email", ""),
                        password=password,
                        first_name=row.get("first_name", ""),
                        last_name=row.get("last_name", ""),
                        phone=row.get("phone", ""),
                        role="candidate",
                    )

                    logger.info(
                        "Candidate created successfully",
                        extra={
                            "candidate_id": user.id,
                            "username": user.username,
                        }
                    )

                    if user.email:
                        send_candidate_credentials(
                            user,
                            password
                        )

                        logger.info(
                            "Candidate credentials email sent",
                            extra={
                                "candidate_id": user.id,
                                "email": user.email,
                            }
                        )

                    created += 1

                except Exception as exc:

                    logger.exception(
                        "Error creating candidate during bulk upload",
                        extra={
                            "row_number": index + 2,
                            "error": str(exc),
                        }
                    )

                    errors.append(
                        f"Row {index + 2}: {exc}"
                    )

        else:

            logger.info("Starting bulk question upload")

            for index, row in enumerate(rows):

                try:
                    category, _ = (
                        Category.objects.get_or_create(
                            name=row["category"],
                            defaults={
                                "description":
                                    (
                                        "Imported category: "
                                        f"{row['category']}"
                                    )
                            },
                        )
                    )

                    logger.info(
                        "Category fetched/created successfully",
                        extra={
                            "category_name":
                                category.name,
                        }
                    )

                    question = Question.objects.create(
                        title=row["title"],
                        question_type=row["question_type"],
                        category=category,
                        difficulty=row.get(
                            "difficulty",
                            "medium"
                        ),
                        marks=row.get("marks", 1),
                        description=row.get(
                            "description",
                            ""
                        ),
                        option1=row.get("option1", ""),
                        option2=row.get("option2", ""),
                        option3=row.get("option3", ""),
                        option4=row.get("option4", ""),
                        correct_answer=row.get(
                            "correct_answer",
                            ""
                        ),
                        created_by=request.user,
                        tags=row.get("tags", ""),
                    )

                    logger.info(
                        "Question created successfully",
                        extra={
                            "question_id": question.id,
                            "title": question.title,
                        }
                    )

                    created += 1

                except Exception as exc:

                    logger.exception(
                        "Error creating question during bulk upload",
                        extra={
                            "row_number": index + 2,
                            "error": str(exc),
                        }
                    )

                    errors.append(
                        f"Row {index + 2}: {exc}"
                    )

        logger.info(
            "Bulk upload completed",
            extra={
                "model_type": model_type,
                "created_count": created,
                "error_count": len(errors),
            }
        )

        return Response({
            "created": created,
            "errors": errors[:50]
        })


class ExportCandidatesView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, *args, **kwargs):

        logger.info(
            "Export candidates API called",
            extra={"user_id": request.user.id}
        )

        excel_file = export_candidates_to_excel()

        logger.info(
            "Candidates Excel file generated successfully"
        )

        response = HttpResponse(
            excel_file.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

        response["Content-Disposition"] = (
            'attachment; filename="candidates.xlsx"'
        )

        logger.info(
            "Candidates export response prepared successfully"
        )

        return response


class ExportQuestionsView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, *args, **kwargs):

        logger.info(
            "Export questions API called",
            extra={"user_id": request.user.id}
        )

        excel_file = export_questions_to_excel()

        logger.info(
            "Questions Excel file generated successfully"
        )

        response = HttpResponse(
            excel_file.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

        response["Content-Disposition"] = (
            'attachment; filename="questions.xlsx"'
        )

        logger.info(
            "Questions export response prepared successfully"
        )

        return response


class ExportAssessmentResultsView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, assessment_id: int, *args, **kwargs):

        logger.info(
            "Export assessment results API called",
            extra={
                "assessment_id": assessment_id,
                "user_id": request.user.id,
            }
        )

        excel_file = export_results_to_excel(
            assessment_id
        )

        logger.info(
            "Assessment results Excel file generated successfully",
            extra={"assessment_id": assessment_id}
        )

        response = HttpResponse(
            excel_file.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

        response["Content-Disposition"] = (
            'attachment; '
            'filename="assessment_results.xlsx"'
        )

        logger.info(
            "Assessment results export response prepared successfully",
            extra={"assessment_id": assessment_id}
        )

        return response


class PrintAssessmentResultsView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, assessment_id: int, *args, **kwargs):

        logger.info(
            "Print assessment results API called",
            extra={
                "assessment_id": assessment_id,
                "user_id": request.user.id,
            }
        )

        assessment = get_object_or_404(
            Assessment,
            pk=assessment_id
        )

        logger.info(
            "Assessment fetched successfully",
            extra={
                "assessment_id": assessment.id,
                "assessment_title": assessment.title,
            }
        )

        results = CandidateAssessment.objects.filter(
            assessment=assessment
        )

        logger.info(
            "Candidate assessment results fetched",
            extra={
                "assessment_id": assessment.id,
                "results_count": results.count(),
            }
        )

        payload = [
            {
                "candidate":
                    (
                        candidate_assessment.candidate
                        .get_full_name()
                        or
                        candidate_assessment.candidate
                        .username
                    ),
                "score":
                    candidate_assessment.score,
                "total_marks":
                    candidate_assessment.total_marks,
                "percentage":
                    candidate_assessment.percentage,
                "status":
                    candidate_assessment.status,
            }
            for candidate_assessment in results
        ]

        logger.info(
            "Assessment results payload prepared successfully",
            extra={
                "assessment_id": assessment.id,
                "payload_count": len(payload),
            }
        )

        return Response({
            "assessment": assessment.title,
            "results": payload
        })


class TakeAssessmentView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, assessment_id: int = None, candidate_assessment_id: int = None, *args, **kwargs):

        logger.info(
            "Take assessment API called",
            extra={
                "candidate_id": request.user.id,
                "assessment_id": assessment_id,
                "candidate_assessment_id": candidate_assessment_id,
            }
        )

        if candidate_assessment_id:
            logger.info(
                "Fetching candidate assessment using candidate_assessment_id",
                extra={"candidate_assessment_id": candidate_assessment_id}
            )

            candidate_assessment = get_object_or_404(
            CandidateAssessment, id=candidate_assessment_id,candidate=request.user
        )
        else:
            logger.info(
                "Fetching candidate assessment using assessment_id",
                extra={"assessment_id": assessment_id}
            )

            candidate_assessment = get_object_or_404(
            CandidateAssessment, assessment_id=assessment_id, candidate=request.user
        )

        logger.info(
            "Candidate assessment fetched successfully",
            extra={
                "candidate_assessment_id": candidate_assessment.id,
                "status": candidate_assessment.status,
            }
        )

    # def get(self, request, assessment_id: int, *args, **kwargs):
    #     candidate_assessment = get_object_or_404(
    #         CandidateAssessment, assessment_id=assessment_id, candidate=request.user
    #     )
        assessment = candidate_assessment.assessment

        logger.info(
            "Assessment fetched successfully",
            extra={
                "assessment_id": assessment.id,
                "assessment_title": assessment.title,
            }
        )

        if candidate_assessment.status == "completed":

            logger.warning(
                "Assessment already submitted",
                extra={
                    "candidate_assessment_id": candidate_assessment.id
                }
            )

            return Response(
                {"detail": "Assessment already submitted."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if assessment.is_upcoming():

            logger.warning(
                "Assessment has not started yet",
                extra={"assessment_id": assessment.id}
            )

            return Response(
                {"detail": "Assessment has not started yet."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if assessment.is_expired():

            logger.warning(
                "Assessment expired",
                extra={"assessment_id": assessment.id}
            )

            candidate_assessment.status = "expired"
            candidate_assessment.save(update_fields=["status"])

            logger.info(
                "Candidate assessment marked as expired",
                extra={
                    "candidate_assessment_id": candidate_assessment.id
                }
            )

            return Response(
                {"detail": "Assessment has expired."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if candidate_assessment.status == "assigned":

            # Check subscription limits for individual users
            if getattr(request.user, 'is_individual', False):
                from core.utils import check_subscription_limits
                allowed, message = check_subscription_limits(request.user, 'assessment')
                if not allowed:
                    return Response(
                        {"detail": message},
                        status=status.HTTP_403_FORBIDDEN,
                    )

            logger.info(
                "Starting assessment",
                extra={
                    "candidate_assessment_id": candidate_assessment.id
                }
            )

            candidate_assessment.status = "in_progress"
            candidate_assessment.start_time = timezone.now()
            candidate_assessment.end_time = candidate_assessment.start_time + timedelta(
                minutes=assessment.duration
            )

            candidate_assessment.save(update_fields=["status", "start_time", "end_time"])

            # Track usage for subscription enforcement
            from core.usage_tracking import increment_assessment_usage
            increment_assessment_usage(request.user)

            logger.info(
                "Assessment moved to in_progress",
                extra={
                    "candidate_assessment_id": candidate_assessment.id,
                    "start_time": str(candidate_assessment.start_time),
                    "end_time": str(candidate_assessment.end_time),
                }
            )

        elif not candidate_assessment.end_time:
            if not candidate_assessment.start_time:
                candidate_assessment.start_time = timezone.now()
                logger.warning(
                    "Candidate assessment missing start_time; setting it now",
                    extra={"candidate_assessment_id": candidate_assessment.id},
                )

            candidate_assessment.end_time = candidate_assessment.start_time + timedelta(
                minutes=assessment.duration
            )

            candidate_assessment.save(update_fields=["end_time"])

            logger.info(
                "End time updated for candidate assessment",
                extra={
                    "candidate_assessment_id": candidate_assessment.id,
                    "end_time": str(candidate_assessment.end_time),
                }
            )

        assessment_questions = AssessmentQuestion.objects.filter(
            assessment=assessment
        ).select_related("question")

        logger.info(
            "Assessment questions fetched",
            extra={
                "assessment_id": assessment.id,
                "question_count": assessment_questions.count(),
            }
        )

        if assessment.shuffle_questions:

            logger.info(
                "Shuffling assessment questions",
                extra={"assessment_id": assessment.id}
            )

            assessment_questions = assessment_questions.order_by("?")

        assessment_questions = list(assessment_questions)

        question_ids = [aq.question_id for aq in assessment_questions]

        existing_question_ids = set(
            CandidateResponse.objects.filter(
                candidate=request.user,
                assessment=assessment,
                question_id__in=question_ids,
            ).values_list("question_id", flat=True)
        )

        logger.info(
            "Fetched existing responses",
            extra={
                "existing_response_count": len(existing_question_ids)
            }
        )

        missing_responses = [
            CandidateResponse(
                candidate=request.user,
                assessment=assessment,
                question=aq.question,
                answer="",
                code_language="",
            )
            for aq in assessment_questions
            if aq.question_id not in existing_question_ids
        ]

        if missing_responses:

            CandidateResponse.objects.bulk_create(missing_responses)

            logger.info(
                "Missing responses created",
                extra={
                    "created_response_count": len(missing_responses)
                }
            )

        responses = {
            response.question_id: response
            for response in CandidateResponse.objects.filter(
                candidate=request.user, assessment=assessment
            )
        }

        logger.info(
            "Responses mapped successfully",
            extra={
                "response_count": len(responses)
            }
        )

        questions_payload = []

        for aq in assessment_questions:
            question = aq.question
            resp = responses.get(question.id)

            questions_payload.append(
                {
                    "question_id": question.id,
                    "title": question.title,
                    "description": question.description,
                    "question_type": question.question_type,
                    "marks": question.marks,
                    "options": [
                        question.option1,
                        question.option2,
                        question.option3,
                        question.option4,
                        question.option5,
                    ],
                    "sample_input": question.sample_input,
                    "sample_output": question.sample_output,
                    "answer": resp.answer if resp else "",
                    "code_language": resp.code_language if resp else "",
                }
            )

        logger.info(
            "Questions payload prepared successfully",
            extra={
                "payload_count": len(questions_payload)
            }
        )

        return Response(
            {
                "candidate_assessment": {
                    "id": candidate_assessment.id,
                    "assessment_id": assessment.id,   
                    "status": candidate_assessment.status,
                    "start_time": candidate_assessment.start_time,
                    "end_time": candidate_assessment.end_time,
                    "duration_minutes": assessment.duration,
                    "assessment_title": assessment.title,
                },
                "questions": questions_payload,
            }
        )


class SubmitAssessmentView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, assessment_id: int = None, candidate_assessment_id: int = None, *args, **kwargs):

        logger.info(
            "Submit assessment API called",
            extra={
                "candidate_id": request.user.id,
                "assessment_id": assessment_id,
                "candidate_assessment_id": candidate_assessment_id,
            }
        )

        if candidate_assessment_id:

            logger.info(
                "Fetching candidate assessment using candidate_assessment_id",
                extra={"candidate_assessment_id": candidate_assessment_id}
            )

            candidate_assessment = get_object_or_404(
                CandidateAssessment, id=candidate_assessment_id, candidate=request.user
            )
        else:

            logger.info(
                "Fetching candidate assessment using assessment_id",
                extra={"assessment_id": assessment_id}
            )

            candidate_assessment = get_object_or_404(
                CandidateAssessment, assessment_id=assessment_id, candidate=request.user
            )

        logger.info(
            "Candidate assessment fetched successfully",
            extra={
                "candidate_assessment_id": candidate_assessment.id,
                "status": candidate_assessment.status,
            }
        )

    # def post(self, request, assessment_id: int, *args, **kwargs):
    #     candidate_assessment = get_object_or_404(
    #         CandidateAssessment, assessment_id=assessment_id, candidate=request.user
    #     )

        if candidate_assessment.status != "completed":

            logger.info(
                "Submitting assessment",
                extra={
                    "candidate_assessment_id": candidate_assessment.id
                }
            )

            candidate_assessment.status = "completed"
            candidate_assessment.end_time = timezone.now()

            candidate_assessment.save(update_fields=["status", "end_time"])

            logger.info(
                "Assessment marked as completed",
                extra={
                    "candidate_assessment_id": candidate_assessment.id,
                    "end_time": str(candidate_assessment.end_time),
                }
            )

            candidate_assessment.evaluate_all_mcqs()

            logger.info(
                "All MCQ questions evaluated",
                extra={
                    "candidate_assessment_id": candidate_assessment.id
                }
            )

            candidate_assessment.calculate_score()

            logger.info(
                "Assessment score calculated",
                extra={
                    "candidate_assessment_id": candidate_assessment.id,
                    "score": candidate_assessment.score,
                    "percentage": candidate_assessment.percentage,
                }
            )

            try:

                send_assessment_completion_email(candidate_assessment)

                logger.info(
                    "Assessment completion email sent successfully",
                    extra={
                        "candidate_assessment_id": candidate_assessment.id
                    }
                )

            except Exception as exc:

                logger.exception(
                    "Failed to send assessment completion email",
                    extra={
                        "candidate_assessment_id": candidate_assessment.id,
                        "error": str(exc),
                    },
                )

        else:

            logger.warning(
                "Assessment already submitted",
                extra={
                    "candidate_assessment_id": candidate_assessment.id
                }
            )

        return Response(
            {
                "score": candidate_assessment.score,
                "total_marks": candidate_assessment.total_marks,
                "percentage": candidate_assessment.percentage,
            }
        )


# class SubmitAndLogoutView(APIView):
#     permission_classes = [CandidatePermission]

#     def post(self, request, assessment_id: int, *args, **kwargs):
#         response = SubmitAssessmentView().post(request, assessment_id, *args, **kwargs)
#         logout(request)
#         payload = response.data
#         payload.update({"message": "Assessment submitted and user logged out."})
#         return Response(payload)


class CandidateAssessmentResultView(APIView):
    permission_classes = [CandidatePermission]

    def get(self, request, assessment_id: int, *args, **kwargs):

        logger.info(
            "Candidate assessment result API called",
            extra={
                "candidate_id": request.user.id,
                "assessment_id": assessment_id,
            }
        )

        assessment = get_object_or_404(Assessment, pk=assessment_id)

        logger.info(
            "Assessment fetched successfully",
            extra={
                "assessment_id": assessment.id,
                "assessment_title": assessment.title,
            }
        )

        candidate_assessment = CandidateAssessment.objects.filter(
            candidate=request.user, assessment=assessment
        ).first()

        if not candidate_assessment:

            logger.warning(
                "Candidate assessment result not found",
                extra={
                    "candidate_id": request.user.id,
                    "assessment_id": assessment.id,
                }
            )

            return Response(
                {"detail": "Result not available."}, status=status.HTTP_404_NOT_FOUND
            )

        logger.info(
            "Candidate assessment fetched successfully",
            extra={
                "candidate_assessment_id": candidate_assessment.id,
                "status": candidate_assessment.status,
            }
        )

        # Ensure score is up-to-date
        candidate_assessment.calculate_score()

        logger.info(
            "Candidate assessment score recalculated",
            extra={
                "candidate_assessment_id": candidate_assessment.id,
                "score": candidate_assessment.score,
                "percentage": candidate_assessment.percentage,
            }
        )

        # Fetch all candidate responses for this assessment
        responses_qs = CandidateResponse.objects.filter(
            candidate=request.user, assessment=assessment
        ).select_related("question")

        logger.info(
            "Candidate responses fetched",
            extra={
                "candidate_id": request.user.id,
                "response_count": responses_qs.count(),
            }
        )

        responses_dict = {r.question_id: r for r in responses_qs}

        response_payload = []
        total_questions = assessment.questions.count()
        attempted = 0

        for question in assessment.questions.all():

            resp = responses_dict.get(question.id)

            if resp:

                response_data = {
                    "question_id": resp.question_id,
                    "question_title": resp.question.title,
                    "question_description": resp.question.description,
                    "question_marks": resp.question.marks,
                    "question_type": resp.question.question_type,
                    "answer": resp.answer,
                    "is_correct": resp.is_correct,
                    "marks_obtained": resp.marks_obtained,
                }

                if resp.answer not in (None, ""):
                    attempted += 1

            else:

                logger.info(
                    "Question not attempted",
                    extra={
                        "question_id": question.id,
                        "candidate_id": request.user.id,
                    }
                )

                # Placeholder for unattempted question
                response_data = {
                    "question_id": question.id,
                    "question_title": question.title,
                    "question_description": question.description,
                    "question_marks": question.marks,
                    "question_type": question.question_type,
                    "answer": None,
                    "is_correct": False,
                    "marks_obtained": 0,
                }

            response_payload.append(response_data)

        logger.info(
            "Response payload prepared successfully",
            extra={
                "candidate_id": request.user.id,
                "response_count": len(response_payload),
            }
        )

        feedback = Feedback.objects.filter(
            candidate=request.user, assessment=assessment
        ).first()

        if feedback:

            logger.info(
                "Feedback fetched successfully",
                extra={
                    "candidate_id": request.user.id,
                    "assessment_id": assessment.id,
                }
            )

        return Response(
            {
                "assessment": AssessmentSerializer(assessment).data,
                "candidate_assessment": CandidateAssessmentSerializer(
                    candidate_assessment
                ).data,
                "responses": response_payload,
                "stats": {
                    "attempted": attempted,
                    "not_attempted": max(0, total_questions - attempted),
                    "correct": sum(1 for r in response_payload if r["is_correct"]),
                    "incorrect": sum(1 for r in response_payload if r["answer"] not in (None, "") and not r["is_correct"]),
                },
                "feedback": {
                    "rating": feedback.rating,
                    "comments": feedback.comments,
                }
                if feedback
                else None,
            }
        )

    def post(self, request, assessment_id: int, *args, **kwargs):

        logger.info(
            "Candidate feedback submission API called",
            extra={
                "candidate_id": request.user.id,
                "assessment_id": assessment_id,
            }
        )

        assessment = get_object_or_404(Assessment, pk=assessment_id)

        logger.info(
            "Assessment fetched successfully for feedback",
            extra={
                "assessment_id": assessment.id,
                "assessment_title": assessment.title,
            }
        )

        candidate_assessment = CandidateAssessment.objects.filter(
            candidate=request.user, assessment=assessment
        ).first()

        if not candidate_assessment:

            logger.warning(
                "Candidate assessment not found for feedback",
                extra={
                    "candidate_id": request.user.id,
                    "assessment_id": assessment.id,
                }
            )

            return Response({"detail": "Assessment not found."}, status=status.HTTP_404_NOT_FOUND)

        if Feedback.objects.filter(candidate=request.user, assessment=assessment).exists():

            logger.warning(
                "Feedback already submitted",
                extra={
                    "candidate_id": request.user.id,
                    "assessment_id": assessment.id,
                }
            )

            return Response(
                {"detail": "Feedback already submitted."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rating = request.data.get("rating")
        comments = request.data.get("comments", "")

        logger.info(
            "Feedback data received",
            extra={
                "candidate_id": request.user.id,
                "rating": rating,
            }
        )

        try:
            rating = int(rating)

            if rating < 1 or rating > 5:
                raise ValueError

        except (TypeError, ValueError):

            logger.warning(
                "Invalid feedback rating provided",
                extra={
                    "candidate_id": request.user.id,
                    "rating": rating,
                }
            )

            return Response(
                {"detail": "Rating must be between 1 and 5."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        Feedback.objects.create(
            candidate=request.user, assessment=assessment, rating=rating, comments=comments
        )

        logger.info(
            "Feedback submitted successfully",
            extra={
                "candidate_id": request.user.id,
                "assessment_id": assessment.id,
                "rating": rating,
            }
        )

        return Response({"message": "Feedback submitted."}, status=status.HTTP_201_CREATED)


class CandidateResultByCandidateAssessmentView(APIView):
    """
    Candidate-facing result endpoint that accepts candidate_assessment_id.
    This allows the frontend to use the same ID throughout the assessment flow
    (from /take/ to submission to results) without needing to track assessment_id separately.
    """
    permission_classes = [CandidatePermission]

    def get(self, request, candidate_assessment_id: int, *args, **kwargs):

        logger.info(
            "Candidate result by candidate assessment API called",
            extra={
                "candidate_id": request.user.id,
                "candidate_assessment_id": candidate_assessment_id,
            }
        )

        candidate_assessment = get_object_or_404(
            CandidateAssessment, pk=candidate_assessment_id, candidate=request.user
        )

        logger.info(
            "Candidate assessment fetched successfully",
            extra={
                "candidate_assessment_id": candidate_assessment.id,
                "status": candidate_assessment.status,
            }
        )

        assessment = candidate_assessment.assessment

        logger.info(
            "Assessment fetched successfully",
            extra={
                "assessment_id": assessment.id,
                "assessment_title": assessment.title,
            }
        )

        # Ensure score is up-to-date
        candidate_assessment.calculate_score()

        logger.info(
            "Candidate assessment score recalculated",
            extra={
                "candidate_assessment_id": candidate_assessment.id,
                "score": candidate_assessment.score,
                "percentage": candidate_assessment.percentage,
            }
        )

        # Fetch all candidate responses for this assessment
        responses_qs = CandidateResponse.objects.filter(
            candidate=request.user, assessment=assessment
        ).select_related("question")

        logger.info(
            "Candidate responses fetched successfully",
            extra={
                "candidate_id": request.user.id,
                "response_count": responses_qs.count(),
            }
        )

        responses_dict = {r.question_id: r for r in responses_qs}

        response_payload = []
        total_questions = assessment.questions.count()
        attempted = 0

        for question in assessment.questions.all():

            resp = responses_dict.get(question.id)

            if resp:

                response_data = {
                    "question_id": resp.question_id,
                    "question_title": resp.question.title,
                    "question_description": resp.question.description,
                    "question_marks": resp.question.marks,
                    "question_type": resp.question.question_type,
                    "answer": resp.answer,
                    "is_correct": resp.is_correct,
                    "marks_obtained": resp.marks_obtained,
                    "correct_answer": resp.question.correct_answer or "",
                    "correct_answer_text": resp.question.correct_answer or "",
                    "answer_text": resp.answer or "",
                    "question_options": [
                        {"label": chr(65 + i), "value": opt}
                        for i, opt in enumerate([
                            resp.question.option1,
                            resp.question.option2,
                            resp.question.option3,
                            resp.question.option4,
                            resp.question.option5,
                        ])
                        if opt and opt.strip()
                    ],
                }

                if resp.answer not in (None, ""):
                    attempted += 1

            else:

                logger.info(
                    "Question not attempted",
                    extra={
                        "question_id": question.id,
                        "candidate_id": request.user.id,
                    }
                )

                response_data = {
                    "question_id": question.id,
                    "question_title": question.title,
                    "question_description": question.description,
                    "question_marks": question.marks,
                    "question_type": question.question_type,
                    "answer": None,
                    "is_correct": False,
                    "marks_obtained": 0,
                    "correct_answer": "",
                    "correct_answer_text": "",
                    "answer_text": "",
                    "question_options": [],
                }

            response_payload.append(response_data)

        logger.info(
            "Response payload prepared successfully",
            extra={
                "candidate_id": request.user.id,
                "response_count": len(response_payload),
            }
        )

        feedback = Feedback.objects.filter(
            candidate=request.user, assessment=assessment
        ).first()

        if feedback:

            logger.info(
                "Feedback fetched successfully",
                extra={
                    "candidate_id": request.user.id,
                    "assessment_id": assessment.id,
                }
            )

        return Response(
            {
                "assessment": AssessmentSerializer(assessment).data,
                "candidate_assessment": CandidateAssessmentSerializer(
                    candidate_assessment
                ).data,
                "responses": response_payload,
                "stats": {
                    "attempted": attempted,
                    "not_attempted": max(0, total_questions - attempted),
                    "correct": sum(1 for r in response_payload if r["is_correct"]),
                    "incorrect": sum(
                        1
                        for r in response_payload
                        if r["answer"] not in (None, "") and not r["is_correct"]
                    ),
                },
                "feedback": {
                    "rating": feedback.rating,
                    "comments": feedback.comments,
                }
                if feedback
                else None,
            }
        )

    def post(self, request, candidate_assessment_id: int, *args, **kwargs):

        logger.info(
            "Candidate feedback submission by candidate assessment API called",
            extra={
                "candidate_id": request.user.id,
                "candidate_assessment_id": candidate_assessment_id,
            }
        )

        candidate_assessment = get_object_or_404(
            CandidateAssessment, pk=candidate_assessment_id, candidate=request.user
        )

        logger.info(
            "Candidate assessment fetched for feedback",
            extra={
                "candidate_assessment_id": candidate_assessment.id,
                "status": candidate_assessment.status,
            }
        )

        assessment = candidate_assessment.assessment

        if Feedback.objects.filter(candidate=request.user, assessment=assessment).exists():

            logger.warning(
                "Feedback already submitted",
                extra={
                    "candidate_id": request.user.id,
                    "assessment_id": assessment.id,
                }
            )

            return Response(
                {"detail": "Feedback already submitted."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rating = request.data.get("rating")
        comments = request.data.get("comments", "")

        logger.info(
            "Feedback data received",
            extra={
                "candidate_id": request.user.id,
                "rating": rating,
            }
        )

        try:
            rating = int(rating)

            if rating < 1 or rating > 5:
                raise ValueError

        except (TypeError, ValueError):

            logger.warning(
                "Invalid feedback rating provided",
                extra={
                    "candidate_id": request.user.id,
                    "rating": rating,
                }
            )

            return Response(
                {"detail": "Rating must be between 1 and 5."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        Feedback.objects.create(
            candidate=request.user, assessment=assessment, rating=rating, comments=comments
        )

        logger.info(
            "Feedback submitted successfully",
            extra={
                "candidate_id": request.user.id,
                "assessment_id": assessment.id,
                "rating": rating,
            }
        )

        return Response({"message": "Feedback submitted."}, status=status.HTTP_201_CREATED)


class CategoryListCreateView(APIView):

    def get(self, request):

        logger.info(
            "Category list API called",
            extra={"user_id": request.user.id}
        )

        categories = Category.objects.all()

        # logger.info(
        #     f"Categories fetched successfully. Count: {count}",
        #     extra={"category_count": count}
        # )
        serializer = CategorySerializer(categories, many=True)

        logger.info(
            "Category list serialized successfully"
        )

        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):

        logger.info(
            "Category create API called",
            extra={"user_id": request.user.id}
        )

        serializer = CategorySerializer(data=request.data)

        if not serializer.is_valid():

            logger.warning(
                "Category create validation failed",
                extra={"errors": serializer.errors}
            )

            serializer.is_valid(raise_exception=True)

        category = serializer.save()
        
        logger.info(
            f"Category created successfully. ID: {category.id}, Name: {category.name}",
            extra={
                "category_id": category.id,
                "category_name": category.name,
            }
        )

        return Response(serializer.data, status=status.HTTP_201_CREATED)



class CategoryDetailView(APIView):

    def get_object(self, category_id):

        logger.info(
            "Fetching category object",
            extra={"category_id": category_id}
        )

        try:

            category = Category.objects.get(id=category_id)

            logger.info(
                "Category fetched successfully",
                extra={
                    "category_id": category.id,
                    "category_name": category.name,
                }
            )

            return category

        except Category.DoesNotExist:

            logger.warning(
                "Category not found",
                extra={"category_id": category_id}
            )

            return None

    def get(self, request, category_id):

        logger.info(
            "Category detail API called",
            extra={"category_id": category_id}
        )

        category = self.get_object(category_id)

        if not category:

            logger.warning(
                "Category detail fetch failed",
                extra={"category_id": category_id}
            )

            return Response({"message": "Category not found"}, status=404)

        serializer = CategorySerializer(category)

        logger.info(
            "Category detail serialized successfully",
            extra={"category_id": category.id}
        )

        return Response(serializer.data)

    def patch(self, request, category_id):

        logger.info(
            "Category update API called",
            extra={"category_id": category_id}
        )

        category = self.get_object(category_id)

        if not category:

            logger.warning(
                "Category update failed - category not found",
                extra={"category_id": category_id}
            )

            return Response({"message": "Category not found"}, status=404)

        serializer = CategorySerializer(category, data=request.data, partial=True)

        if not serializer.is_valid():

            logger.warning(
                "Category update validation failed",
                extra={
                    "category_id": category_id,
                    "errors": serializer.errors,
                }
            )

            serializer.is_valid(raise_exception=True)

        updated_category = serializer.save()

        logger.info(
            "Category updated successfully",
            extra={
                "category_id": updated_category.id,
                "category_name": updated_category.name,
            }
        )

        return Response(serializer.data)

    def delete(self, request, category_id):

        logger.info(
            "Category delete API called",
            extra={"category_id": category_id}
        )

        category = self.get_object(category_id)

        if not category:

            logger.warning(
                "Category delete failed - category not found",
                extra={"category_id": category_id}
            )

            return Response({"message": "Category not found"}, status=404)

        category.delete()

        logger.info(
            "Category deleted successfully",
            extra={
                "category_id": category_id,
                "category_name": category.name,
            }
        )

        return Response({"message": "Category deleted successfully"}, status=204)




class SQLDatasetListView(APIView):
    permission_classes = [AdminPermission]   # optional

    def get(self, request):

        logger.info(
            "SQL dataset list API called",
            extra={"user_id": request.user.id}
        )

        datasets = SQLDataset.objects.all().order_by("-created_at")

        logger.info(
            "SQL datasets fetched successfully",
            extra={"dataset_count": datasets.count()}
        )

        serializer = SQLDatasetSerializer(datasets, many=True)

        logger.info(
            "SQL dataset list serialized successfully"
        )

        return Response(serializer.data, status=200)


class SQLDatasetDetailView(APIView):
    permission_classes = [AdminPermission]

    def get(self, request, pk):

        logger.info(
            "SQL dataset detail API called",
            extra={
                "dataset_id": pk,
                "user_id": request.user.id,
            }
        )

        try:

            dataset = SQLDataset.objects.get(pk=pk)

            logger.info(
                "SQL dataset fetched successfully",
                extra={
                    "dataset_id": dataset.id,
                    "dataset_name": dataset.name,
                }
            )

        except SQLDataset.DoesNotExist:

            logger.warning(
                "SQL dataset not found",
                extra={"dataset_id": pk}
            )

            return Response({"error": "Dataset not found"}, status=404)

        serializer = SQLDatasetSerializer(dataset)

        logger.info(
            "SQL dataset serialized successfully",
            extra={"dataset_id": dataset.id}
        )

        return Response(serializer.data, status=200)


# ==========================================
# CANDIDATE ASSESSMENT VIEWS
# ==========================================

class CandidateAssignedAssessmentsView(APIView):
    """
    Get all assigned assessments for the candidate.
    Includes assessments that are assigned or in_progress.
    """
    permission_classes = [CandidatePermission]

    def get(self, request, *args, **kwargs):

        logger.info(
            "Candidate assigned assessments API called",
            extra={"candidate_id": request.user.id}
        )

        now = timezone.now()
        
        from organization.context import current_organization_id, current_user_is_individual, current_user_is_super_admin
        from core.models import User
        u_qs = User.objects.filter(id=request.user.id)
        logger.info(f"DEBUG: user_sql={u_qs.query}")
        u_from_db = u_qs.first()
        logger.info(f"DEBUG: user={request.user.email}, user_is_indiv={getattr(request.user, 'is_individual', 'N/A')}, db_is_indiv={getattr(u_from_db, 'is_individual', 'N/A')}, ctx_is_indiv={current_user_is_individual.get()}")

        candidate_assessments = CandidateAssessment.objects.filter(
            candidate=request.user,
            status__in=['assigned', 'in_progress']
        ).select_related('assessment').order_by('-assigned_date')

        logger.info(
            "Regular candidate assessments fetched successfully",
            extra={
                "candidate_id": request.user.id,
                "regular_assessment_count": candidate_assessments.count(),
            }
        )

        assessments_data = []

        for ca in candidate_assessments:

            assessment = ca.assessment

            is_currently_active = (
                ca.status in ['assigned', 'in_progress'] and
                (assessment.start_date is None or assessment.start_date <= now) and
                (assessment.end_date is None or assessment.end_date >= now)
            )

            assessments_data.append({
                'candidate_assessment_id': ca.id,
                'assessment_id': assessment.id,
                'title': assessment.title,
                'description': assessment.description,
                'duration_minutes': assessment.duration,
                'start_date': assessment.start_date,
                'end_date': assessment.end_date,
                'status': ca.status,
                'assigned_date': ca.assigned_date,
                'start_time': ca.start_time,
                'total_questions': assessment.questions.count(),
                'is_active': assessment.is_active,
                'is_currently_active': is_currently_active,
                'assessment_type': 'regular',
            })

        logger.info(
            "Regular assessment payload prepared successfully",
            extra={
                "candidate_id": request.user.id,
                "regular_payload_count": len(assessments_data),
            }
        )

        ai_candidate_assessments = (
            CandidateAIAssessment.objects.filter(
                candidate=request.user,
                status__in=['assigned', 'in_progress'],
                # Exclude assessments the candidate has already submitted — once
                # submitted we stamp end_time, even though status stays
                # in_progress until the async report is ready. This is what
                # stops a completed assessment from showing "Start" again.
                end_time__isnull=True,
            )
            .select_related('ai_assessment')
            .order_by('-assigned_date')
        )

        logger.info(
            "AI candidate assessments fetched successfully",
            extra={
                "candidate_id": request.user.id,
                "ai_assessment_count": ai_candidate_assessments.count(),
            }
        )

        ai_assessments_data = []

        for ai_ca in ai_candidate_assessments:

            ai_assessment = ai_ca.ai_assessment

            total_questions = len(ai_ca.generated_questions or [])

            if not total_questions:
                total_questions = ai_assessment.num_questions

            ai_is_currently_active = (
                ai_ca.status in ['assigned', 'in_progress'] and
                (ai_assessment.start_date is None or ai_assessment.start_date <= now) and
                (ai_assessment.end_date is None or ai_assessment.end_date >= now)
            )

            ai_assessments_data.append({
                'candidate_ai_assessment_id': ai_ca.id,
                'assessment_id': ai_assessment.id,
                'title': ai_assessment.title,
                'description': ai_assessment.description,
                'duration_minutes': getattr(ai_assessment, 'duration', None),
                'start_date': ai_assessment.start_date,
                'end_date': ai_assessment.end_date,
                'status': ai_ca.status,
                'assigned_date': ai_ca.assigned_date,
                'start_time': ai_ca.start_time,
                'total_questions': total_questions,
                'is_active': ai_assessment.is_active,
                'is_currently_active': ai_is_currently_active,
                'assessment_type': 'ai',
            })

        logger.info(
            "AI assessment payload prepared successfully",
            extra={
                "candidate_id": request.user.id,
                "ai_payload_count": len(ai_assessments_data),
            }
        )

        total_count = len(assessments_data) + len(ai_assessments_data)

        logger.info(
            "Candidate assigned assessments response prepared successfully",
            extra={
                "candidate_id": request.user.id,
                "total_count": total_count,
                "regular_count": len(assessments_data),
                "ai_count": len(ai_assessments_data),
            }
        )

        return Response({
            'count': total_count,
            'regular_count': len(assessments_data),
            'ai_count': len(ai_assessments_data),
            'assigned_assessments': assessments_data,
            'ai_assigned_assessments': ai_assessments_data,
        })


class CandidateCompletedAssessmentsView(APIView):
    """
    Get all completed assessments for the candidate.
    Includes score, percentage, and submission details.
    """
    permission_classes = [CandidatePermission]

    def get(self, request, *args, **kwargs):

        logger.info(
            "Candidate completed assessments API called",
            extra={"candidate_id": request.user.id}
        )

        candidate_assessments = CandidateAssessment.objects.filter(
            candidate=request.user,
            status='completed'
        ).select_related('assessment').order_by('-end_time')

        logger.info(
            "Completed regular assessments fetched successfully",
            extra={
                "candidate_id": request.user.id,
                "regular_completed_count": candidate_assessments.count(),
            }
        )

        completed_ai_assessments = (
            CandidateAIAssessment.objects.filter(
                candidate=request.user,
                status='completed'
            )
            .select_related('ai_assessment')
            .prefetch_related('aiinterviewresponse_set')
            .order_by('-end_time')
        )

        logger.info(
            "Completed AI assessments fetched successfully",
            extra={
                "candidate_id": request.user.id,
                "ai_completed_count": completed_ai_assessments.count(),
            }
        )

        assessments_data = []

        for ca in candidate_assessments:

            assessment = ca.assessment

            responses = CandidateResponse.objects.filter(
                candidate=request.user,
                assessment=assessment
            )

            logger.info(
                "Processing completed regular assessment",
                extra={
                    "candidate_assessment_id": ca.id,
                    "assessment_id": assessment.id,
                }
            )
            
            assessments_data.append({
                'candidate_assessment_id': ca.id,
                'assessment_id': assessment.id,
                'title': assessment.title,
                'description': assessment.description,
                'duration_minutes': assessment.duration,
                'score': ca.score,
                'total_marks': ca.total_marks,
                'percentage': round(ca.percentage, 2),
                'status': ca.status,
                'assigned_date': ca.assigned_date,
                'start_time': ca.start_time,
                'end_time': ca.end_time,
                'total_questions': assessment.questions.count(),
                'attempted_questions': responses.exclude(answer__isnull=True).exclude(answer='').count(),
                'correct_questions': responses.filter(is_correct=True).count(),
                'assessment_type': 'regular',
            })

        logger.info(
            "Regular completed assessments payload prepared successfully",
            extra={
                "candidate_id": request.user.id,
                "regular_payload_count": len(assessments_data),
            }
        )

        ai_completed_data = []

        for ai_ca in completed_ai_assessments:

            ai_assessment = ai_ca.ai_assessment

            responses = ai_ca.aiinterviewresponse_set.all()

            attempted_questions = (
                responses.exclude(answer_text__isnull=True)
                .exclude(answer_text='')
                .count()
            )

            total_questions = len(ai_ca.generated_questions or [])

            if not total_questions:
                total_questions = ai_assessment.num_questions

            logger.info(
                "Processing completed AI assessment",
                extra={
                    "candidate_ai_assessment_id": ai_ca.id,
                    "assessment_id": ai_assessment.id,
                }
            )

            ai_completed_data.append({
                'candidate_ai_assessment_id': ai_ca.id,
                'assessment_id': ai_assessment.id,
                'title': ai_assessment.title,
                'description': ai_assessment.description,
                'duration_minutes': getattr(ai_assessment, 'duration', None),
                'overall_score': ai_ca.overall_score,
                'technical_score': ai_ca.technical_score,
                'communication_score': ai_ca.communication_score,
                'problem_solving_score': ai_ca.problem_solving_score,
                'overall_feedback': ai_ca.overall_feedback,
                'technical_feedback': ai_ca.technical_feedback,
                'communication_feedback': ai_ca.communication_feedback,
                'problem_solving_feedback': ai_ca.problem_solving_feedback,
                'strengths_feedback': ai_ca.strengths_feedback,
                'improvement_feedback': ai_ca.improvement_feedback,
                'ai_feedback': ai_ca.ai_feedback,
                'status': ai_ca.status,
                'assigned_date': ai_ca.assigned_date,
                'start_time': ai_ca.start_time,
                'end_time': ai_ca.end_time,
                'total_questions': total_questions,
                'attempted_questions': attempted_questions,
                'assessment_type': 'ai',
            })

        logger.info(
            "AI completed assessments payload prepared successfully",
            extra={
                "candidate_id": request.user.id,
                "ai_payload_count": len(ai_completed_data),
            }
        )

        total_count = len(assessments_data) + len(ai_completed_data)

        logger.info(
            "Candidate completed assessments response prepared successfully",
            extra={
                "candidate_id": request.user.id,
                "total_count": total_count,
                "regular_count": len(assessments_data),
                "ai_count": len(ai_completed_data),
            }
        )

        return Response({
            'count': total_count,
            'regular_count': len(assessments_data),
            'ai_count': len(ai_completed_data),
            'completed_assessments': assessments_data,
            'ai_completed_assessments': ai_completed_data,
        })


class CandidateUpcomingAssessmentsView(APIView):
    """
    Get all upcoming assessments for the candidate.
    Includes assessments that are assigned but not yet started.
    """
    permission_classes = [CandidatePermission]

    def get(self, request, *args, **kwargs):

        logger.info(
            "Candidate upcoming assessments API called",
            extra={"candidate_id": request.user.id}
        )

        now = timezone.now()
        
        # Get assigned assessments that haven't started yet
        candidate_assessments = CandidateAssessment.objects.filter(
            candidate=request.user,
            status='assigned'
        ).select_related('assessment').order_by('assessment__start_date')

        logger.info(
            "Upcoming regular assessments fetched successfully",
            extra={
                "candidate_id": request.user.id,
                "regular_upcoming_count": candidate_assessments.count(),
            }
        )

        ai_candidate_assessments = (
            CandidateAIAssessment.objects.filter(
                candidate=request.user,
                status='assigned'
            )
            .select_related('ai_assessment')
            .order_by('ai_assessment__start_date')
        )

        logger.info(
            "Upcoming AI assessments fetched successfully",
            extra={
                "candidate_id": request.user.id,
                "ai_upcoming_count": ai_candidate_assessments.count(),
            }
        )

        assessments_data = []

        for ca in candidate_assessments:

            assessment = ca.assessment
            
            # Only include if assessment hasn't expired yet
            if assessment.end_date > now:

                logger.info(
                    "Processing upcoming regular assessment",
                    extra={
                        "candidate_assessment_id": ca.id,
                        "assessment_id": assessment.id,
                    }
                )

                time_until_start = assessment.start_date - now
                days_remaining = max(0, time_until_start.days)
                hours_remaining = max(0, time_until_start.seconds // 3600)
                
                assessments_data.append({
                    'candidate_assessment_id': ca.id,
                    'assessment_id': assessment.id,
                    'title': assessment.title,
                    'description': assessment.description,
                    'duration_minutes': assessment.duration,
                    'start_date': assessment.start_date,
                    'end_date': assessment.end_date,
                    'status': ca.status,
                    'assigned_date': ca.assigned_date,
                    'total_questions': assessment.questions.count(),
                    'is_active': assessment.is_active,
                    'time_remaining': {
                        'days': days_remaining,
                        'hours': hours_remaining,
                    },
                    'assessment_type': 'regular',
                })

        logger.info(
            "Upcoming regular assessment payload prepared successfully",
            extra={
                "candidate_id": request.user.id,
                "regular_payload_count": len(assessments_data),
            }
        )

        ai_upcoming_data = []

        for ai_ca in ai_candidate_assessments:

            ai_assessment = ai_ca.ai_assessment

            if ai_assessment.end_date > now:

                logger.info(
                    "Processing upcoming AI assessment",
                    extra={
                        "candidate_ai_assessment_id": ai_ca.id,
                        "assessment_id": ai_assessment.id,
                    }
                )

                time_until_start = ai_assessment.start_date - now
                days_remaining = max(0, time_until_start.days)
                hours_remaining = max(0, time_until_start.seconds // 3600)

                total_questions = len(ai_ca.generated_questions or [])

                if not total_questions:
                    total_questions = ai_assessment.num_questions

                ai_upcoming_data.append({
                    'candidate_ai_assessment_id': ai_ca.id,
                    'assessment_id': ai_assessment.id,
                    'title': ai_assessment.title,
                    'description': ai_assessment.description,
                    'duration_minutes': getattr(ai_assessment, 'duration', None),
                    'start_date': ai_assessment.start_date,
                    'end_date': ai_assessment.end_date,
                    'status': ai_ca.status,
                    'assigned_date': ai_ca.assigned_date,
                    'total_questions': total_questions,
                    'is_active': ai_assessment.is_active,
                    'time_remaining': {
                        'days': days_remaining,
                        'hours': hours_remaining,
                    },
                    'assessment_type': 'ai',
                })

        logger.info(
            "Upcoming AI assessment payload prepared successfully",
            extra={
                "candidate_id": request.user.id,
                "ai_payload_count": len(ai_upcoming_data),
            }
        )

        total_count = len(assessments_data) + len(ai_upcoming_data)

        logger.info(
            "Candidate upcoming assessments response prepared successfully",
            extra={
                "candidate_id": request.user.id,
                "total_count": total_count,
                "regular_count": len(assessments_data),
                "ai_count": len(ai_upcoming_data),
            }
        )

        return Response({
            'count': total_count,
            'regular_count': len(assessments_data),
            'ai_count': len(ai_upcoming_data),
            'upcoming_assessments': assessments_data,
            'ai_upcoming_assessments': ai_upcoming_data,
        })
